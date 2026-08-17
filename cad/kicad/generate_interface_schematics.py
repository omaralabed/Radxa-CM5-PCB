#!/usr/bin/env python3
"""Generate native KiCad interface-contract schematics for CM5-CARRIER and AUDIO-8X8."""

from __future__ import annotations

import json
from contextlib import contextmanager
from pathlib import Path
import re
import shutil
import uuid as _uuid


_uuid_counter = 0


def _deterministic_uuid4() -> _uuid.UUID:
    """Return stable KiCad UUIDs for reproducible generated schematics."""
    global _uuid_counter
    _uuid_counter += 1
    return _uuid.uuid5(
        _uuid.NAMESPACE_URL,
        f"radxa-cm5-procomm:interface-schematics:{_uuid_counter}",
    )


# kicad-sch-api uses uuid4 throughout its object factories. Install the stable
# source before importing it so repeated generation produces reviewable diffs.
_uuid.uuid4 = _deterministic_uuid4


@contextmanager
def isolated_uuid_namespace(tag: str):
    """Allocate UUIDs for new blocks without shifting established sheet IDs."""
    counter = 0
    previous_uuid4 = _uuid.uuid4

    def isolated_uuid4() -> _uuid.UUID:
        nonlocal counter
        counter += 1
        return _uuid.uuid5(
            _uuid.NAMESPACE_URL,
            f"radxa-cm5-procomm:interface-schematics:extension:{tag}:{counter}",
        )

    _uuid.uuid4 = isolated_uuid4
    try:
        yield
    finally:
        _uuid.uuid4 = previous_uuid4


def lock_component_pin_uuids(component) -> None:
    """Assign pin UUIDs now so KiCad's serializer cannot shift later sheets."""
    if component.pin_uuids:
        return
    for pin in component.pins:
        component.pin_uuids[str(pin.number)] = str(_uuid.uuid4())

from kicad_sch_api import create_schematic, get_symbol_cache
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
WORKSPACE = ROOT.parent.parent
CM5_PINOUT = WORKSPACE / "docs" / "radxa_cm5_v2210_pinout.xlsx"
CM5_LOCAL_LIBRARY = ROOT / "CM5-CARRIER" / "CM5Carrier.kicad_sym"
CM5_WURTH_LIBRARY = ROOT / "CM5-CARRIER" / "WurthRJ45.kicad_sym"
CM5_LOCAL_FOOTPRINTS = ROOT / "CM5-CARRIER" / "CM5Carrier.pretty"
CM5_LOCAL_3DMODELS = ROOT / "CM5-CARRIER" / "CM5Carrier.3dshapes"
COMPONENT_REFERENCES = WORKSPACE / "references" / "components"
WURTH_74991114412 = COMPONENT_REFERENCES / "wurth" / "74991114412"
LEGACY_CELLULAR_FOOTPRINTS = (
    WORKSPACE.parent / "ProComm enclosure and PCB boards" /
    "ADVANCED_SCHEMATIC_WORK" / "ProComm_RevF.pretty"
)
KICAD_TEMPLATE = Path(
    "/Applications/KiCad/KiCad.app/Contents/SharedSupport/template/"
    "EuroCard160mmX100mm/EuroCard160mmX100mm.kicad_pro"
)

MF_2X2 = "Connector_Molex:Molex_Micro-Fit_3.0_43045-0412_2x02_P3.00mm_Vertical"
MF_2X4 = "Connector_Molex:Molex_Micro-Fit_3.0_43045-0812_2x04_P3.00mm_Vertical"
JST_GH_4 = "Connector_JST:JST_GH_BM04B-GHS-TBT_1x04-1MP_P1.25mm_Vertical"
JST_GH_5 = "Connector_JST:JST_GH_BM05B-GHS-TBT_1x05-1MP_P1.25mm_Vertical"
JST_GH_6 = "Connector_JST:JST_GH_BM06B-GHS-TBT_1x06-1MP_P1.25mm_Vertical"
PICO_6 = "Connector_Molex:Molex_PicoBlade_53047-0610_1x06_P1.25mm_Vertical"
PICO_8 = "Connector_Molex:Molex_PicoBlade_53047-0810_1x08_P1.25mm_Vertical"
R_0603 = "Resistor_SMD:R_0603_1608Metric"
C_0402 = "Capacitor_SMD:C_0402_1005Metric"
C_0603 = "Capacitor_SMD:C_0603_1608Metric"
C_0805 = "Capacitor_SMD:C_0805_2012Metric"
C_1206 = "Capacitor_SMD:C_1206_3216Metric"
C_1210 = "Capacitor_SMD:C_1210_3225Metric"
TEST_POINT_2MM = "TestPoint:TestPoint_Pad_D2.0mm"
MILLIGRID_2X15 = (
    "Connector_Molex_Milligrid:"
    "Molex_8783230xx_2x15_P2.0mm_Header_Vertical_Polarized_MountingPegs"
)
CM5_DF40_FOOTPRINT = (
    "Connector_Hirose_DF40:"
    "Hirose_DF40C-100DS-0.4V_2x50_P0.4mm"
)

# These mappings are package-code or manufacturer-drawing backed. Keep this
# list exact: an unlisted MPN must remain blocked until its package is checked.
MPN_FOOTPRINTS = {
    "GRM188R71H473KA61D": C_0603,
    "GRM1885C1H682JA01D": C_0603,
    "GRM1885C1H151JA01D": C_0603,
    "GRM188R71E104KA01D": C_0603,
    "GRM188R71A105KA61D": C_0603,
    "GRM21BR71A475KA73L": C_0805,
    "GRM21BR71A226ME44L": C_0805,
    "GRM21BR71E106KA73L": C_0805,
    "GRM31CR71H475KA12L": C_1206,
    "GRM32ER71A476KE15L": C_1210,
    "GCM32ER70J476KE19L": C_1210,
    "16SVP330M": "Capacitor_SMD:CP_Elec_10x12.6",
    "6SVP330M": "Capacitor_SMD:CP_Elec_10x7.9",
    "6SVP470M": "Capacitor_SMD:CP_Elec_8x11.9",
    "BAT54WS-7-F": "Diode_SMD:D_SOD-323",
    "0453003.MR": "Fuse:Fuse_Littelfuse-NANO2-451_453",
    "0453002.MR": "Fuse:Fuse_Littelfuse-NANO2-451_453",
    "0453.250MR": "Fuse:Fuse_Littelfuse-NANO2-451_453",
    "XAL7070-472MEC": "Inductor_SMD:L_Coilcraft_XAL7070-XXX",
    "XGL4030-222MEC": "Inductor_SMD:L_Coilcraft_XxL4040",
    "SPM10065VC-3R3M-D": "CM5Carrier:TDK_SPM10065VC",
    "74439370047": "CM5Carrier:Wurth_74439370047",
    "KRL6432E-M-R006-F-T1": "CM5Carrier:Susumu_KRL6432E_6mR",
    "KRL11050-C-R004-F-T1": "CM5Carrier:Susumu_KRL11050_4mR",
    "NVMFS6B25NLT1G": "CM5Carrier:onsemi_DFN5_5x6_488AA_GSD",
    "FDWS86068-F085": "CM5Carrier:onsemi_DFNW8_5p2x6p3_507AU_GSD",
    "CSD18532Q5B": "CM5Carrier:TI_DNK0008A_GSD",
    "CSD17573Q5B": "CM5Carrier:TI_DNK0008A_GSD",
    "LM61460RJR": "CM5Carrier:TI_RJR0014A",
    "LM61440RJR": "CM5Carrier:TI_RJR0014A",
    "TPS22990DMLR": "CM5Carrier:TI_DML0010A",
    "RC0603FR-0740K2L": R_0603,
    "RC0603FR-077K5L": R_0603,
    "RC0603FR-07499RL": R_0603,
    "RC0603FR-0769K8L": R_0603,
    "RC0603FR-0710KL": R_0603,
    "RC0603JR-070RL": R_0603,
    "TNPW060324K0BEEA": R_0603,
    "TNPW060326K1BEEA": R_0603,
    "TNPW06034K42BEEA": R_0603,
    "TNPW06031K24BEEA": R_0603,
    "TNPW06031K00BEEA": R_0603,
    "TNPW06034K99BEEA": R_0603,
    "TNPW06036K19BEEA": R_0603,
    "TNPW060315K6BEEA": R_0603,
    "TNPW060320K0BEEA": R_0603,
    "TNPW060321K5BEEA": R_0603,
    "TNPW060329K4BEEA": R_0603,
    "TNPW060335K7BEEA": R_0603,
    "TNPW060343K2BEEA": R_0603,
    "TNPW0603100KBEEA": R_0603,
    "TNPW0603280KBEEA": R_0603,
    "RC0603FR-0727K4L": R_0603,
    "RC0603FR-0733K2L": R_0603,
    "RC0603FR-0773K2L": R_0603,
    "RC0603FR-0793K1L": R_0603,
    "RC0603FR-07255RL": R_0603,
    "GRM31CR71H105KA61L": C_1206,
    "GRM188R71H103KA01D": C_0603,
    "GRM188R71H222KA01D": C_0603,
    "GRM188R71H333KA01D": C_0603,
    "GRM188R71H472KA01D": C_0603,
    "GRM1885C1H221JA01D": C_0603,
    "GRM1885C1H561JA01D": C_0603,
    "GRM188R71E474KA12D": C_0603,
    "PCA9306DP,118": "Package_SO:TSSOP-8_3x3mm_P0.65mm",
    "PCA9517ADP,118": "Package_SO:TSSOP-8_3x3mm_P0.65mm",
    "TCA9535PWR": "Package_SO:TSSOP-24_4.4x7.8mm_P0.65mm",
    "EMC2305-1-AP-TR": "Package_DFN_QFN:QFN-16-1EP_4x4mm_P0.65mm_EP2.1x2.1mm",
    "TMP117AIDRVR": "Package_SON:WSON-6-1EP_2x2mm_P0.65mm_EP1x1.6mm",
    "2N7002K-7": "Package_TO_SOT_SMD:SOT-23",
    "2920L300/15DR": "Fuse:Fuse_2920_7451Metric",
    "1812L110/33DR": "Fuse:Fuse_1812_4532Metric",
    "VLS3012HBX-3R3M-N": "CM5Carrier:TDK_VLS3012HBX",
    "ABM8-25.000MHZ-10-D1G-T": "Crystal:Crystal_SMD_3225-4Pin_3.2x2.5mm",
    "C1005X7R1H104K050BB": C_0402,
    "GRM188R61E106MA73D": C_0603,
    "GRM155R6YA105KE11D": C_0402,
    "GRM1555C1H150JA01D": C_0402,
    "RC0603FR-071K43L": R_0603,
    "RC0603FR-07475RL": R_0603,
    "RC0603FR-07200RL": R_0603,
    "RC0603FR-07200KL": R_0603,
    "RC0603FR-076K04L": R_0603,
    "RC0603FR-07330RL": R_0603,
    "RC0603FR-072K2L": R_0603,
    "RC0603FR-074K7L": R_0603,
    "RC0603FR-075K1L": R_0603,
    "RC0603FR-07100KL": R_0603,
    "RC0603FR-071ML": R_0603,
    "RC0603FR-071KL": R_0603,
    "TPD4E05U06DQAR": "Package_SON:USON-10_2.5x1.0mm_P0.5mm",
    "SN65LVDS047PWR": "Package_SO:TSSOP-16_4.4x5mm_P0.65mm",
    "SN65LVDT2DR": "Package_SO:SOIC-8_3.9x4.9mm_P1.27mm",
    "SN74AVC4T245PWR": "Package_SO:TSSOP-16_4.4x5mm_P0.65mm",
    "SN74LVC1T45DCKR": "Package_TO_SOT_SMD:SOT-363_SC-70-6",
    "SN74LVC1G11DBVR": "Package_TO_SOT_SMD:SOT-23-6",
    "LP5907MFX-3.3/NOPB": "Package_TO_SOT_SMD:SOT-23-5",
    "LP5907MFX-1.8/NOPB": "Package_TO_SOT_SMD:SOT-23-5",
    "TPA6132A2RTER": (
        "Package_DFN_QFN:"
        "WQFN-16-1EP_3x3mm_P0.5mm_EP1.68x1.68mm_ThermalVias"
    ),
    "87832-6423": MILLIGRID_2X15,
    "STX-353K7A-6N-KTTR": "CM5Carrier:Kycon_STX-353K7A-6N-KTTR_PRELIMINARY",
    "GRM188R71A225KE15": C_0603,
    "RC0603FR-07330KL": R_0603,
    "0603L010YR": "Fuse:Fuse_0603_1608Metric",
    "1206L110/16WR": "Fuse:Fuse_1206_3216Metric",
    "TPD3F303DPVR": (
        "Package_SON:Texas_R-PUSON-N8_USON-8-1EP_1.6x2.1mm_P0.5mm_EP0.4x1.7mm"
    ),
    "AK5558VN": "Package_DFN_QFN:QFN-64-1EP_9x9mm_P0.5mm_EP6x6mm_ThermalVias",
    "AK4458VN": "Package_DFN_QFN:QFN-48-1EP_7x7mm_P0.5mm_EP5.15x5.15mm_ThermalVias",
    "THAT1206S08-U": "Package_SO:SOIC-8_3.9x4.9mm_P1.27mm",
    "THAT1646S08-U": "Package_SO:SOIC-8_3.9x4.9mm_P1.27mm",
    "OPA1652AIDR": "Package_SO:SOIC-8_3.9x4.9mm_P1.27mm",
    "SN65LVDS1DR": "Package_SO:SOIC-8_3.9x4.9mm_P1.27mm",
    "LT3045IMSE#TRPBF": (
        "Package_SO:MSOP-12-1EP_3x4.039mm_P0.65mm_EP1.651x2.845mm_ThermalVias"
    ),
    "TPS62913RPUT": "Package_DFN_QFN:Texas_RPU0010A_VQFN-HR-10_2x2mm_P0.5mm",
    "TRI 20-1223": "CM5Carrier:TRACO_TRI20_DUAL",
    "ULN2803ADWR": "Package_SO:SOIC-18W_7.5x11.6mm_P1.27mm",
    "TQ2-12V": "CM5Carrier:Panasonic_TQ2-12V_PRELIMINARY",
    "S1G-13-F": "Diode_SMD:D_SMA",
    "BZT52C12-7-F": "Diode_SMD:D_SOD-123",
    "BLM21PG221SN1D": "Inductor_SMD:L_0805_2012Metric",
    "UES1H100MDM1TD": "Capacitor_THT:CP_Radial_D5.0mm_P2.00mm",
    "UES0J221MHM": "Capacitor_THT:CP_Radial_D10.0mm_P5.00mm",
    "6SVP100M": "Capacitor_SMD:CP_Elec_6.3x5.8",
    "6SVP220MX": "Capacitor_SMD:CP_Elec_8x6.9",
    "GRM1885C1H680JA01D": C_0603,
    "GRM1555C1H221JA01D": C_0402,
    "GRM1555C1H680JA01D": C_0402,
    "GRM1555C1H9R1BA01D": C_0402,
    "GRM1555C1H4R7BA01D": C_0402,
    "SMF4L5.0AT1G": "Diode_SMD:D_SOD-123F",
    "RC0603FR-0710RL": R_0603,
    "RC0603FR-07100RL": R_0603,
    "RC0603FR-0713KL": R_0603,
    "RC0603FR-073K9L": R_0603,
    "RC0603FR-07150RL": R_0603,
    "RC0603FR-0720RL": R_0603,
    "RC0603FR-07220RL": R_0603,
    "RC0603FR-0720KL": R_0603,
    "TNPW060349K9BEEA": R_0603,
    "GRM1885C1H471JA01D": C_0603,
    "GRM1885C1H101JA01D": C_0603,
    "GRM1885C1H332JA01D": C_0603,
    "GRM1885C1H392JA01D": C_0603,
    "GRM188R71A106KA73D": C_0603,
    "GRM188R71E474KA12D": C_0603,
    "GRM21BR71E106KA73L": C_0805,
    "RC0603FR-074K99L": R_0603,
    "RC0603FR-0729K4L": R_0603,
    "1N4148W-7-F": "Diode_SMD:D_SOD-123",
    "TPS7A2033PDBVR": "Package_TO_SOT_SMD:SOT-23-5",
    "GRM32ER71E226KE15L": C_1210,
    "RC0603FR-07140KL": R_0603,
    "RC2512JK-070RL": "Resistor_SMD:R_2512_6332Metric",
    "BM04B-GHS-TBT(LF)(SN)": JST_GH_4,
    "BM05B-GHS-TBT(LF)(SN)": JST_GH_5,
    "BM06B-GHS-TBT(LF)(SN)": JST_GH_6,
    "43045-0812": MF_2X4,
    "VY1472M63Y5UQ6TV0": "Capacitor_THT:C_Disc_D16.0mm_W5.0mm_P10.00mm",
}

POWER_PASSIVE_PARTS = {
    "1uF 50V X7R": ("Murata", "GRM31CR71H105KA61L"),
    "100nF": ("Murata", "GRM188R71E104KA01D"),
    "100nF 25V": ("Murata", "GRM188R71E104KA01D"),
    "10nF": ("Murata", "GRM188R71H103KA01D"),
    "2.2nF": ("Murata", "GRM188R71H222KA01D"),
    "33nF": ("Murata", "GRM188R71H333KA01D"),
    "4.7nF": ("Murata", "GRM188R71H472KA01D"),
    "220pF": ("Murata", "GRM1885C1H221JA01D"),
    "560pF": ("Murata", "GRM1885C1H561JA01D"),
    "470nF": ("Murata", "GRM188R71E474KA12D"),
    "1uF X7R": ("Murata", "GRM188R71A105KA61D"),
    "4.7uF 10V": ("Murata", "GRM21BR71A475KA73L"),
    "4.7uF 50V": ("Murata", "GRM31CR71H475KA12L"),
    "10k": ("Yageo", "RC0603FR-0710KL"),
    "10k 1%": ("Yageo", "RC0603FR-0710KL"),
    "100k 1%": ("Yageo", "RC0603FR-07100KL"),
    "27.4k": ("Yageo", "RC0603FR-0727K4L"),
    "33.2k / 400kHz": ("Yageo", "RC0603FR-0733K2L"),
    "73.2k 1%": ("Yageo", "RC0603FR-0773K2L"),
    "93.1k": ("Yageo", "RC0603FR-0793K1L"),
    "255R 1% / 6A ILIM": ("Yageo", "RC0603FR-07255RL"),
    "1.24k 0.1%": ("Vishay", "TNPW06031K24BEEA"),
    "4.99k 0.1%": ("Vishay", "TNPW06034K99BEEA"),
    "6.19k 0.1%": ("Vishay", "TNPW06036K19BEEA"),
    "15.6k 0.1%": ("Vishay", "TNPW060315K6BEEA"),
    "20k 0.1%": ("Vishay", "TNPW060320K0BEEA"),
    "29.4k 0.1%": ("Vishay", "TNPW060329K4BEEA"),
    "35.7k 0.1%": ("Vishay", "TNPW060335K7BEEA"),
    "43.2k 0.1%": ("Vishay", "TNPW060343K2BEEA"),
    "100k 0.1%": ("Vishay", "TNPW0603100KBEEA"),
    "280k 0.1%": ("Vishay", "TNPW0603280KBEEA"),
}

THERMAL_PASSIVE_PARTS = {
    "100nF": ("Murata", "GRM188R71E104KA01D"),
    "1k": ("Yageo", "RC0603FR-071KL"),
    "1k ring limit": ("Yageo", "RC0603FR-071KL"),
    "2.2k": ("Yageo", "RC0603FR-072K2L"),
    "4.7k": ("Yageo", "RC0603FR-074K7L"),
    "4.7k 5%": ("Yageo", "RC0603FR-074K7L"),
    "5.1k": ("Yageo", "RC0603FR-075K1L"),
    "10k": ("Yageo", "RC0603FR-0710KL"),
    "100k": ("Yageo", "RC0603FR-07100KL"),
    "200k": ("Yageo", "RC0603FR-07200KL"),
}

NETWORK_PASSIVE_PARTS = {
    "100nF": ("TDK", "C1005X7R1H104K050BB"),
    "10nF": ("Murata", "GRM188R71H103KA01D"),
    "22uF 10V X7R": ("Murata", "GRM21BR71A226ME44L"),
    "10uF 25V X5R": ("Murata", "GRM188R61E106MA73D"),
    "1uF 35V X5R / ESR <1R": ("Murata", "GRM155R6YA105KE11D"),
    "220R ferrite": ("Murata", "BLM21PG221SN1D"),
    "15pF C0G": ("Murata", "GRM1555C1H150JA01D"),
    "4.7k": ("Yageo", "RC0603FR-074K7L"),
    "1.43k 1%": ("Yageo", "RC0603FR-071K43L"),
    "475R 1%": ("Yageo", "RC0603FR-07475RL"),
    "2.2k": ("Yageo", "RC0603FR-072K2L"),
    "0R": ("Yageo", "RC0603JR-070RL"),
    "200R 1%": ("Yageo", "RC0603FR-07200RL"),
    "6.04k 1%": ("Yageo", "RC0603FR-076K04L"),
    "10k": ("Yageo", "RC0603FR-0710KL"),
    "100k": ("Yageo", "RC0603FR-07100KL"),
    "330R": ("Yageo", "RC0603FR-07330RL"),
}

WWAN_PASSIVE_PARTS = {
    "220uF 6.3V polymer": ("Panasonic", "6SVP220MX"),
    "10uF 25V X5R": ("Murata", "GRM188R61E106MA73D"),
    "100nF": ("TDK", "C1005X7R1H104K050BB"),
    "6.8nF C0G": ("Murata", "GRM1885C1H682JA01D"),
    "220pF C0G": ("Murata", "GRM1555C1H221JA01D"),
    "68pF C0G": ("Murata", "GRM1555C1H680JA01D"),
    "15pF C0G": ("Murata", "GRM1555C1H150JA01D"),
    "9.1pF C0G": ("Murata", "GRM1555C1H9R1BA01D"),
    "4.7pF C0G": ("Murata", "GRM1555C1H4R7BA01D"),
    "100k": ("Yageo", "RC0603FR-07100KL"),
}

DISPLAY_PASSIVE_PARTS = {
    "2.2k": ("Yageo", "RC0603FR-072K2L"),
    "7.5k S-CONF": ("Yageo", "RC0603FR-077K5L"),
    "26.1k 0.1%": ("Vishay", "TNPW060326K1BEEA"),
    "4.99k 0.1%": ("Vishay", "TNPW06034K99BEEA"),
    "470nF": ("Murata", "GRM188R71E474KA12D"),
    "10uF 25V X5R": ("Murata", "GRM188R61E106MA73D"),
    "10uF 25V X7R": ("Murata", "GRM21BR71E106KA73L"),
    "22uF 10V X7R": ("Murata", "GRM21BR71A226ME44L"),
    "100nF": ("TDK", "C1005X7R1H104K050BB"),
}

AUDIO_PASSIVE_PARTS = {
    "0R": ("Yageo", "RC0603JR-070RL"),
    "0R star bond": ("Yageo", "RC0603JR-070RL"),
    "0R power star": ("Yageo", "RC2512JK-070RL"),
    "100nF": ("TDK", "C1005X7R1H104K050BB"),
    "1uF": ("Murata", "GRM188R71A105KA61D"),
    "2.2uF": ("Murata", "GRM188R71A225KE15"),
    "4.7uF": ("Murata", "GRM21BR71A475KA73L"),
    "10uF": ("Murata", "GRM188R61E106MA73D"),
    "2.2k": ("Yageo", "RC0603FR-072K2L"),
    "10k": ("Yageo", "RC0603FR-0710KL"),
    "100k": ("Yageo", "RC0603FR-07100KL"),
    "200k": ("Yageo", "RC0603FR-07200KL"),
    "330k": ("Yageo", "RC0603FR-07330KL"),
}

AUDIO8_PASSIVE_PARTS = {
    "0R star bond": ("Yageo", "RC0603JR-070RL"),
    "0R power star": ("Yageo", "RC2512JK-070RL"),
    "20R": ("Yageo", "RC0603FR-0720RL"),
    "10R": ("Yageo", "RC0603FR-0710RL"),
    "100R": ("Yageo", "RC0603FR-07100RL"),
    "150R": ("Yageo", "RC0603FR-07150RL"),
    "220R": ("Yageo", "RC0603FR-07220RL"),
    "1k 0.1%": ("Vishay", "TNPW06031K00BEEA"),
    "3.9k 0.1%": ("Yageo", "RC0603FR-073K9L"),
    "4.7k 0.1%": ("Yageo", "RC0603FR-074K7L"),
    "10k 0.1%": ("Yageo", "RC0603FR-0710KL"),
    "13k 0.1%": ("Yageo", "RC0603FR-0713KL"),
    "20k 0.1%": ("Yageo", "RC0603FR-0720KL"),
    "21.5k 0.1%": ("Vishay", "TNPW060321K5BEEA"),
    "49.9k 0.1%": ("Vishay", "TNPW060349K9BEEA"),
    "29.4k 0.1%": ("Vishay", "TNPW060329K4BEEA"),
    "4.99k 0.1%": ("Yageo", "RC0603FR-074K99L"),
    "7.5k 0.1%": ("Yageo", "RC0603FR-077K5L"),
    "140k 1%": ("Yageo", "RC0603FR-07140KL"),
    "100k 1%": ("Yageo", "RC0603FR-07100KL"),
    "499R 1% / 300mA ILIM": ("Yageo", "RC0603FR-07499RL"),
    "1M": ("Yageo", "RC0603FR-071ML"),
    "100nF": ("TDK", "C1005X7R1H104K050BB"),
    "470nF": ("Murata", "GRM188R71E474KA12D"),
    "100pF C0G": ("Murata", "GRM1885C1H101JA01D"),
    "470pF C0G": ("Murata", "GRM1885C1H471JA01D"),
    "3.3nF C0G": ("Murata", "GRM1885C1H332JA01D"),
    "3.9nF C0G": ("Murata", "GRM1885C1H392JA01D"),
    "1uF": ("Murata", "GRM188R71A105KA61D"),
    "4.7uF": ("Murata", "GRM21BR71A475KA73L"),
    "4.7uF 50V": ("Murata", "GRM31CR71H475KA12L"),
    "10uF": ("Murata", "GRM188R71A106KA73D"),
    "22uF 25V": ("Murata", "GRM32ER71E226KE15L"),
    "10uF bipolar": ("Nichicon", "UES1H100MDM1TD"),
    "22uF 10V": ("Murata", "GRM21BR71A226ME44L"),
    "100uF 6.3V polymer": ("Panasonic", "6SVP100M"),
    "220uF 6.3V polymer": ("Panasonic", "6SVP220MX"),
    "220uF bipolar": ("Nichicon", "UES0J221MHM"),
    "1N4148W": ("Diodes Incorporated", "1N4148W-7-F"),
    "12V zener": ("Diodes Incorporated", "BZT52C12-7-F"),
    "220R ferrite": ("Murata", "BLM21PG221SN1D"),
    "S1G": ("Diodes Incorporated", "S1G-13-F"),
    "2.2uH": ("Coilcraft", "XGL4030-222MEC"),
    "3A fuse": ("Littelfuse", "0453003.MR"),
}


# This is the controlled allocation from the A0 pin audit.  Physical signal
# names for all 300 connector contacts come directly from the official V2.21
# workbook; only these 76 contacts receive product nets in Rev A.
CM5_ALLOCATIONS = (
    ("U13-A", 77, "Input", "SYS_4V0"),
    ("U13-A", 79, "Input", "SYS_4V0"),
    ("U13-A", 81, "Input", "SYS_4V0"),
    ("U13-A", 83, "Input", "SYS_4V0"),
    ("U13-A", 85, "Input", "SYS_4V0"),
    ("U13-A", 87, "Input", "SYS_4V0"),
    ("U13-A", 78, "Input", "LOGIC_3V3"),
    ("U13-B", 106, "Input", "IO_5V0"),
    ("U13-A", 92, "Input", "CM5_RESET_N"),
    ("U13-A", 93, "Input", "CM5_BOOT"),
    ("U13-A", 99, "Input", "CM5_PWRON_N"),
    ("J1", 26, "Input", "CM5_RECOVERY_KEY"),
    ("U13-A", 3, "Bidirectional", "WAN1_MDI3_P"),
    ("U13-A", 5, "Bidirectional", "WAN1_MDI3_N"),
    ("U13-A", 4, "Bidirectional", "WAN1_MDI1_P"),
    ("U13-A", 6, "Bidirectional", "WAN1_MDI1_N"),
    ("U13-A", 11, "Bidirectional", "WAN1_MDI2_P"),
    ("U13-A", 9, "Bidirectional", "WAN1_MDI2_N"),
    ("U13-A", 12, "Bidirectional", "WAN1_MDI0_P"),
    ("U13-A", 10, "Bidirectional", "WAN1_MDI0_N"),
    ("U13-A", 15, "Output", "WAN1_LED2"),
    ("U13-A", 17, "Output", "WAN1_LED1"),
    ("U13-A", 19, "Output", "WAN1_LED0"),
    ("U13-A", 24, "Input", "PCIE_UP_WAKE_N"),
    ("U13-B", 102, "Input", "PCIE_UP_CLKREQ_N"),
    ("U13-B", 109, "Output", "PCIE_UP_PERST_CMD_N"),
    ("U13-B", 110, "Output", "PCIE_UP_REFCLK_P"),
    ("U13-B", 112, "Output", "PCIE_UP_REFCLK_N"),
    ("U13-B", 116, "Input", "PCIE_UP_RX_P_CM5"),
    ("U13-B", 118, "Input", "PCIE_UP_RX_N_CM5"),
    ("U13-B", 122, "Output", "PCIE_UP_TX_P_CM5"),
    ("U13-B", 124, "Output", "PCIE_UP_TX_N_CM5"),
    ("J1", 51, "Output", "WWAN_USB3_TX_P"),
    ("J1", 53, "Output", "WWAN_USB3_TX_N"),
    ("J1", 57, "Input", "WWAN_USB3_RX_P"),
    ("J1", 59, "Input", "WWAN_USB3_RX_N"),
    ("J1", 63, "Bidirectional", "WWAN_USB2_DP"),
    ("J1", 65, "Bidirectional", "WWAN_USB2_DM"),
    ("J1", 45, "Bidirectional", "TOUCH_USB_DP"),
    ("J1", 47, "Bidirectional", "TOUCH_USB_DM"),
    ("U13-B", 101, "Input", "REC_USB_ID"),
    ("U13-B", 103, "Bidirectional", "REC_USB_DM"),
    ("U13-B", 105, "Bidirectional", "REC_USB_DP"),
    ("J1", 97, "Input", "REC_USB_VBUS_DET"),
    ("U13-B", 170, "Output", "HDMI_D2_P"),
    ("U13-B", 172, "Output", "HDMI_D2_N"),
    ("U13-B", 176, "Output", "HDMI_D1_P"),
    ("U13-B", 178, "Output", "HDMI_D1_N"),
    ("U13-B", 182, "Output", "HDMI_D0_P"),
    ("U13-B", 184, "Output", "HDMI_D0_N"),
    ("U13-B", 188, "Output", "HDMI_CLK_P"),
    ("U13-B", 190, "Output", "HDMI_CLK_N"),
    ("U13-B", 151, "Bidirectional", "HDMI_CEC"),
    ("U13-B", 153, "Input", "HDMI_HPD"),
    ("U13-B", 199, "Bidirectional", "HDMI_DDC_SDA"),
    ("U13-B", 200, "Bidirectional", "HDMI_DDC_SCL"),
    ("U13-B", 145, "Bidirectional", "HDMI_HEAC_P"),
    ("U13-B", 147, "Bidirectional", "HDMI_HEAC_N"),
    ("U13-A", 34, "Output", "AUD_DAC_SDIN"),
    ("U13-A", 46, "Output", "AUD_BCLK"),
    ("U13-A", 48, "Output", "AUD_FSYNC"),
    ("U13-A", 50, "Output", "AUD_MCLK"),
    ("U13-A", 54, "Input", "AUD_ADC_SDOUT"),
    ("U13-A", 80, "Bidirectional", "SYS_I2C7_SCL"),
    ("U13-A", 82, "Bidirectional", "SYS_I2C7_SDA"),
    ("U13-A", 20, "Input", "AUD_IRQ_N"),
    ("U13-A", 100, "Output", "HS_MCLK"),
    ("U13-A", 28, "Output", "HS_BCLK"),
    ("U13-A", 30, "Output", "HS_LRCK"),
    ("U13-A", 31, "Output", "HS_SDOUT_TO_CODEC"),
    ("J1", 40, "Input", "HS_SDIN_FROM_CODEC"),
    ("J1", 4, "Bidirectional", "HS_I2C_SDA"),
    ("J1", 6, "Bidirectional", "HS_I2C_SCL"),
    ("J1", 36, "Input", "HS_JACK_DET_N"),
    ("U13-A", 51, "Input", "DBG_UART_RX"),
    ("U13-A", 55, "Output", "DBG_UART_TX"),
)

# These contacts remain owned by the allocation ledger but intentionally have
# no copper in Rev A: the selected two-LED WAN jack does not use native LED2,
# and the selected display does not use the HEAC/ARC return sideband.
CM5_ASSIGNED_NC = {("U13-A", 15), ("U13-B", 147)}

LAN7430_PINS = (
    (1, "AVDDH_1"), (2, "TXRXP_A"), (3, "TXRXM_A"), (4, "AVDDL_1"),
    (5, "TXRXP_B"), (6, "TXRXM_B"), (7, "TXRXP_C"), (8, "TXRXM_C"),
    (9, "AVDDL_2"), (10, "TXRXP_D"), (11, "TXRXM_D"), (12, "AVDDH_2"),
    (13, "VDD12CORE"), (14, "VP"), (15, "GD_1"), (16, "PCIE_RX_P"),
    (17, "PCIE_RX_M"), (18, "GD_2"), (19, "PCIE_TX_P"), (20, "VPTX"),
    (21, "PCIE_TX_M"), (22, "GD_3"), (23, "VPH"), (24, "RESREF"),
    (25, "PCIE_CLK_P"), (26, "PCIE_CLK_M"), (27, "VDD25_REG_OUT"),
    (28, "VDD_REG_IN"), (29, "RESET_N"), (30, "TEST"), (31, "VDD12CORE"),
    (32, "VDD12_SW_OUT"), (33, "VDD_SW_IN"), (34, "VDD12_SW_FB"),
    (35, "VAUX_DET/GPIO3/LED3/TCK"),
    (36, "EECLK/GPIO2/LED2/TMS/ADV_PM_DISABLE"),
    (37, "EEDIO/GPIO1/LED1/TDO"), (38, "EECS/GPIO0/LED0/TDI"),
    (39, "VDDVARIO"), (40, "VDD12CORE"), (41, "VDD_OTP"),
    (42, "CLKREQ#"), (43, "WAKE#"), (44, "PERST#"), (45, "AVDD12"),
    (46, "XO"), (47, "XI"), (48, "ISET"), (49, "VSS_EP"),
)

PI7_BALL_ROWS = {
    "A": ("CLKREQ_L[5]", "CLKREQ_L[4]", "PERP[7]", "PETP[7]", "PERP[6]", "PETP[6]", "REFCLKN[1]", "REXT[1]", "PETP[5]", "PERP[5]", "PETP[4]", "PERP[4]", "SHCL_I2C", "LNKSTS[4]"),
    "B": ("LNKSTS[2]", "CLKREQ_L[6]", "PERN[7]", "PETN[7]", "PERN[6]", "PETN[6]", "REFCLKP[1]", "REXT_GND[1]", "PETN[5]", "PERN[5]", "PETN[4]", "PERN[4]", "SHDA_I2C", "LNKSTS[3]/LNKSTS_DIS"),
    "C": ("NC", "LNKSTS[5]", "NC", "NC", "AGND", "AGND", "NC", "AVDDH", "AGND", "AGND", "AGND", "PERST_L", "CLKREQ_L[3]", "SHPCINT_L"),
    "D": ("GPIO[1]", "GPIO[0]", "NC", "VDDR", "VDDC", "VDDC", "AVDD", "AVDD", "VDDC", "VDDC", "VDDR", "CLKREQ_L[2]", "REFCLKOP[1]", "REFCLKON[1]"),
    "E": ("GPIO[4]", "GPIO[3]", "GPIO[2]", "VDDC", "DGND", "DGND", "DGND", "DGND", "DGND", "DGND", "VDDC", "IREF", "REFCLKOP[2]", "REFCLKON[2]"),
    "F": ("UPS_PORTSEL[1]/DBO[4]", "GPIO[5]", "UPS_PORTSEL[0]/DBO[3]", "NC", "DGND", "DGND", "DGND", "DGND", "DGND", "DGND", "CLKREQ_L[1]", "SDA_I2C", "REFCLKOP[3]", "REFCLKON[3]"),
    "G": ("NC", "UPS_PORTSEL[3]", "AVDDH", "VDDC", "DGND", "DGND", "DGND", "DGND", "DGND", "DGND", "VDDC", "CVDDR", "REFCLKIP", "REFCLKIN"),
    "H": ("NC", "UPS_PORTSEL[2]", "AGND", "VDDC", "DGND", "DGND", "DGND", "DGND", "DGND", "DGND", "VDDC", "CVDDR", "REFCLKOP[4]", "REFCLKON[4]"),
    "J": ("GPIO[7]", "RXPOLINV_DIS", "DBO[1]", "GPIO[6]", "DGND", "DGND", "DGND", "DGND", "DGND", "DGND", "SCL_I2C", "SWG_LVL", "REFCLKOP[5]", "REFCLKON[5]"),
    "K": ("NC", "NC", "DBO[2]", "VDDC", "DGND", "DGND", "DGND", "DGND", "DGND", "DGND", "VDDC", "EEDO", "REFCLKOP[6]", "REFCLKON[6]"),
    "L": ("SMBUS_EN_L", "CLKBUF_PD", "NC", "VDDR", "VDDC", "VDDC", "AVDD", "AVDD", "VDDC", "VDDC", "VDDR", "EECK", "REFCLKOP[7]", "REFCLKON[7]"),
    "M": ("NC", "DWNRST_L[1]", "DWNRST_L[4]", "DWNRST_L[5]", "AGND", "AGND", "AVDDH", "NC", "AGND", "AGND", "AGND", "EECS_L", "LNKSTS[1]/DBG_DIS", "EEDI"),
    "N": ("PL_512B/DBO[0]", "DWNRST_L[2]", "PERN[0]", "PETN[0]", "PERN[1]", "PETN[1]", "REXT_GND[0]", "REFCLKP[0]", "PETN[2]", "PERN[2]", "PETN[3]", "PERN[3]", "VC1_EN", "NC"),
    "P": ("NC", "DWNRST_L[3]", "PERP[0]", "PETP[0]", "PERP[1]", "PETP[1]", "REXT[0]", "REFCLKN[0]", "PETP[2]", "PERP[2]", "PETP[3]", "PERP[3]", "PWR_SAV_DIS", "TEST/LNKSTS[0]"),
}


def allocation_map() -> dict[tuple[str, int], tuple[str, str]]:
    return {(connector, pin): (direction, net) for connector, pin, direction, net in CM5_ALLOCATIONS}


def load_cm5_pinout() -> dict[str, list[tuple[int, str]]]:
    workbook = load_workbook(CM5_PINOUT, data_only=True, read_only=True)
    signal_columns = {"U13-A": 5, "U13-B": 3, "J1": 2}
    pinout: dict[str, list[tuple[int, str]]] = {}
    for connector, signal_column in signal_columns.items():
        sheet = workbook[connector]
        rows = []
        for row in sheet.iter_rows(min_row=3, max_row=102, values_only=True):
            pin = int(row[0])
            signal = str(row[signal_column - 1] or "NC")
            rows.append((pin, signal))
        if len(rows) != 100:
            raise RuntimeError(f"{connector} source has {len(rows)} pins, expected 100")
        pinout[connector] = rows
    return pinout


def _escape_symbol_text(value: str) -> str:
    return value.replace("\\", "\\\\").replace('"', '\\"')


def write_cm5_local_library() -> dict[str, list[tuple[int, str]]]:
    """Build exact 100-contact CM5 mating symbols from Radxa's V2.21 pinout."""
    pinout = load_cm5_pinout()
    lines = ["(kicad_symbol_lib (version 20231120) (generator procomm_cm5_generator)"]
    for connector, pins in pinout.items():
        symbol_name = f"Radxa_CM5_{connector.replace('-', '_')}"
        start_pin = pins[0][0]
        lines.extend(
            [
                f'  (symbol "{symbol_name}" (pin_names (offset 0.9)) (in_bom yes) (on_board yes)',
                '    (property "Reference" "J" (at -25.4 67.31 0) (effects (font (size 1.27 1.27)) (justify left)))',
                f'    (property "Value" "RADXA CM5 {connector} MATE" (at -25.4 64.77 0) (effects (font (size 1.27 1.27)) (justify left)))',
                f'    (property "Footprint" "{CM5_DF40_FOOTPRINT}" (at 0 0 0) (effects (font (size 1.27 1.27)) hide))',
                '    (property "Datasheet" "https://dl.radxa.com/cm5/v2210/radxa_cm5_v2210_pinout.xlsx" (at 0 0 0) (effects (font (size 1.27 1.27)) hide))',
                f'    (property "Description" "Radxa CM5 {connector}, all 100 physical contacts" (at 0 0 0) (effects (font (size 1.27 1.27)) hide))',
                f'    (symbol "{symbol_name}_0_1"',
                '      (rectangle (start -25.4 63.5) (end 25.4 -63.5) (stroke (width 0.254) (type default)) (fill (type background)))',
                "    )",
                f'    (symbol "{symbol_name}_1_1"',
            ]
        )
        for pin, signal in pins:
            row = (pin - start_pin) // 2
            y = 62.23 - row * 2.54
            is_left = pin % 2 == 1
            x = -27.94 if is_left else 27.94
            orientation = 0 if is_left else 180
            # These are physical carrier-side connector contacts.  Keep them
            # passive for ERC; direction at the CM5 is controlled separately
            # in CM5_ALLOCATIONS and the allocation workbook.
            pin_type = "passive"
            escaped = _escape_symbol_text(signal)
            lines.append(
                f'      (pin {pin_type} line (at {x:.2f} {y:.2f} {orientation}) (length 2.54) '
                f'(name "{escaped}" (effects (font (size 0.75 0.75)))) '
                f'(number "{pin}" (effects (font (size 0.75 0.75)))))'
            )
        lines.extend(["    )", "  )"])

    lines.extend(
        [
            '  (symbol "LAN7430" (pin_names (offset 0.9)) (in_bom yes) (on_board yes)',
            '    (property "Reference" "U" (at -22.86 34.29 0) (effects (font (size 1.27 1.27)) (justify left)))',
            '    (property "Value" "LAN7430" (at -22.86 31.75 0) (effects (font (size 1.27 1.27)) (justify left)))',
            '    (property "Footprint" "Package_DFN_QFN:QFN-48-1EP_7x7mm_P0.5mm_EP5.3x5.3mm_ThermalVias" (at 0 0 0) (effects (font (size 1.27 1.27)) hide))',
            '    (property "Datasheet" "https://ww1.microchip.com/downloads/aemDocuments/documents/UNG/ProductDocuments/DataSheets/LAN7430-LAN7431-Data-Sheet-DS00002631.pdf" (at 0 0 0) (effects (font (size 1.27 1.27)) hide))',
            '    (property "Description" "48-pin VQFN PCIe to Gigabit Ethernet controller with integrated PHY" (at 0 0 0) (effects (font (size 1.27 1.27)) hide))',
            '    (symbol "LAN7430_0_1"',
            '      (rectangle (start -22.86 30.48) (end 22.86 -30.48) (stroke (width 0.254) (type default)) (fill (type background)))',
            '    )',
            '    (symbol "LAN7430_1_1"',
        ]
    )
    for index, (pin, signal) in enumerate(LAN7430_PINS):
        row = index // 2
        y = 29.21 - row * 2.54
        is_left = index % 2 == 0
        x = -25.4 if is_left else 25.4
        orientation = 0 if is_left else 180
        pin_type = "passive"
        lines.append(
            f'      (pin {pin_type} line (at {x:.2f} {y:.2f} {orientation}) (length 2.54) '
            f'(name "{_escape_symbol_text(signal)}" (effects (font (size 0.72 0.72)))) '
            f'(number "{pin}" (effects (font (size 0.72 0.72)))))'
        )
    lines.extend(["    )", "  )"])

    lines.extend(
        [
            '  (symbol "Crystal_GND24_3225" (pin_names (offset 0.9)) (in_bom yes) (on_board yes)',
            '    (property "Reference" "Y" (at -5.08 7.62 0) (effects (font (size 1.27 1.27))))',
            '    (property "Value" "Crystal_GND24_3225" (at 0 -7.62 0) (effects (font (size 1.27 1.27))))',
            '    (property "Footprint" "Crystal:Crystal_SMD_3225-4Pin_3.2x2.5mm" (at 0 0 0) (effects (font (size 1.27 1.27)) hide))',
            '    (property "Datasheet" "https://abracon.com/Resonators/abm8.pdf" (at 0 0 0) (effects (font (size 1.27 1.27)) hide))',
            '    (property "Description" "Four-pad 3225 crystal with grounded case pads 2 and 4" (at 0 0 0) (effects (font (size 1.27 1.27)) hide))',
            '    (symbol "Crystal_GND24_3225_0_1"',
            '      (rectangle (start -2.54 1.27) (end 2.54 -1.27) (stroke (width 0.254) (type default)) (fill (type background)))',
            '    )',
            '    (symbol "Crystal_GND24_3225_1_1"',
            '      (pin passive line (at -5.08 0 0) (length 2.54) (name "XI" (effects (font (size 0.72 0.72)))) (number "1" (effects (font (size 0.72 0.72)))))',
            '      (pin power_in line (at 0 -5.08 90) (length 3.81) (name "GND" (effects (font (size 0.72 0.72)))) (number "2" (effects (font (size 0.72 0.72)))))',
            '      (pin passive line (at 5.08 0 180) (length 2.54) (name "XO" (effects (font (size 0.72 0.72)))) (number "3" (effects (font (size 0.72 0.72)))))',
            '      (pin power_in line (at 0 5.08 270) (length 3.81) (name "GND" (effects (font (size 0.72 0.72)))) (number "4" (effects (font (size 0.72 0.72)))))',
            '    )',
            '  )',
        ]
    )

    power_names = {"VDDC", "VDDR", "CVDDR", "AVDD", "AVDDH", "AGND", "DGND"}
    pi7_units: dict[int, list[tuple[str, str]]] = {1: [], 2: [], 3: []}
    for row_name, names in PI7_BALL_ROWS.items():
        for column, signal in enumerate(names, start=1):
            ball = f"{row_name}{column}"
            if signal in power_names:
                unit = 3
            elif signal.startswith(("PER", "PET", "REFCLK", "CLKREQ", "DWNRST")) or signal == "PERST_L":
                unit = 1
            else:
                unit = 2
            pi7_units[unit].append((ball, signal))
    lines.extend(
        [
            '  (symbol "PI7C9X2G608GP" (pin_names (offset 0.9)) (in_bom yes) (on_board yes)',
            '    (property "Reference" "U" (at -25.4 2.54 0) (effects (font (size 1.27 1.27)) (justify left)))',
            '    (property "Value" "PI7C9X2G608GP" (at -25.4 0 0) (effects (font (size 1.27 1.27)) (justify left)))',
            '    (property "Footprint" "Package_BGA:BGA-196_15x15mm_Layout14x14_P1.0mm" (at 0 0 0) (effects (font (size 1.27 1.27)) hide))',
            '    (property "Datasheet" "https://www.diodes.com/datasheet/download/PI7C9X2G608GP.pdf" (at 0 0 0) (effects (font (size 1.27 1.27)) hide))',
            '    (property "Description" "196-ball LBGA, six-port/eight-lane PCIe Gen2 packet switch" (at 0 0 0) (effects (font (size 1.27 1.27)) hide))',
            '    (property "Manufacturer" "Diodes Incorporated" (at 0 0 0) (effects (font (size 1.27 1.27)) hide))',
            '    (property "MPN" "PI7C9X2G608GPCNJEX" (at 0 0 0) (effects (font (size 1.27 1.27)) hide))',
        ]
    )
    for unit, pins in pi7_units.items():
        rows = (len(pins) + 1) // 2
        half_height = max(12.7, rows * 1.27 + 2.54)
        lines.extend(
            [
                f'    (symbol "PI7C9X2G608GP_{unit}_1"',
                f'      (rectangle (start -25.4 {half_height:.2f}) (end 25.4 {-half_height:.2f}) (stroke (width 0.254) (type default)) (fill (type background)))',
            ]
        )
        for index, (ball, signal) in enumerate(pins):
            row = index // 2
            y = half_height - 2.54 - row * 2.54
            is_left = index % 2 == 0
            x = -27.94 if is_left else 27.94
            orientation = 0 if is_left else 180
            pin_type = "passive"
            lines.append(
                f'      (pin {pin_type} line (at {x:.2f} {y:.2f} {orientation}) (length 2.54) '
                f'(name "{_escape_symbol_text(signal)}" (effects (font (size 0.68 0.68)))) '
                f'(number "{ball}" (effects (font (size 0.68 0.68)))))'
            )
        lines.append("    )")
    lines.append("  )")

    local_symbols = (
        (
            "LM5146RGY",
            "U",
            "LM5146RGYR",
            "Package_DFN_QFN:Texas_RGY_R-PVQFN-N20_EP2.05x3.05mm_ThermalVias",
            "https://www.ti.com/lit/ds/symlink/lm5146.pdf",
            "100 V synchronous buck controller in 20-pin RGY VQFN",
            (
                (1, "EN/UVLO"), (2, "RT"), (3, "SS/TRK"), (4, "COMP"),
                (5, "FB"), (6, "AGND"), (7, "SYNCOUT"), (8, "SYNCIN"),
                (9, "NC"), (10, "PGOOD"), (11, "ILIM"), (12, "PGND"),
                (13, "LO"), (14, "VCC"), (15, "EP"), (16, "NC"),
                (17, "BST"), (18, "HO"), (19, "SW"), (20, "VIN"),
            ),
        ),
        (
            "LM61460RJR",
            "U",
            "LM61460RJR",
            "",
            "https://www.ti.com/lit/ds/symlink/lm61460.pdf",
            "36 V 6 A synchronous buck converter in 14-pin RJR VQFN-HR",
            (
                (1, "BIAS"), (2, "VCC"), (3, "AGND"), (4, "FB"),
                (5, "PGOOD"), (6, "RT"), (7, "EN/SYNC"), (8, "VIN1"),
                (9, "PGND1"), (10, "SW"), (11, "PGND2"), (12, "VIN2"),
                (13, "RBOOT"), (14, "CBOOT"),
            ),
        ),
        (
            "LM61440RJR",
            "U",
            "LM61440RJR",
            "",
            "https://www.ti.com/lit/ds/symlink/lm61440.pdf",
            "36 V 4 A synchronous buck converter in 14-pin RJR VQFN-HR",
            (
                (1, "BIAS"), (2, "VCC"), (3, "AGND"), (4, "FB"),
                (5, "PGOOD"), (6, "RT"), (7, "EN/SYNC"), (8, "VIN1"),
                (9, "PGND1"), (10, "SW"), (11, "PGND2"), (12, "VIN2"),
                (13, "RBOOT"), (14, "CBOOT"),
            ),
        ),
        (
            "LM5176PWP_A1",
            "U",
            "LM5176PWP",
            "Package_SO:HTSSOP-28-1EP_4.4x9.7mm_P0.65mm_EP3.4x9.5mm_ThermalVias",
            "https://www.ti.com/lit/ds/symlink/lm5176.pdf",
            "55 V four-switch buck-boost controller in 28-pin PWP HTSSOP",
            (
                (1, "EN/UVLO"), (2, "VIN"), (3, "VISNS"), (4, "MODE"),
                (5, "DITH"), (6, "RT/SYNC"), (7, "SLOPE"), (8, "SS"),
                (9, "COMP"), (10, "AGND"), (11, "FB"), (12, "VOSNS"),
                (13, "ISNS-"), (14, "ISNS+"), (15, "CSG"), (16, "CS"),
                (17, "PGOOD"), (18, "SW2"), (19, "HDRV2"), (20, "BOOT2"),
                (21, "LDRV2"), (22, "PGND"), (23, "VCC"), (24, "BIAS"),
                (25, "LDRV1"), (26, "BOOT1"), (27, "HDRV1"), (28, "SW1"),
                (29, "EP"),
            ),
        ),
        (
            "TPS62913RPU",
            "U",
            "TPS62913RPU",
            "Package_DFN_QFN:Texas_RPU0010A_VQFN-HR-10_2x2mm_P0.5mm",
            "https://www.ti.com/lit/ds/symlink/tps62913.pdf",
            "17 V 3 A low-noise buck converter in exact 10-pin RPU QFN pinout",
            (
                (1, "EN/SYNC"), (2, "SW"), (3, "VO"), (4, "PGND"),
                (5, "PG"), (6, "VIN"), (7, "PSNS"), (8, "NR/SS"),
                (9, "FB"), (10, "S-CONF"),
            ),
        ),
        (
            "TPS259827LNRGE",
            "U",
            "TPS259827LNRGER",
            "Package_DFN_QFN:Texas_RGE0024H_VQFN-24-1EP_4x4mm_P0.5mm_EP2.7x2.7mm_ThermalVias",
            "https://www.ti.com/lit/ds/symlink/tps25982.pdf",
            "24 V 15 A eFuse, no fixed OVLO, active current-limiter variant",
            (
                (1, "IN1"), (2, "IN2"), (3, "IN3"), (4, "GND1"),
                (5, "GND2"), (6, "EN/UVLO"), (7, "ITIMER"), (8, "ILIM"),
                (9, "IMON"), (10, "RETRY_DLY"), (11, "NRETRY"), (12, "LDSTRT"),
                (13, "PG"), (14, "GND3"), (15, "dVdt"), (16, "IN4"),
                (17, "OUT1"), (18, "OUT2"), (19, "OUT3"), (20, "OUT4"),
                (21, "OUT5"), (22, "OUT6"), (23, "OUT7"), (24, "OUT8"),
            ),
        ),
        (
            "TPS22990DML",
            "U",
            "TPS22990DMLR",
            "",
            "https://www.ti.com/lit/ds/symlink/tps22990.pdf",
            "5.5 V 10 A controlled-rise-time load switch",
            (
                (1, "CT"), (2, "NC"), (3, "VIN"), (4, "VBIAS"),
                (5, "ON"), (6, "GND"), (7, "PG"), (8, "VOUT1"),
                (9, "VOUT2"), (10, "VOUT3"),
            ),
        ),
        (
            "TRI20_1223",
            "U",
            "TRI 20-1223",
            "",
            "https://www.tracopower.com/tri20-datasheet",
            "20 W reinforced-isolation 12 V to dual 15 V DC/DC module",
            (
                (1, "+VIN"), (2, "-VIN"), (3, "+VOUT"),
                (4, "COMMON"), (5, "-VOUT"),
            ),
        ),
        (
            "TPS7A20DBV_A1",
            "U",
            "TPS7A20 fixed DBV",
            "Package_TO_SOT_SMD:SOT-23-5",
            "https://www.ti.com/lit/ds/symlink/tps7a20.pdf",
            "300 mA low-noise fixed-output LDO in SOT-23-5",
            (
                (1, "IN"), (2, "GND"), (3, "EN"), (4, "NC"), (5, "OUT"),
            ),
        ),
        (
            "PCA9517ADP_A1",
            "U",
            "PCA9517ADP",
            "Package_SO:TSSOP-8_3x3mm_P0.65mm",
            "https://www.nxp.com/docs/en/data-sheet/PCA9517A.pdf",
            "Fast-mode I2C repeater used to isolate the off-board audio-control bus",
            (
                (1, "VCCA"), (2, "SCLA"), (3, "SDAA"), (4, "GND"),
                (5, "EN"), (6, "SDAB"), (7, "SCLB"), (8, "VCCB"),
            ),
        ),
        (
            "FSA2567MPX",
            "U",
            "FSA2567MPX",
            "Package_DFN_QFN:WQFN-16-1EP_3x3mm_P0.5mm_EP1.68x1.68mm",
            "https://www.onsemi.com/pdf/datasheet/fsa2567-d.pdf",
            "Dual-SIM four-pole analog switch",
            (
                (1, "1VSIM"), (2, "SEL"), (3, "2RST"), (4, "RST"),
                (5, "1RST"), (6, "GND"), (7, "2CLK"), (8, "CLK"),
                (9, "1CLK"), (10, "NC"), (11, "2DAT"), (12, "DAT"),
                (13, "1DAT"), (14, "VCC"), (15, "2VSIM"), (16, "VSIM"),
                (17, "EP"),
            ),
        ),
        (
            "Mini_PCIe_52",
            "J",
            "Molex 0679101002 / AW7915-NP1",
            "CM5Carrier:Molex_0679101002_Mini_PCIe",
            "https://www.molex.com/en-us/products/part-detail/0679101002",
            "52-contact Mini PCIe socket with the standard PCIe/USB/UIM pin assignment",
            (
                (1, "WAKE#"), (2, "+3V3AUX"), (3, "COEX1"), (4, "GND"),
                (5, "COEX2"), (6, "+1V5"), (7, "CLKREQ#"), (8, "UIM_PWR"),
                (9, "GND"), (10, "UIM_DATA"), (11, "REFCLK-"), (12, "UIM_CLK"),
                (13, "REFCLK+"), (14, "UIM_RESET"), (15, "GND"), (16, "UIM_VPP"),
                (17, "RESERVED"), (18, "GND"), (19, "RESERVED"), (20, "W_DISABLE#"),
                (21, "GND"), (22, "PERST#"), (23, "PERn0"), (24, "+3V3AUX"),
                (25, "PERp0"), (26, "GND"), (27, "GND"), (28, "+1V5"),
                (29, "GND"), (30, "SMB_CLK"), (31, "PETn0"), (32, "SMB_DATA"),
                (33, "PETp0"), (34, "GND"), (35, "GND"), (36, "USB_D-"),
                (37, "GND"), (38, "USB_D+"), (39, "+3V3AUX"), (40, "GND"),
                (41, "+3V3AUX"), (42, "LED_WWAN#"), (43, "GND"), (44, "LED_WLAN#"),
                (45, "RESERVED"), (46, "LED_WPAN#"), (47, "RESERVED"), (48, "+1V5"),
                (49, "RESERVED"), (50, "GND"), (51, "RESERVED"), (52, "+3V3AUX"),
            ),
        ),
        (
            "Wurth_Nano_SIM_693043020611",
            "J",
            "693043020611",
            "CM5Carrier:J_Wurth_WR-CRD_693043020611",
            "https://www.we-online.com/components/products/datasheet/693043020611.pdf",
            "Wurth WR-CRD push-push nano-SIM holder",
            (
                ("C1", "VCC"), ("C2", "RESET"), ("C3", "CLOCK"),
                ("C5", "GND"), ("C6", "VPP_NC"), ("C7", "I/O"),
                ("S1", "SHIELD"), ("S2", "SHIELD"), ("S3", "SHIELD"),
                ("S4", "SHIELD"), ("S5", "SHIELD"), ("S6", "SHIELD"),
            ),
        ),
        (
            "ES8316",
            "U",
            "ES8316",
            "Package_DFN_QFN:QFN-32-1EP_4x4mm_P0.4mm_EP2.9x2.9mm_ThermalVias",
            "https://dl.radxa.com/cm5/radxa_cm5_io_board_v2200_schematic.pdf",
            "Low-power stereo audio codec; exact pinout follows Radxa CM5 IO V2.2",
            (
                (1, "CCLK"), (2, "MCLK"), (3, "DVDD"), (4, "PVDD"),
                (5, "DGND"), (6, "SCLK"), (7, "DSDIN"), (8, "DLRCK"),
                (9, "ASDOUT"), (10, "GPIO1"), (11, "GPIO2"), (12, "GPIO3"),
                (13, "CPVSSP"), (14, "CPVDD"), (15, "CPTOP"), (16, "CPBOT"),
                (17, "CPGND"), (18, "CPGNDREF"), (19, "ROUT"), (20, "LOUT"),
                (21, "DACVREF"), (22, "AVDD"), (23, "AGND"), (24, "ADCVREF"),
                (25, "VMID"), (26, "MICBIAS"), (27, "RIN2"), (28, "LIN2"),
                (29, "RIN1"), (30, "LIN1"), (31, "CE"), (32, "CDATA"),
                (33, "E-PAD"),
            ),
        ),
        (
            "AK5558VN",
            "U",
            "AK5558VN",
            "Package_DFN_QFN:QFN-64-1EP_9x9mm_P0.5mm_EP6x6mm_ThermalVias",
            "https://www.akm-semi.com/pdf-0f/ak5558vn.pdf",
            "Eight-channel 32-bit ADC; exact 64-QFN pin assignment",
            (
                (1, "AVSS1"), (2, "AVDD1"), (3, "AIN3P"), (4, "AIN3N"),
                (5, "VREFL2"), (6, "VREFH2"), (7, "AIN4N"), (8, "AIN4P"),
                (9, "AIN5P"), (10, "AIN5N"), (11, "VREFH3"), (12, "VREFL3"),
                (13, "AIN6N"), (14, "AIN6P"), (15, "AVDD2"), (16, "AVSS2"),
                (17, "AIN7P"), (18, "AIN7N"), (19, "VREFH4"), (20, "VREFL4"),
                (21, "AIN8N"), (22, "AIN8P"), (23, "TEST"), (24, "MCLK"),
                (25, "TVDD"), (26, "DVSS"), (27, "VDD18"), (28, "PDN"),
                (29, "PW0"), (30, "PW1"), (31, "PW2"), (32, "MSN"),
                (33, "BICK/DCLK"), (34, "LRCK/DSDOL1"), (35, "TDMIN/DSDOR1"),
                (36, "SDTO1/DSDOL2"), (37, "SDTO2/DSDOR2"),
                (38, "SDTO3/DSDOL3"), (39, "SDTO4/DSDOR3"),
                (40, "DSDOL4"), (41, "DSDOR4"), (42, "OVF"),
                (43, "CKS0/SDA/CDTI"), (44, "CKS1/CAD0_I2C/CSN"),
                (45, "CKS2/SCL/CCLK"), (46, "CKS3/CAD1"),
                (47, "SLOW/DCKB"), (48, "SD/PMOD"), (49, "DIF0/DSDSEL0"),
                (50, "DIF1/DSDSEL1"), (51, "TDM0"), (52, "TDM1"),
                (53, "PSN/CAD0_SPI"), (54, "I2C"), (55, "DP"),
                (56, "HPFE/DCKS"), (57, "LDOE"), (58, "ODP"),
                (59, "AIN1P"), (60, "AIN1N"), (61, "VREFL1"),
                (62, "VREFH1"), (63, "AIN2N"), (64, "AIN2P"),
                (65, "EP"),
            ),
        ),
        (
            "AK4458VN",
            "U",
            "AK4458VN",
            "Package_DFN_QFN:QFN-48-1EP_7x7mm_P0.5mm_EP5.15x5.15mm_ThermalVias",
            "https://www.akm-semi.com/pdf-c1/ak4458vn.pdf",
            "Eight-channel premium DAC; exact 48-QFN pin assignment",
            (
                (1, "MCLK"), (2, "BICK/DCLK"), (3, "LRCK/DSDL1"),
                (4, "SDTI1/DSDR1"), (5, "SDTI2/DSDL2"),
                (6, "SDTI3/DSDR2/TDMO1"), (7, "SDTI4/DSDL3/TDMO2"),
                (8, "DSDR3"), (9, "DSDL4"), (10, "DSDR4"),
                (11, "DZF/SMUTE"), (12, "CAD1/DCHAIN"),
                (13, "SDA/CDTI/TDM0"), (14, "SCL/CCLK/TDM1"),
                (15, "CAD0_I2C/CSN/DIF"), (16, "PS/CAD0_SPI"), (17, "I2C"),
                (18, "AOUTL1P"), (19, "AOUTL1N"), (20, "VREFL1"),
                (21, "VREFH1"), (22, "AOUTR1N"), (23, "AOUTR1P"),
                (24, "AOUTL2P"), (25, "AOUTL2N"), (26, "VREFL2"),
                (27, "VREFH2"), (28, "AOUTR2N"), (29, "AOUTR2P"),
                (30, "AVSS"), (31, "AVDD"), (32, "AOUTL3P"),
                (33, "AOUTL3N"), (34, "VREFH3"), (35, "VREFL3"),
                (36, "AOUTR3N"), (37, "AOUTR3P"), (38, "AOUTL4P"),
                (39, "AOUTL4N"), (40, "VREFH4"), (41, "VREFL4"),
                (42, "AOUTR4N"), (43, "AOUTR4P"), (44, "LDOE"),
                (45, "TVDD"), (46, "DVSS"), (47, "VDD18"), (48, "PDN"),
                (49, "EP"),
            ),
        ),
        (
            "THAT1206",
            "U",
            "THAT1206S08-U",
            "Package_SO:SOIC-8_3.9x4.9mm_P1.27mm",
            "https://www.thatcorp.com/datashts/THAT_1200-Series_Datasheet.pdf",
            "Minus-six-decibel InGenius balanced line receiver",
            ((1, "REF"), (2, "IN-"), (3, "IN+"), (4, "VEE"),
             (5, "CM_IN"), (6, "VOUT"), (7, "VCC"), (8, "CM_OUT")),
        ),
        (
            "THAT1646",
            "U",
            "THAT1646S08-U",
            "Package_SO:SOIC-8_3.9x4.9mm_P1.27mm",
            "https://thatcorp.com/datashts/THAT_1606-1646_Datasheet.pdf",
            "OutSmarts balanced line driver with six-decibel gain",
            ((1, "OUT-"), (2, "SNS-"), (3, "GND"), (4, "IN"),
             (5, "VEE"), (6, "VCC"), (7, "SNS+"), (8, "OUT+")),
        ),
        (
            "OPA1652",
            "U",
            "OPA1652AIDR",
            "Package_SO:SOIC-8_3.9x4.9mm_P1.27mm",
            "https://www.ti.com/lit/ds/symlink/opa1654.pdf",
            "Dual low-noise FET-input audio operational amplifier",
            ((1, "OUT_A"), (2, "IN_A-"), (3, "IN_A+"), (4, "V-"),
             (5, "IN_B+"), (6, "IN_B-"), (7, "OUT_B"), (8, "V+")),
        ),
        (
            "Panasonic_TQ2_12V",
            "K",
            "TQ2-12V",
            "CM5Carrier:Panasonic_TQ2-12V_PRELIMINARY",
            "https://na.industrial.panasonic.com/products/relays-contactors/mechanical-signal-relays/series/119572/model/119888",
            "Sealed 12 V DPDT signal relay; footprint remains drawing/coupon gated",
            ((1, "COIL+"), (2, "NC1"), (3, "COM1"), (4, "NO1"),
             (7, "NO2"), (8, "COM2"), (9, "NC2"), (10, "COIL-")),
        ),
        (
            "Kycon_STX_353K7A_6N",
            "J",
            "STX-353K7A-6N-KTTR",
            "CM5Carrier:Kycon_STX-353K7A-6N-KTTR_PRELIMINARY",
            "https://www.kycon.com/Catalog_PDF/STX-353K7A.pdf",
            "Kycon vertical four-pole jack with an isolated insertion switch",
            (
                (1, "SLEEVE_MIC"), (2, "R2_GND"), (3, "R1_HP_R"),
                (4, "TIP_HP_L"), (5, "SWITCH_A"), (6, "SWITCH_B"),
            ),
        ),
    )
    for name, reference, value, footprint, datasheet, description, pins in local_symbols:
        rows = (len(pins) + 1) // 2
        half_height = max(10.16, rows * 1.27 + 2.54)
        lines.extend(
            [
                f'  (symbol "{name}" (pin_names (offset 0.9)) (in_bom yes) (on_board yes)',
                f'    (property "Reference" "{reference}" (at {-17.78:.2f} {half_height + 3.81:.2f} 0) (effects (font (size 1.27 1.27)) (justify left)))',
                f'    (property "Value" "{value}" (at {-17.78:.2f} {half_height + 1.27:.2f} 0) (effects (font (size 1.27 1.27)) (justify left)))',
                f'    (property "Footprint" "{footprint}" (at 0 0 0) (effects (font (size 1.27 1.27)) hide))',
                f'    (property "Datasheet" "{datasheet}" (at 0 0 0) (effects (font (size 1.27 1.27)) hide))',
                f'    (property "Description" "{description}" (at 0 0 0) (effects (font (size 1.27 1.27)) hide))',
                f'    (symbol "{name}_0_1"',
                f'      (rectangle (start -17.78 {half_height:.2f}) (end 17.78 {-half_height:.2f}) (stroke (width 0.254) (type default)) (fill (type background)))',
                "    )",
                f'    (symbol "{name}_1_1"',
            ]
        )
        for index, (pin, signal) in enumerate(pins):
            row = index // 2
            y = half_height - 2.54 - row * 2.54
            is_left = index % 2 == 0
            x = -20.32 if is_left else 20.32
            orientation = 0 if is_left else 180
            lines.append(
                f'      (pin passive line (at {x:.2f} {y:.2f} {orientation}) (length 2.54) '
                f'(name "{_escape_symbol_text(signal)}" (effects (font (size 0.72 0.72)))) '
                f'(number "{pin}" (effects (font (size 0.72 0.72)))))'
            )
        lines.extend(["    )", "  )"])

    cm5_audio_port_pins = (
        (1, "AUD_MCLK", "output"), (2, "AUD_BCLK", "output"),
        (3, "AUD_FSYNC", "output"), (4, "AUD_DAC_SDIN", "output"),
        (5, "AUD_ADC_SDOUT", "passive"), (6, "SYS_I2C7_SCL", "bidirectional"),
        (7, "SYS_I2C7_SDA", "bidirectional"), (8, "AUD_IRQ_N", "passive"),
        (9, "HS_MCLK", "output"), (10, "HS_BCLK", "output"),
        (11, "HS_LRCK", "output"), (12, "HS_SDOUT_TO_CODEC", "output"),
        (13, "HS_SDIN_FROM_CODEC", "passive"), (14, "HS_I2C_SCL", "bidirectional"),
        (15, "HS_I2C_SDA", "bidirectional"), (16, "HS_JACK_DET_N", "passive"),
    )
    lines.extend(
        [
            '  (symbol "CM5_Audio_Port" (pin_names (offset 0.9)) (in_bom no) (on_board no)',
            '    (property "Reference" "U" (at -20.32 25.4 0) (effects (font (size 1.27 1.27)) (justify left)))',
            '    (property "Value" "CM5 AUDIO OFF-SHEET PORT" (at -20.32 22.86 0) (effects (font (size 1.27 1.27)) (justify left)))',
            '    (property "Footprint" "" (at 0 0 0) (effects (font (size 1.27 1.27)) hide))',
            '    (property "Datasheet" "" (at 0 0 0) (effects (font (size 1.27 1.27)) hide))',
            '    (property "Description" "Typed off-sheet CM5 audio interface for ERC" (at 0 0 0) (effects (font (size 1.27 1.27)) hide))',
            '    (symbol "CM5_Audio_Port_0_1"',
            '      (rectangle (start -20.32 20.32) (end 20.32 -20.32) (stroke (width 0.254) (type default)) (fill (type background)))',
            '    )',
            '    (symbol "CM5_Audio_Port_1_1"',
        ]
    )
    for index, (pin, signal, pin_type) in enumerate(cm5_audio_port_pins):
        row = index // 2
        y = 17.78 - row * 5.08
        is_left = index % 2 == 0
        x = -22.86 if is_left else 22.86
        orientation = 0 if is_left else 180
        lines.append(
            f'      (pin {pin_type} line (at {x:.2f} {y:.2f} {orientation}) (length 2.54) '
            f'(name "{signal}" (effects (font (size 0.75 0.75)))) '
            f'(number "{pin}" (effects (font (size 0.75 0.75)))))'
        )
    lines.extend(["    )", "  )"])
    lines.append(")")
    CM5_LOCAL_LIBRARY.parent.mkdir(parents=True, exist_ok=True)
    CM5_LOCAL_LIBRARY.write_text("\n".join(lines) + "\n")
    shutil.copyfile(
        WURTH_74991114412 / "Transformer_Wurth_WE-RJ45LAN.kicad_sym",
        CM5_WURTH_LIBRARY,
    )
    (CM5_LOCAL_LIBRARY.parent / "sym-lib-table").write_text(
        '(sym_lib_table\n'
        '  (version 7)\n'
        '  (lib (name "CM5Carrier")(type "KiCad")(uri "${KIPRJMOD}/CM5Carrier.kicad_sym")(options "")(descr "Radxa CM5 carrier-local symbols"))\n'
        '  (lib (name "WurthRJ45")(type "KiCad")(uri "${KIPRJMOD}/WurthRJ45.kicad_sym")(options "")(descr "Wurth manufacturer RJ45 symbols"))\n'
        ')\n'
    )
    audio_folder = ROOT / "AUDIO-8X8"
    audio_folder.mkdir(parents=True, exist_ok=True)
    (audio_folder / "sym-lib-table").write_text(
        '(sym_lib_table\n'
        '  (version 7)\n'
        '  (lib (name "CM5Carrier")(type "KiCad")(uri "${KIPRJMOD}/../CM5-CARRIER/CM5Carrier.kicad_sym")(options "")(descr "Shared Radxa CM5 carrier and audio symbols"))\n'
        ')\n'
    )
    return pinout


def write_cm5_local_footprints() -> None:
    """Copy controlled connector mechanics into this project's local library."""
    CM5_LOCAL_FOOTPRINTS.mkdir(parents=True, exist_ok=True)
    CM5_LOCAL_3DMODELS.mkdir(parents=True, exist_ok=True)
    filenames = (
        "TE_2199230-3_M2_Key_B_4.2mm.kicad_mod",
        "J_Wurth_WR-CRD_693043020611.kicad_mod",
    )
    for filename in filenames:
        source = LEGACY_CELLULAR_FOOTPRINTS / filename
        if not source.exists():
            raise FileNotFoundError(f"Controlled connector footprint missing: {source}")
        shutil.copyfile(source, CM5_LOCAL_FOOTPRINTS / filename)
    shutil.copyfile(
        WURTH_74991114412 / "T_Wurth_WE-RJ45LAN_74991114412.kicad_mod",
        CM5_LOCAL_FOOTPRINTS / "T_Wurth_WE-RJ45LAN_74991114412.kicad_mod",
    )
    shutil.copyfile(
        WURTH_74991114412 / "T_Wurth_WE-RJ45LAN_74991114412.step",
        CM5_LOCAL_3DMODELS / "T_Wurth_WE-RJ45LAN_74991114412.step",
    )
    write_molex_0679101002_footprint()
    write_kycon_stx_353k7a_footprint()
    write_panasonic_tq2_footprint()
    write_traco_tri20_footprint()
    write_power_magnetic_and_shunt_footprints()
    write_power_semiconductor_footprints()
    (CM5_LOCAL_LIBRARY.parent / "fp-lib-table").write_text(
        '(fp_lib_table\n'
        '  (version 7)\n'
        '  (lib (name "CM5Carrier")(type "KiCad")(uri "${KIPRJMOD}/CM5Carrier.pretty")(options "")(descr "Radxa CM5 carrier-local footprints"))\n'
        ')\n'
    )
    (ROOT / "AUDIO-8X8" / "fp-lib-table").write_text(
        '(fp_lib_table\n'
        '  (version 7)\n'
        '  (lib (name "CM5Carrier")(type "KiCad")(uri "${KIPRJMOD}/../CM5-CARRIER/CM5Carrier.pretty")(options "")(descr "Shared Radxa CM5 carrier and audio footprints"))\n'
        ')\n'
    )


def write_kycon_stx_353k7a_footprint() -> None:
    """Generate a drawing-derived, coupon-gated land pattern for the CTIA jack.

    Kycon's catalog drawing defines the terminal center locations and component
    envelope but does not publish a recommended PCB land.  This footprint is
    intentionally marked preliminary and must be checked against a physical
    STX-353K7A-6N-KTTR sample and a plated-hole coupon before production.
    """

    name = "Kycon_STX-353K7A-6N-KTTR_PRELIMINARY"

    def fp_uuid(label: str) -> str:
        return str(
            _uuid.uuid5(
                _uuid.NAMESPACE_URL,
                f"radxa-cm5-procomm:kycon-stx-353k7a:{label}",
            )
        )

    lines = [
        f'(footprint "{name}"',
        '  (version 20240108)',
        '  (generator "procomm_cm5_generator")',
        '  (layer "F.Cu")',
        '  (descr "Kycon STX-353K7A-6N-KTTR vertical four-pole jack; terminal centers reconstructed from the Kycon component drawing; PRELIMINARY - SAMPLE AND COUPON VERIFICATION REQUIRED")',
        '  (tags "Kycon STX-353K7A CTIA preliminary coupon required")',
        '  (attr through_hole)',
        '  (duplicate_pad_numbers_are_jumpers no)',
        f'  (property "Reference" "J**" (at 0 -14.25 0) (layer "F.SilkS") (uuid "{fp_uuid("reference")}") (effects (font (size 1 1) (thickness 0.15))))',
        f'  (property "Value" "{name}" (at 0 2.75 0) (layer "F.Fab") (uuid "{fp_uuid("value")}") (effects (font (size 1 1) (thickness 0.15))))',
        '  (property "Datasheet" "https://www.kycon.com/Catalog_PDF/STX-353K7A.pdf" (at 0 0 0) (layer "F.Fab") (hide yes) '
        f'(uuid "{fp_uuid("datasheet")}") (effects (font (size 1.27 1.27))))',
        '  (property "Description" "Drawing-derived preliminary footprint; physical sample and coupon are release gates" (at 0 0 0) (layer "F.Fab") (hide yes) '
        f'(uuid "{fp_uuid("description")}") (effects (font (size 1.27 1.27))))',
        f'  (fp_rect (start -3.825 -12.50) (end 3.825 0.00) (stroke (width 0.10) (type default)) (fill none) (layer "F.Fab") (uuid "{fp_uuid("fab-body")}"))',
        f'  (fp_rect (start -4.15 -12.85) (end 4.15 1.05) (stroke (width 0.05) (type default)) (fill none) (layer "F.CrtYd") (uuid "{fp_uuid("courtyard")}"))',
        f'  (fp_line (start -3.825 -12.50) (end 3.825 -12.50) (stroke (width 0.15) (type default)) (layer "F.SilkS") (uuid "{fp_uuid("silk-back")}"))',
        f'  (fp_line (start -3.825 -12.50) (end -3.825 -0.70) (stroke (width 0.15) (type default)) (layer "F.SilkS") (uuid "{fp_uuid("silk-left")}"))',
        f'  (fp_line (start 3.825 -12.50) (end 3.825 -0.70) (stroke (width 0.15) (type default)) (layer "F.SilkS") (uuid "{fp_uuid("silk-right")}"))',
        f'  (fp_text user "PRELIMINARY - COUPON REQUIRED" (at 0 -8.8 0) (layer "F.Fab") (uuid "{fp_uuid("coupon-note")}") (effects (font (size 0.8 0.8) (thickness 0.12))))',
    ]
    # Pin 1 is the drawing datum.  Pins 5/6/1/4 form the rear terminal row;
    # pins 2/3 form the forward row at the drawing's 5.10 mm row spacing.
    pads = (
        (1, 0.000, 0.000),
        (2, -1.725, -5.100),
        (3, 1.725, -5.100),
        (4, 3.070, 0.000),
        (5, -1.300, 0.000),
        (6, -3.350, 0.000),
    )
    for pin, x, y in pads:
        shape = "roundrect" if pin == 1 else "circle"
        lines.extend(
            [
                f'  (pad "{pin}" thru_hole {shape}',
                f'    (at {x:.3f} {y:.3f}) (size 1.80 1.80) (drill 1.00)',
                '    (layers "*.Cu" "*.Mask")',
            ]
        )
        if pin == 1:
            lines.append('    (roundrect_rratio 0.25)')
        lines.extend([f'    (uuid "{fp_uuid(f"pad-{pin}")}")', '  )'])
    lines.extend(['  (embedded_fonts no)', ')'])
    (CM5_LOCAL_FOOTPRINTS / f"{name}.kicad_mod").write_text(
        "\n".join(lines) + "\n"
    )


def write_panasonic_tq2_footprint() -> None:
    """Generate the preliminary TQ2 through-hole pattern for capture review.

    The 2.54 mm contact pitch and 7.62 mm row spacing are held as provisional
    dimensions until Panasonic's current engineering drawing and a relay
    insertion coupon are signed off.  Routing is intentionally gated on that
    verification in the footprint audit.
    """
    name = "Panasonic_TQ2-12V_PRELIMINARY"

    def fp_uuid(label: str) -> str:
        return str(
            _uuid.uuid5(
                _uuid.NAMESPACE_URL,
                f"radxa-cm5-procomm:panasonic-tq2:{label}",
            )
        )

    lines = [
        f'(footprint "{name}"',
        '  (version 20240108)',
        '  (generator "procomm_cm5_generator")',
        '  (layer "F.Cu")',
        '  (descr "Panasonic TQ2 standard PC-board relay; PRELIMINARY - current engineering drawing and insertion coupon required")',
        '  (tags "Panasonic TQ2 preliminary coupon required")',
        '  (attr through_hole)',
        '  (duplicate_pad_numbers_are_jumpers no)',
        f'  (property "Reference" "K**" (at 0 -5.8 0) (layer "F.SilkS") (uuid "{fp_uuid("reference")}") (effects (font (size 1 1) (thickness 0.15))))',
        f'  (property "Value" "{name}" (at 0 5.8 0) (layer "F.Fab") (uuid "{fp_uuid("value")}") (effects (font (size 1 1) (thickness 0.15))))',
        '  (property "Datasheet" "https://na.industrial.panasonic.com/products/relays-contactors/mechanical-signal-relays/series/119572/model/119888" (at 0 0 0) (layer "F.Fab") (hide yes) '
        f'(uuid "{fp_uuid("datasheet")}") (effects (font (size 1.27 1.27))))',
        f'  (fp_rect (start -7.1 -4.5) (end 7.1 4.5) (stroke (width 0.10) (type default)) (fill none) (layer "F.Fab") (uuid "{fp_uuid("fab-body")}"))',
        f'  (fp_rect (start -7.35 -4.75) (end 7.35 4.75) (stroke (width 0.05) (type default)) (fill none) (layer "F.CrtYd") (uuid "{fp_uuid("courtyard")}"))',
        f'  (fp_line (start -7.1 -4.5) (end 7.1 -4.5) (stroke (width 0.15) (type default)) (layer "F.SilkS") (uuid "{fp_uuid("silk-top")}"))',
        f'  (fp_text user "PRELIMINARY - DRAWING/COUPON" (at 0 0 0) (layer "F.Fab") (uuid "{fp_uuid("gate-note")}") (effects (font (size 0.7 0.7) (thickness 0.1))))',
    ]
    pads = (
        (1, -3.81, -3.81), (2, -1.27, -3.81),
        (3, 1.27, -3.81), (4, 3.81, -3.81),
        (10, -3.81, 3.81), (9, -1.27, 3.81),
        (8, 1.27, 3.81), (7, 3.81, 3.81),
    )
    for pin, x, y in pads:
        shape = "roundrect" if pin == 1 else "circle"
        lines.extend(
            [
                f'  (pad "{pin}" thru_hole {shape}',
                f'    (at {x:.2f} {y:.2f}) (size 1.7 1.7) (drill 0.9)',
                '    (layers "*.Cu" "*.Mask")',
            ]
        )
        if pin == 1:
            lines.append('    (roundrect_rratio 0.25)')
        lines.extend([f'    (uuid "{fp_uuid(f"pad-{pin}")}")', '  )'])
    lines.extend(['  (embedded_fonts no)', ')'])
    (CM5_LOCAL_FOOTPRINTS / f"{name}.kicad_mod").write_text(
        "\n".join(lines) + "\n"
    )


def write_traco_tri20_footprint() -> None:
    """Generate the TRI 20 dual-output footprint from Traco's Rev 2024 drawing."""
    name = "TRACO_TRI20_DUAL"

    def fp_uuid(label: str) -> str:
        return str(
            _uuid.uuid5(
                _uuid.NAMESPACE_URL,
                f"radxa-cm5-procomm:traco-tri20:{label}",
            )
        )

    lines = [
        f'(footprint "{name}"',
        '  (version 20240108)',
        '  (generator "procomm_cm5_generator")',
        '  (layer "F.Cu")',
        '  (descr "Traco Power TRI 20 dual-output DC/DC; exact 50.8 x 25.4 mm outline and five-pin pattern from Rev August 7 2024 drawing")',
        '  (tags "Traco TRI20 2x1 inch DIP dual output")',
        '  (attr through_hole)',
        '  (duplicate_pad_numbers_are_jumpers no)',
        f'  (property "Reference" "U**" (at 0 -14.7 0) (layer "F.SilkS") (uuid "{fp_uuid("reference")}") (effects (font (size 1 1) (thickness 0.15))))',
        f'  (property "Value" "{name}" (at 0 14.7 0) (layer "F.Fab") (uuid "{fp_uuid("value")}") (effects (font (size 1 1) (thickness 0.15))))',
        '  (property "Datasheet" "https://www.tracopower.com/tri20-datasheet" (at 0 0 0) (layer "F.Fab") (hide yes) '
        f'(uuid "{fp_uuid("datasheet")}") (effects (font (size 1.27 1.27))))',
        f'  (fp_rect (start -25.4 -12.7) (end 25.4 12.7) (stroke (width 0.10) (type default)) (fill none) (layer "F.Fab") (uuid "{fp_uuid("fab-body")}"))',
        f'  (fp_rect (start -25.7 -13.0) (end 25.7 13.0) (stroke (width 0.05) (type default)) (fill none) (layer "F.CrtYd") (uuid "{fp_uuid("courtyard")}"))',
        f'  (fp_line (start -25.4 -12.7) (end 25.4 -12.7) (stroke (width 0.15) (type default)) (layer "F.SilkS") (uuid "{fp_uuid("silk-top")}"))',
        f'  (fp_line (start -25.4 12.7) (end 25.4 12.7) (stroke (width 0.15) (type default)) (layer "F.SilkS") (uuid "{fp_uuid("silk-bottom")}"))',
        f'  (fp_line (start -25.4 -12.7) (end -25.4 12.7) (stroke (width 0.15) (type default)) (layer "F.SilkS") (uuid "{fp_uuid("silk-left")}"))',
        f'  (fp_line (start 25.4 -12.7) (end 25.4 12.7) (stroke (width 0.15) (type default)) (layer "F.SilkS") (uuid "{fp_uuid("silk-right")}"))',
        f'  (fp_circle (center -23.7 -6.8) (end -23.1 -6.8) (stroke (width 0.15) (type default)) (fill none) (layer "F.SilkS") (uuid "{fp_uuid("pin-one")}"))',
        f'  (fp_text user "BOTTOM-VIEW PIN DATUM" (at 0 0 0) (layer "F.Fab") (uuid "{fp_uuid("datum-note")}") (effects (font (size 0.8 0.8) (thickness 0.12))))',
    ]
    # Manufacturer bottom view: 45.72 mm column pitch. Input pins are 5.08 mm
    # apart; dual-output pins are on 5.08 mm pitch around the center common.
    pads = (
        (1, -22.86, -2.54),
        (2, -22.86, 2.54),
        (3, 22.86, -5.08),
        (4, 22.86, 0.00),
        (5, 22.86, 5.08),
    )
    for pin, x, y in pads:
        shape = "roundrect" if pin == 1 else "circle"
        lines.extend(
            [
                f'  (pad "{pin}" thru_hole {shape}',
                f'    (at {x:.2f} {y:.2f}) (size 2.20 2.20) (drill 1.20)',
                '    (layers "*.Cu" "*.Mask")',
            ]
        )
        if pin == 1:
            lines.append('    (roundrect_rratio 0.25)')
        lines.extend([f'    (uuid "{fp_uuid(f"pad-{pin}")}")', '  )'])
    lines.extend(['  (embedded_fonts no)', ')'])
    (CM5_LOCAL_FOOTPRINTS / f"{name}.kicad_mod").write_text(
        "\n".join(lines) + "\n"
    )


def write_power_magnetic_and_shunt_footprints() -> None:
    """Generate simple two-pad lands from controlled manufacturer drawings."""

    def write_footprint(
        name: str,
        reference_prefix: str,
        body: tuple[float, float],
        pad_size: tuple[float, float],
        pad_centers: tuple[tuple[float, float], tuple[float, float]],
        datasheet: str,
        description: str,
    ) -> None:
        def fp_uuid(label: str) -> str:
            return str(
                _uuid.uuid5(
                    _uuid.NAMESPACE_URL,
                    f"radxa-cm5-procomm:power-footprints:{name}:{label}",
                )
            )

        body_x, body_y = body
        pad_x, pad_y = pad_size
        extent_x = max(body_x / 2, *(abs(x) + pad_x / 2 for x, _ in pad_centers))
        extent_y = max(body_y / 2, *(abs(y) + pad_y / 2 for _, y in pad_centers))
        courtyard_x = extent_x + 0.25
        courtyard_y = extent_y + 0.25
        lines = [
            f'(footprint "{name}"',
            '  (version 20240108)',
            '  (generator "procomm_cm5_generator")',
            '  (layer "F.Cu")',
            f'  (descr "{description}")',
            '  (attr smd)',
            '  (duplicate_pad_numbers_are_jumpers no)',
            f'  (property "Reference" "{reference_prefix}**" (at 0 {-courtyard_y - 1.0:.2f} 0) (layer "F.SilkS") (uuid "{fp_uuid("reference")}") (effects (font (size 1 1) (thickness 0.15))))',
            f'  (property "Value" "{name}" (at 0 {courtyard_y + 1.0:.2f} 0) (layer "F.Fab") (uuid "{fp_uuid("value")}") (effects (font (size 1 1) (thickness 0.15))))',
            f'  (property "Datasheet" "{datasheet}" (at 0 0 0) (layer "F.Fab") (hide yes) (uuid "{fp_uuid("datasheet")}") (effects (font (size 1.27 1.27))))',
            f'  (fp_rect (start {-body_x / 2:.2f} {-body_y / 2:.2f}) (end {body_x / 2:.2f} {body_y / 2:.2f}) (stroke (width 0.10) (type default)) (fill none) (layer "F.Fab") (uuid "{fp_uuid("fab")}"))',
            f'  (fp_rect (start {-courtyard_x:.2f} {-courtyard_y:.2f}) (end {courtyard_x:.2f} {courtyard_y:.2f}) (stroke (width 0.05) (type default)) (fill none) (layer "F.CrtYd") (uuid "{fp_uuid("courtyard")}"))',
            f'  (fp_line (start {-body_x / 2:.2f} {-body_y / 2:.2f}) (end {body_x / 2:.2f} {-body_y / 2:.2f}) (stroke (width 0.15) (type default)) (layer "F.SilkS") (uuid "{fp_uuid("silk-top")}"))',
            f'  (fp_line (start {-body_x / 2:.2f} {body_y / 2:.2f}) (end {body_x / 2:.2f} {body_y / 2:.2f}) (stroke (width 0.15) (type default)) (layer "F.SilkS") (uuid "{fp_uuid("silk-bottom")}"))',
        ]
        for index, (x, y) in enumerate(pad_centers, start=1):
            lines.extend(
                [
                    f'  (pad "{index}" smd roundrect',
                    f'    (at {x:.3f} {y:.3f}) (size {pad_x:.3f} {pad_y:.3f})',
                    '    (layers "F.Cu" "F.Paste" "F.Mask")',
                    '    (roundrect_rratio 0.10)',
                    f'    (uuid "{fp_uuid(f"pad-{index}")}")',
                    '  )',
                ]
            )
        lines.extend(['  (embedded_fonts no)', ')'])
        (CM5_LOCAL_FOOTPRINTS / f"{name}.kicad_mod").write_text("\n".join(lines) + "\n")

    write_footprint(
        "TDK_SPM10065VC",
        "L",
        (10.5, 10.0),
        (2.95, 4.50),
        ((-4.525, 0.0), (4.525, 0.0)),
        "https://product.tdk.com/en/search/inductor/inductor/smd/info?part_no=SPM10065VC-3R3M-D",
        "TDK SPM10065VC; 10.5 x 10.0 mm body; A=2.95, B=6.10, C=4.50 mm recommended land",
    )
    write_footprint(
        "Wurth_74439370047",
        "L",
        (16.4, 15.4),
        (14.1, 3.3),
        ((0.0, -5.3), (0.0, 5.3)),
        "https://www.we-online.com/components/products/datasheet/74439370047.pdf",
        "Wurth 74439370047 WE-XHMI; manufacturer recommended 14.1 x 3.3 mm lands with 7.3 mm gap",
    )
    write_footprint(
        "Susumu_KRL6432E_6mR",
        "R",
        (3.1, 6.3),
        (1.0, 6.6),
        ((-1.6, 0.0), (1.6, 0.0)),
        "https://www.susumu.co.jp/en/tech/rep-data-rand/",
        "Susumu KRL6432E long-side terminal, 2 mOhm and above land: a=2.20, b=4.20, c=6.60 mm",
    )
    write_footprint(
        "Susumu_KRL11050_4mR",
        "R",
        (5.0, 11.0),
        (1.0, 11.2),
        ((-2.3, 0.0), (2.3, 0.0)),
        "https://www.susumu.co.jp/en/tech/rep-data-rand/",
        "Susumu KRL11050 long-side terminal, 2 mOhm and above land: a=3.60, b=5.60, c=11.20 mm",
    )


def write_power_semiconductor_footprints() -> None:
    """Generate drawing-derived power IC and MOSFET land patterns.

    The MOSFET symbols expose G/S/D as pins 1/2/3.  Their physical packages
    use repeated source and drain terminals, so these project footprints map
    every physical terminal onto the corresponding three schematic nets.
    """

    def fp_uuid(name: str, label: str) -> str:
        return str(
            _uuid.uuid5(
                _uuid.NAMESPACE_URL,
                f"radxa-cm5-procomm:power-semiconductor:{name}:{label}",
            )
        )

    def base_lines(
        name: str,
        body: tuple[float, float],
        courtyard: tuple[float, float],
        datasheet: str,
        description: str,
        reference_prefix: str,
    ) -> list[str]:
        body_x, body_y = body
        courtyard_x, courtyard_y = courtyard
        return [
            f'(footprint "{name}"',
            '  (version 20240108)',
            '  (generator "procomm_cm5_generator")',
            '  (layer "F.Cu")',
            f'  (descr "{description}")',
            '  (attr smd)',
            '  (duplicate_pad_numbers_are_jumpers no)',
            f'  (property "Reference" "{reference_prefix}**" (at 0 {-courtyard_y - 1.0:.2f} 0) (layer "F.SilkS") (uuid "{fp_uuid(name, "reference")}") (effects (font (size 1 1) (thickness 0.15))))',
            f'  (property "Value" "{name}" (at 0 {courtyard_y + 1.0:.2f} 0) (layer "F.Fab") (uuid "{fp_uuid(name, "value")}") (effects (font (size 1 1) (thickness 0.15))))',
            f'  (property "Datasheet" "{datasheet}" (at 0 0 0) (layer "F.Fab") (hide yes) (uuid "{fp_uuid(name, "datasheet")}") (effects (font (size 1.27 1.27))))',
            f'  (fp_rect (start {-body_x / 2:.3f} {-body_y / 2:.3f}) (end {body_x / 2:.3f} {body_y / 2:.3f}) (stroke (width 0.10) (type default)) (fill none) (layer "F.Fab") (uuid "{fp_uuid(name, "fab")}"))',
            f'  (fp_rect (start {-courtyard_x:.3f} {-courtyard_y:.3f}) (end {courtyard_x:.3f} {courtyard_y:.3f}) (stroke (width 0.05) (type default)) (fill none) (layer "F.CrtYd") (uuid "{fp_uuid(name, "courtyard")}"))',
            f'  (fp_line (start {-body_x / 2:.3f} {-body_y / 2:.3f}) (end {body_x / 2:.3f} {-body_y / 2:.3f}) (stroke (width 0.15) (type default)) (layer "F.SilkS") (uuid "{fp_uuid(name, "silk-top")}"))',
            f'  (fp_line (start {-body_x / 2:.3f} {body_y / 2:.3f}) (end {body_x / 2:.3f} {body_y / 2:.3f}) (stroke (width 0.15) (type default)) (layer "F.SilkS") (uuid "{fp_uuid(name, "silk-bottom")}"))',
            f'  (fp_circle (center {-body_x / 2 - 0.35:.3f} {-body_y / 2 - 0.35:.3f}) (end {-body_x / 2 - 0.20:.3f} {-body_y / 2 - 0.35:.3f}) (stroke (width 0.15) (type default)) (fill none) (layer "F.SilkS") (uuid "{fp_uuid(name, "pin-one")}"))',
        ]

    def add_rect_pad(
        lines: list[str],
        name: str,
        number: str,
        x: float,
        y: float,
        width: float,
        height: float,
        label: str,
        layers: str = '"F.Cu" "F.Paste" "F.Mask"',
        shape: str = "roundrect",
    ) -> None:
        lines.extend(
            [
                f'  (pad "{number}" smd {shape}',
                f'    (at {x:.3f} {y:.3f}) (size {width:.3f} {height:.3f})',
                f'    (layers {layers})',
            ]
        )
        if shape == "roundrect":
            lines.append('    (roundrect_rratio 0.10)')
        lines.extend([f'    (uuid "{fp_uuid(name, label)}")', '  )'])

    def add_segmented_paste(
        lines: list[str],
        name: str,
        x_centers: tuple[float, ...],
        y_centers: tuple[float, ...],
        size: tuple[float, float],
        label_prefix: str,
    ) -> None:
        for row, y in enumerate(y_centers):
            for column, x in enumerate(x_centers):
                add_rect_pad(
                    lines,
                    name,
                    "",
                    x,
                    y,
                    size[0],
                    size[1],
                    f"{label_prefix}-{row}-{column}",
                    layers='"F.Paste"',
                    shape="rect",
                )

    def finish(name: str, lines: list[str]) -> None:
        lines.extend(['  (embedded_fonts no)', ')'])
        (CM5_LOCAL_FOOTPRINTS / f"{name}.kicad_mod").write_text("\n".join(lines) + "\n")

    # TI RJR0014A VQFN-HR.  Coordinates are reconstructed directly from the
    # 4223976/H example board layout dimensions (June 2026).  Duplicate
    # rectangles form the four L-shaped corner lands without losing pad nets.
    name = "TI_RJR0014A"
    lines = base_lines(
        name,
        (4.0, 3.5),
        (2.45, 2.05),
        "https://www.ti.com/lit/ml/mpqf507g/mpqf507g.pdf",
        "TI RJR0014A 14-pin VQFN-HR; drawing 4223976/H manufacturer example land pattern",
        "U",
    )
    rjr_rects = (
        ("1", -1.675, -1.450, 0.35, 1.00, "pin-1-vertical"),
        ("1", -1.850, -1.150, 0.70, 0.40, "pin-1-horizontal"),
        ("2", -1.850, -0.525, 0.70, 0.25, "pin-2"),
        ("3", -1.850, 0.000, 0.70, 0.25, "pin-3"),
        ("4", -1.850, 0.525, 0.70, 0.25, "pin-4"),
        ("5", -1.675, 1.450, 0.35, 1.00, "pin-5-vertical"),
        ("5", -1.850, 1.150, 0.70, 0.40, "pin-5-horizontal"),
        ("6", -1.125, 1.600, 0.25, 0.70, "pin-6"),
        ("7", -0.625, 1.600, 0.25, 0.70, "pin-7"),
        ("8", 0.450, 1.450, 0.40, 1.00, "pin-8"),
        ("9", 1.600, 1.450, 0.40, 1.00, "pin-9-vertical"),
        ("9", 1.800, 1.150, 0.80, 0.40, "pin-9-horizontal"),
        ("10", 1.000, 0.000, 2.40, 0.40, "pin-10"),
        ("11", 1.600, -1.450, 0.40, 1.00, "pin-11-vertical"),
        ("11", 1.800, -1.150, 0.80, 0.40, "pin-11-horizontal"),
        ("12", 0.450, -1.450, 0.40, 1.00, "pin-12"),
        ("13", -0.625, -1.600, 0.25, 0.70, "pin-13"),
        ("14", -1.125, -1.600, 0.25, 0.70, "pin-14"),
    )
    for number, x, y, width, height, label in rjr_rects:
        add_rect_pad(lines, name, number, x, y, width, height, label)
    finish(name, lines)

    # TPS22990 DML0010A WSON.  The exposed thermal land is electrically VIN,
    # so it deliberately repeats schematic pad 3.
    name = "TI_DML0010A"
    lines = base_lines(
        name,
        (2.0, 3.0),
        (1.60, 1.80),
        "https://www.ti.com/lit/ds/symlink/tps22990.pdf",
        "TI DML0010A 10-pin WSON; drawing 4222524/A manufacturer land pattern; exposed pad mapped to VIN pin 3",
        "U",
    )
    for pin, y in enumerate((-1.0, -0.5, 0.0, 0.5, 1.0), start=1):
        add_rect_pad(lines, name, str(pin), -1.30, y, 0.45, 0.24, f"pin-{pin}")
    for pin, y in zip((6, 7, 8, 9, 10), (1.0, 0.5, 0.0, -0.5, -1.0)):
        height = 0.28 if pin >= 8 else 0.24
        add_rect_pad(lines, name, str(pin), 1.30, y, 0.45, height, f"pin-{pin}")
    add_rect_pad(lines, name, "3", 0.02, 0.0, 1.10, 1.95, "vin-exposed", layers='"F.Cu" "F.Mask"', shape="rect")
    add_segmented_paste(lines, name, (-0.255, 0.295), (-0.4875, 0.4875), (0.45, 0.75), "vin-paste")
    finish(name, lines)

    # TI DNK0008A VSON-CLIP used by both CSD power MOSFETs.  The copper and
    # paste geometry follows KiCad's TI CSD18531Q5A drawing-backed footprint;
    # only the pad numbers are remapped to the G/S/D schematic convention.
    name = "TI_DNK0008A_GSD"
    lines = base_lines(
        name,
        (5.0, 6.0),
        (3.40, 3.40),
        "https://www.ti.com/lit/ds/symlink/csd18532q5b.pdf",
        "TI DNK0008A VSON-CLIP 5 x 6 mm; physical source/gate/drain terminals remapped to schematic G/S/D pins 1/2/3",
        "Q",
    )
    for y in (-1.905, -0.635, 0.635):
        add_rect_pad(lines, name, "2", -2.80, y, 0.70, 0.70, f"source-{y}")
    add_rect_pad(lines, name, "1", -2.80, 1.905, 0.70, 0.70, "gate")
    for y in (-1.905, -0.635, 0.635, 1.905):
        add_rect_pad(lines, name, "3", 2.80, y, 0.70, 0.70, f"drain-{y}", shape="rect")
    add_rect_pad(lines, name, "3", 0.33, 0.0, 4.35, 4.51, "drain-exposed", layers='"F.Cu" "F.Mask"', shape="rect")
    add_segmented_paste(lines, name, (-0.4525, 1.1125), (-1.095, 1.095), (1.585, 1.57), "drain-paste")
    finish(name, lines)

    # onsemi DFN5 5x6 (SO-8FL), case 488AA.  Physical pins 1-3 are source,
    # pin 4 is gate, and pin 5 plus the large exposed area are drain.
    name = "onsemi_DFN5_5x6_488AA_GSD"
    lines = base_lines(
        name,
        (5.15, 6.15),
        (2.65, 3.65),
        "https://www.onsemi.com/download/data-sheet/pdf/nvmfs6b25nl-d.pdf",
        "onsemi case 488AA DFN5 5 x 6 mm; recommended SO-8FL land remapped to schematic G/S/D pins 1/2/3",
        "Q",
    )
    for physical_pin, x in enumerate((-1.905, -0.635, 0.635, 1.905), start=1):
        number = "2" if physical_pin <= 3 else "1"
        add_rect_pad(lines, name, number, x, 2.765, 0.75, 1.00, f"terminal-{physical_pin}")
    add_rect_pad(lines, name, "3", 0.0, -0.665, 4.56, 3.20, "drain-main", layers='"F.Cu" "F.Mask"', shape="rect")
    for x in (-2.0425, 2.0425):
        add_rect_pad(lines, name, "3", x, 1.600, 0.475, 1.33, f"drain-side-{x}")
    add_segmented_paste(lines, name, (-1.45, 0.0, 1.45), (-1.20, -0.10), (1.20, 0.85), "drain-paste")
    finish(name, lines)

    # onsemi DFNW8 5.2x6.3, case 507AU.  Physical pins 1-3 are source, pin 4
    # is gate, pins 5-8 and the central thermal land are drain.
    name = "onsemi_DFNW8_5p2x6p3_507AU_GSD"
    lines = base_lines(
        name,
        (5.20, 6.30),
        (2.85, 3.75),
        "https://www.onsemi.com/download/data-sheet/pdf/fdws86068-f085-d.pdf",
        "onsemi case 507AU DFNW8 5.2 x 6.3 mm; manufacturer land remapped to schematic G/S/D pins 1/2/3",
        "Q",
    )
    for physical_pin, x in enumerate((-1.905, -0.635, 0.635, 1.905), start=1):
        number = "2" if physical_pin <= 3 else "1"
        add_rect_pad(lines, name, number, x, 2.745, 0.61, 1.42, f"terminal-{physical_pin}")
    add_rect_pad(lines, name, "3", 0.0, 0.0, 4.42, 3.75, "drain-main", layers='"F.Cu" "F.Mask"', shape="rect")
    for physical_pin, x in zip((8, 7, 6, 5), (-1.905, -0.635, 0.635, 1.905)):
        add_rect_pad(lines, name, "3", x, -2.745, 0.66, 1.42, f"drain-terminal-{physical_pin}")
    add_segmented_paste(lines, name, (-1.40, 0.0, 1.40), (-1.15, 0.0, 1.15), (1.05, 0.85), "drain-paste")
    finish(name, lines)


def write_molex_0679101002_footprint() -> None:
    """Generate the Molex Mini PCIe land pattern from SD-67910-001 C2."""

    def fp_uuid(label: str) -> str:
        return str(
            _uuid.uuid5(
                _uuid.NAMESPACE_URL,
                f"radxa-cm5-procomm:molex-0679101002:{label}",
            )
        )

    lines = [
        '(footprint "Molex_0679101002_Mini_PCIe"',
        '  (version 20240108)',
        '  (generator "pcbnew")',
        '  (layer "F.Cu")',
        '  (descr "Molex 0679101002 Mini PCI Express socket, 52 contacts, 0.8 mm pitch, 4.0 mm height; manufacturer recommended PCB layout from production drawing SD-67910-001 revision C2")',
        '  (tags "Molex 0679101002 67910-1002 Mini PCIe socket")',
        '  (property "Reference" "J**"',
        '    (at 12.5 -7.0 0)',
        '    (layer "F.SilkS")',
        f'    (uuid "{fp_uuid("reference")}")',
        '    (effects (font (size 1 1) (thickness 0.15)))',
        '  )',
        '  (property "Value" "Molex_0679101002_Mini_PCIe"',
        '    (at 12.5 7.0 0)',
        '    (layer "F.Fab")',
        f'    (uuid "{fp_uuid("value")}")',
        '    (effects (font (size 1 1) (thickness 0.15)))',
        '  )',
        '  (property "Datasheet" "https://www.molex.com/content/dam/molex/molex-dot-com/products/automated/en-us/salesdrawingpdf/679/67910/679100002_sd.pdf"',
        '    (at 0 0 0) (layer "F.Fab") (hide yes)',
        f'    (uuid "{fp_uuid("datasheet")}")',
        '    (effects (font (size 1.27 1.27)))',
        '  )',
        '  (property "Description" "Molex 0679101002 right-angle 52-position Mini PCIe card socket"',
        '    (at 0 0 0) (layer "F.Fab") (hide yes)',
        f'    (uuid "{fp_uuid("description")}")',
        '    (effects (font (size 1.27 1.27)))',
        '  )',
        '  (attr smd)',
        '  (duplicate_pad_numbers_are_jumpers no)',
    ]

    def add_line(
        start: tuple[float, float],
        end: tuple[float, float],
        layer: str,
        width: float,
        label: str,
    ) -> None:
        lines.extend(
            [
                '  (fp_line',
                f'    (start {start[0]:.2f} {start[1]:.2f})',
                f'    (end {end[0]:.2f} {end[1]:.2f})',
                f'    (stroke (width {width:.2f}) (type default))',
                f'    (layer "{layer}")',
                f'    (uuid "{fp_uuid(label)}")',
                '  )',
            ]
        )

    # The left locator hole is datum E. The connector drawing specifies a
    # 25.00 mm locator-hole pitch and a 30.00 mm maximum body width.
    for layer, width, suffix in (("F.Fab", 0.10, "fab"), ("F.CrtYd", 0.05, "courtyard")):
        x0, y0, x1, y1 = (
            (-2.50, -5.10, 27.50, 5.10)
            if layer == "F.Fab"
            else (-3.10, -5.60, 28.10, 5.60)
        )
        add_line((x0, y0), (x1, y0), layer, width, f"{suffix}-top")
        add_line((x1, y0), (x1, y1), layer, width, f"{suffix}-right")
        add_line((x1, y1), (x0, y1), layer, width, f"{suffix}-bottom")
        add_line((x0, y1), (x0, y0), layer, width, f"{suffix}-left")

    add_line((-2.50, -5.35), (27.50, -5.35), "F.SilkS", 0.15, "silk-top")
    add_line((-2.50, -5.35), (-2.50, -4.95), "F.SilkS", 0.15, "silk-left")
    add_line((27.50, -5.35), (27.50, -4.95), "F.SilkS", 0.15, "silk-right")
    add_line((-0.80, 5.35), (0.80, 5.35), "F.SilkS", 0.25, "pin-one-marker")

    lines.extend(
        [
            '  (fp_text user "DATUM E"',
            '    (at 0 0 0) (layer "F.Fab")',
            f'    (uuid "{fp_uuid("datum-e-text")}")',
            '    (effects (font (size 0.6 0.6) (thickness 0.10)))',
            '  )',
            '  (fp_text user "SD-67910-001 C2"',
            '    (at 12.5 0 0) (layer "F.Fab")',
            f'    (uuid "{fp_uuid("drawing-text")}")',
            '    (effects (font (size 0.7 0.7) (thickness 0.10)))',
            '  )',
            '  (pad "" np_thru_hole circle',
            '    (at 0 0) (size 1.60 1.60) (drill 1.60)',
            '    (layers "*.Cu" "*.Mask")',
            f'    (uuid "{fp_uuid("locator-e")}")',
            '  )',
            '  (pad "" np_thru_hole circle',
            '    (at 25.00 0) (size 1.10 1.10) (drill 1.10)',
            '    (layers "*.Cu" "*.Mask")',
            f'    (uuid "{fp_uuid("locator-d")}")',
            '  )',
            '  (pad "" smd rect',
            '    (at -2.15 3.50) (size 1.60 3.20)',
            '    (layers "F.Cu" "F.Paste" "F.Mask")',
            f'    (uuid "{fp_uuid("hold-down-left")}")',
            '  )',
            '  (pad "" smd rect',
            '    (at 27.15 3.50) (size 1.60 3.20)',
            '    (layers "F.Cu" "F.Paste" "F.Mask")',
            f'    (uuid "{fp_uuid("hold-down-right")}")',
            '  )',
        ]
    )

    # The contact rows are staggered 0.40 mm. Pins 1-16 occupy the 5.60 mm
    # group; pins 17-52 occupy the 13.60 mm group across the card key.
    contact_positions: list[tuple[int, float, float]] = []
    contact_positions.extend((1 + 2 * index, 0.80 * index, 4.10) for index in range(8))
    contact_positions.extend((2 + 2 * index, 0.40 + 0.80 * index, -4.10) for index in range(8))
    contact_positions.extend((17 + 2 * index, 10.30 + 0.80 * index, 4.10) for index in range(18))
    contact_positions.extend((18 + 2 * index, 10.70 + 0.80 * index, -4.10) for index in range(18))
    for pin, x, y in sorted(contact_positions):
        shape = "roundrect" if pin == 1 else "rect"
        lines.extend(
            [
                f'  (pad "{pin}" smd {shape}',
                f'    (at {x:.2f} {y:.2f}) (size 0.60 2.00)',
                '    (layers "F.Cu" "F.Paste" "F.Mask")',
            ]
        )
        if pin == 1:
            lines.append('    (roundrect_rratio 0.25)')
        lines.extend([f'    (uuid "{fp_uuid(f"pad-{pin}")}")', '  )'])

    lines.append(')')
    output = CM5_LOCAL_FOOTPRINTS / "Molex_0679101002_Mini_PCIe.kicad_mod"
    output.write_text("\n".join(lines) + "\n")


def create_project(folder: Path, name: str) -> None:
    folder.mkdir(parents=True, exist_ok=True)
    project_path = folder / f"{name}.kicad_pro"
    shutil.copyfile(KICAD_TEMPLATE, project_path)
    data = json.loads(project_path.read_text())
    data["meta"]["filename"] = project_path.name
    project_path.write_text(json.dumps(data, indent=2) + "\n")


def pin_xy(component, pin: int | str) -> tuple[float, float]:
    point = component.get_pin_position(str(pin))
    if point is None:
        raise RuntimeError(f"{component.reference} has no pin {pin}")
    return (point.x, 2.0 * component.position.y - point.y)


def add_symbol(
    schematic,
    lib_id: str,
    reference: str,
    value: str,
    position: tuple[float, float],
    footprint: str = "",
    manufacturer: str = "",
    mpn: str = "",
    rotation: int = 0,
):
    if not footprint and mpn:
        footprint = MPN_FOOTPRINTS.get(mpn, "")
    component = schematic.components.add(
        lib_id,
        reference,
        value,
        position=position,
        footprint=footprint or None,
        rotation=rotation,
    )
    if manufacturer:
        component.add_property("Manufacturer", manufacturer, hidden=True)
    if mpn:
        component.add_property("MPN", mpn, hidden=True)
    return component


def label_pin(schematic, component, pin: int, net_name: str, two_row: bool = False) -> None:
    justify = "right" if (not two_row or pin % 2 == 1) else "left"
    add_global_net_label(schematic, net_name, pin_xy(component, pin), 0.95, justify)


def add_global_net_label(
    schematic,
    net_name: str,
    position: tuple[float, float],
    size: float,
    justify: str,
) -> None:
    """Add a project-wide net label so nested board sheets form one netlist."""
    schematic.labels.add(net_name, position, size=size, justify_h=justify)


def save_generated_schematic(schematic, output: Path) -> None:
    """Save and promote all generated net labels to project-wide global labels.

    kicad-sch-api 0.4 accepts global labels but its KiCad 9 writer currently
    omits them. The local-label serializer is stable, so promote those emitted
    records structurally after save. All labels in these generated circuit
    sheets are electrical net labels; free annotations use text objects.
    """
    schematic.save(output)
    text = output.read_text(encoding="utf-8")
    text, count = re.subn(
        r'(?m)^(\t)\(label ("[^"\n]+")$',
        r'\1(global_label \2\n\1\t(shape bidirectional)',
        text,
    )
    if not count:
        raise RuntimeError(f"{output.name} emitted no electrical labels to globalize")
    output.write_text(text, encoding="utf-8")


def label_pin_auto(schematic, component, pin: int | str, net_name: str, size: float = 0.72) -> None:
    position = pin_xy(component, pin)
    justify = "right" if position[0] < component.position.x else "left"
    add_global_net_label(schematic, net_name, position, size, justify)


def label_two_pin_device(schematic, component, net_1: str, net_2: str) -> None:
    label_pin_auto(schematic, component, 1, net_1, 0.7)
    label_pin_auto(schematic, component, 2, net_2, 0.7)


def label_pin_with_stub(
    schematic,
    component,
    pin: int | str,
    net_name: str,
    offsets: tuple[tuple[float, float], ...],
    justify: str,
) -> None:
    """Route a short orthogonal stub so a net label clears the symbol fields."""
    start = pin_xy(component, pin)
    points = [start]
    for offset_x, offset_y in offsets:
        points.append((start[0] + offset_x, start[1] + offset_y))
    for segment_start, segment_end in zip(points, points[1:]):
        schematic.wires.add(start=segment_start, end=segment_end)
    add_global_net_label(schematic, net_name, points[-1], 0.85, justify)


def add_connector(
    schematic,
    lib_id: str,
    reference: str,
    value: str,
    position: tuple[float, float],
    pin_map: dict[int, str],
    footprint: str = "",
    manufacturer: str = "",
    mpn: str = "",
    two_row: bool = False,
):
    component = add_symbol(
        schematic,
        lib_id,
        reference,
        value,
        position,
        footprint,
        manufacturer,
        mpn,
    )
    for pin, net_name in pin_map.items():
        label_pin(schematic, component, pin, net_name, two_row=two_row)
    return component


def heading(schematic, text: str, position: tuple[float, float], size: float = 2.0) -> None:
    schematic.texts.add(text, position, size=size, bold=True)


def note(schematic, text: str, position: tuple[float, float], size: float = 1.0) -> None:
    schematic.texts.add(text, position, size=size)


def add_board_child_sheets(
    schematic,
    project_name: str,
    sheets: tuple[tuple[str, str], ...],
    origin: tuple[float, float],
) -> None:
    """Nest detailed circuit pages under the physical board root schematic."""
    origin_x, origin_y = origin
    for index, (sheet_name, filename) in enumerate(sheets, start=1):
        schematic.add_sheet(
            name=sheet_name,
            filename=filename,
            position=(origin_x, origin_y + (index - 1) * 48),
            size=(165, 30),
            stroke_width=0.3,
            project_name=project_name,
            page_number=str(index + 1),
        )


def add_audio8_passive(
    schematic,
    lib_id: str,
    reference: str,
    value: str,
    position: tuple[float, float],
    rotation: int = 0,
):
    """Add a production-identified passive used by the detailed audio sheets."""
    try:
        manufacturer, mpn = AUDIO8_PASSIVE_PARTS[value]
    except KeyError as error:
        raise RuntimeError(f"No controlled AUDIO-8X8 passive for {value!r}") from error
    return add_symbol(
        schematic,
        lib_id,
        reference,
        value,
        position,
        manufacturer=manufacturer,
        mpn=mpn,
        rotation=rotation,
    )


def build_cm5_core_sheet(pinout: dict[str, list[tuple[int, str]]]) -> Path:
    """Capture all three CM5 connectors and the 76 owned allocation contacts."""
    folder = ROOT / "CM5-CARRIER"
    name = "CM5-Core-Allocated"
    create_project(folder, name)
    schematic = create_schematic(name)
    schematic.set_paper_size("A2")
    schematic.set_title_block(
        title="Radxa CM5 ProComm Carrier - Exact CM5 Connector Allocation",
        date="2026-08-13",
        rev="A1",
        company="ProComm",
        comments={
            1: "Three 100-contact Hirose DF40C mates; all physical pin numbers follow CM5 V2.21",
            2: "76 contacts are owned: 74 connected and 2 assigned no-connect; all other functions are no-connect",
            3: "All connector grounds join the carrier ground plane; allocated GPIO domain is 3.3 V",
            4: "DETAILED CAPTURE BASELINE - verify CM5 voltage limits before release",
        },
    )

    heading(schematic, "1. EXACT RADXA CM5 MATING CONNECTORS", (45, 18), 1.8)
    allocations = allocation_map()
    if len(CM5_ALLOCATIONS) != 76 or len(allocations) != 76:
        raise RuntimeError(
            "CM5 allocation contract must contain exactly 76 unique physical contacts"
        )
    if not CM5_ASSIGNED_NC.issubset(allocations) or len(CM5_ASSIGNED_NC) != 2:
        raise RuntimeError("CM5 assigned-no-connect contract must contain two allocated contacts")
    physical_contacts = {
        (connector, pin)
        for connector, rows in pinout.items()
        for pin, _signal in rows
    }
    missing_contacts = sorted(set(allocations) - physical_contacts)
    if missing_contacts:
        raise RuntimeError(
            f"CM5 allocations reference contacts absent from the V2.21 workbook: {missing_contacts}"
        )
    if len(physical_contacts) != 300:
        raise RuntimeError(
            f"CM5 source must contain 300 unique physical contacts, found {len(physical_contacts)}"
        )
    placements = (
        ("U13-A", "J501", 105.0),
        ("U13-B", "J502", 295.0),
        ("J1", "J503", 485.0),
    )
    for connector, reference, x in placements:
        symbol_name = f"Radxa_CM5_{connector.replace('-', '_')}"
        component = add_symbol(
            schematic,
            f"CM5Carrier:{symbol_name}",
            reference,
            f"RADXA CM5 {connector} MATE",
            (x, 105),
            CM5_DF40_FOOTPRINT,
            "Hirose",
            "DF40C-100DS-0.4V(51)",
        )
        for pin, signal in pinout[connector]:
            position = pin_xy(component, pin)
            if (connector, pin) in CM5_ASSIGNED_NC:
                schematic.no_connects.add(position)
            elif (connector, pin) in allocations:
                net_name = allocations[(connector, pin)][1]
                justify = "right" if position[0] < x else "left"
                add_global_net_label(schematic, net_name, position, 0.68, justify)
            elif signal == "GND":
                justify = "right" if position[0] < x else "left"
                add_global_net_label(schematic, "GND", position, 0.62, justify)
            else:
                schematic.no_connects.add(position)
        note(schematic, f"{connector}: 100 physical contacts", (x - 28, 174), 0.82)

    heading(schematic, "2. INTERNAL POWER / STARTUP / RECOVERY SERVICE", (45, 202), 1.6)
    add_connector(
        schematic,
        "Connector_Generic:Conn_01x06",
        "J510",
        "CM5_POWER_SERVICE",
        (90, 235),
        {
            1: "LOGIC_3V3",
            2: "GND",
            3: "CM5_RESET_N",
            4: "CM5_BOOT",
            5: "CM5_PWRON_N",
            6: "CM5_RECOVERY_KEY",
        },
        JST_GH_6,
        "JST",
        "BM06B-GHS-TBT(LF)(SN)",
    )
    note(schematic, "Internal keyed service header; never exposed on the top panel.", (55, 263), 0.85)

    add_connector(
        schematic,
        "Connector_Generic:Conn_01x04",
        "J511",
        "DEBUG_UART_3V3",
        (215, 232),
        {1: "LOGIC_3V3", 2: "GND", 3: "DBG_UART_TX", 4: "DBG_UART_RX"},
        JST_GH_4,
        "JST",
        "BM04B-GHS-TBT(LF)(SN)",
    )
    note(schematic, "3.3 V UART only; keyed factory/debug cable.", (190, 255), 0.85)

    add_connector(
        schematic,
        "Connector_Generic:Conn_01x05",
        "J512",
        "INTERNAL_USB_RECOVERY",
        (330, 233),
        {
            1: "REC_USB_VBUS_DET",
            2: "REC_USB_DM",
            3: "REC_USB_DP",
            4: "REC_USB_ID",
            5: "GND",
        },
        JST_GH_5,
        "JST",
        "BM05B-GHS-TBT(LF)(SN)",
    )
    note(schematic, "Internal keyed recovery harness; service fixture owns the protected USB-C receptacle and CC network.", (285, 260), 0.85)

    add_connector(
        schematic,
        "Connector_Generic:Conn_02x04_Odd_Even",
        "J513",
        "CM5_POWER_ENTRY",
        (470, 232),
        {
            1: "SYS_4V0",
            2: "SYS_4V0",
            3: "SYS_4V0",
            4: "SYS_4V0",
            5: "GND",
            6: "GND",
            7: "GND",
            8: "GND",
        },
        MF_2X4,
        "Molex",
        "43045-0812",
        two_row=True,
    )
    note(schematic, "Kelvin-sense SYS_4V0 at J501; 4.0 V follows the Radxa CM5 carrier design note.", (420, 260), 0.85)

    heading(schematic, "3. OWNERSHIP AUDIT", (45, 295), 1.6)
    audit_lines = (
        "76 / 76 contacts owned; 74 connected, 2 assigned NC, 0 duplicate physical-pin claims.",
        "WAN1 uses native MDI. PCIe Gen2 x1 feeds the packet switch.",
        "WWAN owns USB30_2 + USB20_HOST0; touch owns USB20_HOST1.",
        "HDMI_TX0, I2S0 TDM, I2S1 headset, I2C7, I2C3 and UART2 are reserved exactly once.",
        "Pin 145 reaches HDMI pin 14; pin 147 is assigned NC because HEAC/ARC is unused. U13-B pin 106 receives IO_5V0.",
        "Unused alternate-function contacts carry explicit no-connect markers on this A1 baseline.",
    )
    for index, text in enumerate(audit_lines):
        note(schematic, text, (45, 312 + index * 9), 0.9)

    output = folder / f"{name}.kicad_sch"
    save_generated_schematic(schematic, output)
    return output


def build_network_pcie_sheet() -> Path:
    """Capture PCIe, four protected 1 GbE ports, and the Wi-Fi M.2 interface."""
    folder = ROOT / "CM5-CARRIER"
    name = "Network-PCIe"
    create_project(folder, name)
    schematic = create_schematic(name)
    schematic.set_paper_size("A1")
    schematic.set_title_block(
        title="Radxa CM5 ProComm Carrier - PCIe / Ethernet / Wi-Fi",
        date="2026-08-14",
        rev="A1",
        company="ProComm",
        comments={
            1: "PI7C9X2G608GP in 606 mode: x1 upstream plus five x1 downstream ports",
            2: "Native WAN1 plus three LAN7430 endpoints use exact Wurth 74991114412 MagJacks",
            3: "Exact package balls/pins from Diodes DS40210 Rev 8 and Microchip DS00002631G",
            4: "Port 5 is disabled and unrouted; generic cable headers are prohibited on PCIe Gen2 lanes",
        },
    )

    def add_network_passive(
        lib_id: str,
        reference: str,
        value: str,
        position: tuple[float, float],
    ):
        try:
            manufacturer, mpn = NETWORK_PASSIVE_PARTS[value]
        except KeyError as error:
            raise RuntimeError(
                f"Network-PCIe {reference} has no locked production passive for {value}"
            ) from error
        return add_symbol(
            schematic,
            lib_id,
            reference,
            value,
            position,
            manufacturer=manufacturer,
            mpn=mpn,
        )

    heading(schematic, "1. PI7C9X2G608GP PCIe GEN2 SWITCH - 606 MODE", (35, 18), 1.8)
    switch_group = schematic.components.add(
        "CM5Carrier:PI7C9X2G608GP",
        "U601",
        "PI7C9X2G608GP",
        position=(104.14, 105.41),
        footprint="Package_BGA:BGA-196_15x15mm_Layout14x14_P1.0mm",
        add_all_units=True,
        unit_spacing=150,
    )
    switch_group.place_unit(1, (104.14, 105.41))
    switch_group.place_unit(2, (104.14, 284.48))
    switch_group.place_unit(3, (104.14, 455.93))
    switch_units = {unit: switch_group.get_unit(unit) for unit in (1, 2, 3)}
    switch_units[1].add_property("Manufacturer", "Diodes Incorporated", hidden=True)
    switch_units[1].add_property("MPN", "PI7C9X2G608GPCNJEX", hidden=True)

    switch_nets = {
        # Upstream lane 0 and the clock-buffer input.
        "C12": "PCIE_UP_PERST_N",
        "G13": "PCIE_UP_REFCLK_P",
        "G14": "PCIE_UP_REFCLK_N",
        "P3": "PCIE_UP_SW_RX_P",
        "N3": "PCIE_UP_SW_RX_N",
        "P4": "PCIE_UP_SW_TX_P",
        "N4": "PCIE_UP_SW_TX_N",
        # Clock-buffer loopbacks for the switch PHY banks.
        "D13": "SW_REFCLK1_RAW_P",
        "D14": "SW_REFCLK1_RAW_N",
        "N8": "SW_REFCLK0_IN_P",
        "P8": "SW_REFCLK0_IN_N",
        "E13": "SW_REFCLK2_RAW_P",
        "E14": "SW_REFCLK2_RAW_N",
        "B7": "SW_REFCLK1_IN_P",
        "A7": "SW_REFCLK1_IN_N",
        # Port 1 / lane 4: WAN2.
        "A11": "WAN2_SW_TX_P",
        "B11": "WAN2_SW_TX_N",
        "A12": "WAN2_PCIE_TX_P",
        "B12": "WAN2_PCIE_TX_N",
        "F13": "WAN2_PCIE_REFCLK_P",
        "F14": "WAN2_PCIE_REFCLK_N",
        "M2": "WAN2_PCIE_PERST_N",
        "F11": "WAN2_PCIE_CLKREQ_N",
        # Port 2 / lane 5: LAN1.
        "A9": "LAN1_SW_TX_P",
        "B9": "LAN1_SW_TX_N",
        "A10": "LAN1_PCIE_TX_P",
        "B10": "LAN1_PCIE_TX_N",
        "H13": "LAN1_PCIE_REFCLK_P",
        "H14": "LAN1_PCIE_REFCLK_N",
        "N2": "LAN1_PCIE_PERST_N",
        "D12": "LAN1_PCIE_CLKREQ_N",
        # Port 3 / lane 6: LAN2.
        "A6": "LAN2_SW_TX_P",
        "B6": "LAN2_SW_TX_N",
        "A5": "LAN2_PCIE_TX_P",
        "B5": "LAN2_PCIE_TX_N",
        "J13": "LAN2_PCIE_REFCLK_P",
        "J14": "LAN2_PCIE_REFCLK_N",
        "P2": "LAN2_PCIE_PERST_N",
        "C13": "LAN2_PCIE_CLKREQ_N",
        # Port 4 / lane 7: Wi-Fi.
        "A4": "WIFI_SW_TX_P",
        "B4": "WIFI_SW_TX_N",
        "A3": "WIFI_PCIE_TX_P",
        "B3": "WIFI_PCIE_TX_N",
        "K13": "WIFI_PCIE_REFCLK_P",
        "K14": "WIFI_PCIE_REFCLK_N",
        "M3": "WIFI_PCIE_PERST_N",
        "A2": "WIFI_PCIE_CLKREQ_N",
        # Configuration and management.
        "D2": "PCIE_MODE_606_HIGH",
        "F12": "SYS_I2C7_SDA",
        "J11": "SYS_I2C7_SCL",
        "A8": "SW_REXT1",
        "B8": "GND",
        "P7": "SW_REXT0",
        "N7": "GND",
        "E12": "SW_IREF",
    }
    power_names = {"VDDC", "VDDR", "CVDDR", "AVDD", "AVDDH", "AGND", "DGND"}
    for row_name, names in PI7_BALL_ROWS.items():
        for column, signal in enumerate(names, start=1):
            ball = f"{row_name}{column}"
            if signal in power_names:
                unit = 3
            elif signal.startswith(("PER", "PET", "REFCLK", "CLKREQ", "DWNRST")) or signal == "PERST_L":
                unit = 1
            else:
                unit = 2
            component = switch_units[unit]
            if ball in switch_nets:
                label_pin_auto(schematic, component, ball, switch_nets[ball], 0.64)
            elif signal in {"AGND", "DGND"}:
                label_pin_auto(schematic, component, ball, "GND", 0.6)
            elif signal in {"VDDC", "AVDD"}:
                label_pin_auto(schematic, component, ball, "PCIE_1V0", 0.6)
            elif signal in {"VDDR", "CVDDR", "AVDDH"}:
                label_pin_auto(schematic, component, ball, "NET_3V3", 0.6)
            else:
                schematic.no_connects.add(pin_xy(component, ball))

    note(schematic, "U601A: exact PCIe lanes, reset and distributed reference clocks", (35, 166), 0.82)
    note(schematic, "U601B: 606-mode strap, I2C and calibration; other straps use documented internal defaults", (35, 346), 0.82)
    note(schematic, "U601C: every 1.0 V / 3.3 V / ground ball shown; one 100 nF capacitor is assigned to every supply ball", (35, 535), 0.82)

    support_parts = (
        ("R601", "Device:R", "4.7k", "PCIE_MODE_606_HIGH", "NET_3V3", (230, 250)),
        ("R602", "Device:R", "1.43k 1%", "SW_REXT1", "GND", (230, 270)),
        ("R603", "Device:R", "1.43k 1%", "SW_REXT0", "GND", (230, 290)),
        ("R604", "Device:R", "475R 1%", "SW_IREF", "GND", (230, 310)),
        ("R605", "Device:R", "2.2k", "SYS_I2C7_SDA", "LOGIC_3V3", (230, 330)),
        ("R606", "Device:R", "2.2k", "SYS_I2C7_SCL", "LOGIC_3V3", (230, 350)),
        ("R607", "Device:R", "0R", "PCIE_UP_CLKREQ_N", "GND", (230, 370)),
        ("C601", "Device:C", "100nF", "SW_REFCLK1_RAW_P", "SW_REFCLK0_IN_P", (230, 410)),
        ("C602", "Device:C", "100nF", "SW_REFCLK1_RAW_N", "SW_REFCLK0_IN_N", (230, 430)),
        ("C603", "Device:C", "100nF", "SW_REFCLK2_RAW_P", "SW_REFCLK1_IN_P", (230, 450)),
        ("C604", "Device:C", "100nF", "SW_REFCLK2_RAW_N", "SW_REFCLK1_IN_N", (230, 470)),
        ("C605", "Device:C", "100nF", "PCIE_UP_TX_P_CM5", "PCIE_UP_SW_RX_P", (285, 410)),
        ("C606", "Device:C", "100nF", "PCIE_UP_TX_N_CM5", "PCIE_UP_SW_RX_N", (285, 430)),
        ("C607", "Device:C", "100nF", "PCIE_UP_SW_TX_P", "PCIE_UP_RX_P_CM5", (285, 450)),
        ("C608", "Device:C", "100nF", "PCIE_UP_SW_TX_N", "PCIE_UP_RX_N_CM5", (285, 470)),
        ("C609", "Device:C", "100nF", "WIFI_SW_TX_P", "WIFI_PCIE_RX_P", (340, 450)),
        ("C610", "Device:C", "100nF", "WIFI_SW_TX_N", "WIFI_PCIE_RX_N", (340, 470)),
    )
    for reference, lib_id, value, net_1, net_2, position in support_parts:
        part = add_network_passive(lib_id, reference, value, position)
        label_two_pin_device(schematic, part, net_1, net_2)
    note(schematic, "GPIO[1:0] = 01 selects 606 mode. GPIO1 uses its internal pulldown; GPIO0 is pulled high.", (200, 390), 0.82)
    heading(schematic, "HARDWARE-QUALIFIED UPSTREAM RESET", (185, 492), 1.35)
    reset_gate = add_symbol(
        schematic,
        "74xGxx:74LVC1G11",
        "U606",
        "SN74LVC1G11DBVR",
        (235, 525),
        manufacturer="Texas Instruments",
        mpn="SN74LVC1G11DBVR",
    )
    for pin, net_name in {
        1: "PCIE_UP_PERST_CMD_N", 2: "GND", 3: "NET_3V3_PG",
        4: "PCIE_UP_PERST_N", 5: "LOGIC_3V3", 6: "PCIE_1V0_PG",
    }.items():
        label_pin_auto(schematic, reset_gate, pin, net_name, 0.52)
    reset_pulldown = add_network_passive("Device:R", "R608", "100k", (290, 515))
    label_two_pin_device(schematic, reset_pulldown, "PCIE_UP_PERST_N", "GND")
    reset_bypass = add_network_passive("Device:C", "C611", "100nF", (290, 540))
    label_two_pin_device(schematic, reset_bypass, "LOGIC_3V3", "GND")
    note(schematic, "U606 asserts switch PERST# unless the CM5 request and both PCIe rail power-good signals are high.", (185, 552), 0.72)
    note(schematic, "CM5 firmware/root-complex timing must keep the request low until REFCLK and rails have been stable for at least 100 ms.", (185, 560), 0.72)

    switch_supply_rails: list[str] = []
    for names in PI7_BALL_ROWS.values():
        for signal in names:
            if signal in {"VDDC", "AVDD"}:
                switch_supply_rails.append("PCIE_1V0")
            elif signal in {"VDDR", "CVDDR", "AVDDH"}:
                switch_supply_rails.append("NET_3V3")
    for index, rail in enumerate(switch_supply_rails, start=1):
        x = 390 + ((index - 1) % 10) * 28
        y = 430 + ((index - 1) // 10) * 25
        capacitor = add_network_passive("Device:C", f"C{6600 + index}", "100nF", (x, y))
        label_two_pin_device(schematic, capacitor, rail, "GND")
    for reference, rail, x in (
        ("C6630", "PCIE_1V0", 665),
        ("C6631", "NET_3V3", 705),
    ):
        capacitor = add_network_passive("Device:C", reference, "22uF 10V X7R", (x, 505))
        label_two_pin_device(schematic, capacitor, rail, "GND")
    note(schematic, "U601 rail bypass: one 100 nF per supply ball plus 22 uF bulk on 1.0 V and 3.3 V.", (390, 525), 0.78)

    heading(schematic, "2. THREE LAN7430 PCIe GIGABIT ENDPOINTS", (295, 18), 1.8)

    def place_ethernet_esd(
        reference: str,
        nets: tuple[str, str, str, str],
        position: tuple[float, float],
    ) -> None:
        esd = add_symbol(
            schematic,
            "Power_Protection:TPD4E05U06DQA",
            reference,
            "TPD4E05U06DQA",
            position,
            "Package_SON:USON-10_2.5x1.0mm_P0.5mm",
            "Texas Instruments",
            "TPD4E05U06DQAR",
        )
        for pin, net_name in zip((1, 2, 4, 5), nets):
            label_pin_auto(schematic, esd, pin, net_name, 0.54)
        for pin in (3, 8):
            label_pin_auto(schematic, esd, pin, "GND", 0.54)
        for pin in (6, 7, 9, 10):
            schematic.no_connects.add(pin_xy(esd, pin))

    def place_magjack(
        reference: str,
        prefix: str,
        x: float,
        y: float,
        mdi_nets: tuple[str, str, str, str, str, str, str, str],
        led_sinks: tuple[str, str],
        esd_references: tuple[str, str],
        resistor_references: tuple[str, str],
    ) -> None:
        jack = add_symbol(
            schematic,
            "WurthRJ45:74991114412",
            reference,
            "74991114412",
            (x, y),
            "CM5Carrier:T_Wurth_WE-RJ45LAN_74991114412",
            "Wurth Elektronik",
            "74991114412",
        )
        label_pin_auto(schematic, jack, 1, "GND", 0.56)
        for pin, net_name in zip(range(2, 10), mdi_nets):
            label_pin_auto(schematic, jack, pin, net_name, 0.54)
        label_pin_auto(schematic, jack, "S1", "CHASSIS_GND", 0.54)
        label_pin_auto(schematic, jack, "S2", "CHASSIS_GND", 0.54)
        label_pin_auto(schematic, jack, 11, f"{prefix}_LED_A1", 0.54)
        label_pin_auto(schematic, jack, 12, led_sinks[0], 0.54)
        label_pin_auto(schematic, jack, 13, f"{prefix}_LED_A2", 0.54)
        label_pin_auto(schematic, jack, 14, led_sinks[1], 0.54)

        place_ethernet_esd(esd_references[0], mdi_nets[:4], (x - 52, y + 4))
        place_ethernet_esd(esd_references[1], mdi_nets[4:], (x + 52, y + 4))
        for reference_r, anode_net, x_offset in (
            (resistor_references[0], f"{prefix}_LED_A1", -14),
            (resistor_references[1], f"{prefix}_LED_A2", 14),
        ):
            resistor = add_network_passive("Device:R", reference_r, "330R", (x + x_offset, y + 43))
            label_two_pin_device(schematic, resistor, "NET_3V3", anode_net)
        note(
            schematic,
            f"{prefix}: 74991114412 voltage-mode 1 GbE; shields direct to chassis, PHY-side residual ESD",
            (x - 72, y + 59),
            0.66,
        )

    heading(schematic, "NATIVE CM5 WAN1 FRONT END", (190, 18), 1.5)
    place_magjack(
        "J610",
        "WAN1",
        240,
        90,
        (
            "WAN1_MDI0_P", "WAN1_MDI0_N", "WAN1_MDI1_P", "WAN1_MDI1_N",
            "WAN1_MDI2_P", "WAN1_MDI2_N", "WAN1_MDI3_P", "WAN1_MDI3_N",
        ),
        ("WAN1_LED0", "WAN1_LED1"),
        ("U630", "U631"),
        ("R650", "R651"),
    )

    def place_lan(reference: str, prefix: str, x: float) -> None:
        component = add_symbol(
            schematic,
            "CM5Carrier:LAN7430",
            reference,
            "LAN7430",
            (x, 100),
            "Package_DFN_QFN:QFN-48-1EP_7x7mm_P0.5mm_EP5.3x5.3mm_ThermalVias",
            "Microchip",
            "LAN7430T-I/Y9X",
        )
        pin_nets = {
            1: f"{prefix}_2V5_A", 2: f"{prefix}_MDI_A_P", 3: f"{prefix}_MDI_A_N",
            4: f"{prefix}_1V2_A", 5: f"{prefix}_MDI_B_P", 6: f"{prefix}_MDI_B_N",
            7: f"{prefix}_MDI_C_P", 8: f"{prefix}_MDI_C_N", 9: f"{prefix}_1V2_A",
            10: f"{prefix}_MDI_D_P", 11: f"{prefix}_MDI_D_N", 12: f"{prefix}_2V5_A",
            13: f"{prefix}_1V2", 14: f"{prefix}_1V2_A", 15: "GND",
            16: f"{prefix}_PCIE_RX_P", 17: f"{prefix}_PCIE_RX_N", 18: "GND",
            19: f"{prefix}_EP_TX_P", 20: f"{prefix}_1V2_A", 21: f"{prefix}_EP_TX_N",
            22: "GND", 23: f"{prefix}_2V5_A", 24: f"{prefix}_RESREF",
            25: f"{prefix}_PCIE_REFCLK_P", 26: f"{prefix}_PCIE_REFCLK_N",
            27: f"{prefix}_2V5_OUT", 28: "NET_3V3", 29: f"{prefix}_RESET_LOCAL_N",
            30: "GND", 31: f"{prefix}_1V2", 32: f"{prefix}_SW_NODE",
            33: "NET_3V3", 34: f"{prefix}_1V2", 35: "GND",
            37: f"{prefix}_LED1", 38: f"{prefix}_LED0",
            39: "NET_3V3", 40: f"{prefix}_1V2", 41: "NET_3V3",
            42: f"{prefix}_PCIE_CLKREQ_N", 43: "PCIE_UP_WAKE_N",
            44: f"{prefix}_PCIE_PERST_N", 45: f"{prefix}_1V2_A",
            46: f"{prefix}_XO", 47: f"{prefix}_XI", 48: f"{prefix}_ISET", 49: "GND",
        }
        for pin, net_name in pin_nets.items():
            label_pin_auto(schematic, component, pin, net_name, 0.58)
        schematic.no_connects.add(pin_xy(component, 36))

        inductor = add_symbol(
            schematic,
            "Device:L",
            f"L{reference[1:]}1",
            "3.3uH / 2.37A min",
            (x - 45, 175),
            manufacturer="TDK",
            mpn="VLS3012HBX-3R3M-N",
        )
        label_two_pin_device(schematic, inductor, f"{prefix}_SW_NODE", f"{prefix}_1V2")
        bulk = add_network_passive("Device:C", f"C{reference[1:]}1", "22uF 10V X7R", (x - 15, 175))
        label_two_pin_device(schematic, bulk, f"{prefix}_1V2", "GND")
        hf = add_network_passive("Device:C", f"C{reference[1:]}2", "100nF", (x + 15, 175))
        label_two_pin_device(schematic, hf, f"{prefix}_1V2", "GND")
        ldo = add_network_passive(
            "Device:C",
            f"C{reference[1:]}3",
            "1uF 35V X5R / ESR <1R",
            (x + 45, 175),
        )
        label_two_pin_device(schematic, ldo, f"{prefix}_2V5_OUT", "GND")
        resref = add_network_passive("Device:R", f"R{reference[1:]}1", "200R 1%", (x - 35, 205))
        label_two_pin_device(schematic, resref, f"{prefix}_RESREF", "GND")
        iset = add_network_passive("Device:R", f"R{reference[1:]}2", "6.04k 1%", (x + 5, 205))
        label_two_pin_device(schematic, iset, f"{prefix}_ISET", "GND")
        crystal = add_symbol(
            schematic,
            "CM5Carrier:Crystal_GND24_3225",
            f"Y{reference[1:]}",
            "25MHz / 10pF / -40..85C",
            (x + 52, 205),
            manufacturer="Abracon",
            mpn="ABM8-25.000MHZ-10-D1G-T",
        )
        label_pin_auto(schematic, crystal, 1, f"{prefix}_XI", 0.58)
        label_pin_auto(schematic, crystal, 3, f"{prefix}_XO", 0.58)
        label_pin_auto(schematic, crystal, 2, "GND", 0.58)
        label_pin_auto(schematic, crystal, 4, "GND", 0.58)
        crystal_caps = (
            (f"C{reference[1:]}4", f"{prefix}_XI", x + 35),
            (f"C{reference[1:]}5", f"{prefix}_XO", x + 69),
        )
        for cap_reference, clock_net, cap_x in crystal_caps:
            load_cap = add_network_passive("Device:C", cap_reference, "15pF C0G", (cap_x, 226))
            label_two_pin_device(schematic, load_cap, clock_net, "GND")
        for index, pin_net in enumerate(("NET_3V3",) * 4, start=6):
            decoupler = add_network_passive(
                "Device:C", f"C{reference[1:]}{index}", "10nF", (x - 75 + (index - 6) * 25, 205)
            )
            label_two_pin_device(schematic, decoupler, pin_net, "GND")
        for bead_index, source_net, load_net, bead_x in (
            (1, f"{prefix}_1V2", f"{prefix}_1V2_A", x - 55),
            (2, f"{prefix}_2V5_OUT", f"{prefix}_2V5_A", x - 20),
        ):
            bead = add_network_passive(
                "Device:FerriteBead", f"FB{reference[1:]}{bead_index}", "220R ferrite", (bead_x, 235)
            )
            label_two_pin_device(schematic, bead, source_net, load_net)
        for index, rail, cap_x in (
            (10, f"{prefix}_1V2_A", x + 15),
            (11, f"{prefix}_2V5_A", x + 45),
        ):
            decoupler = add_network_passive("Device:C", f"C{reference[1:]}{index}", "100nF", (cap_x, 235))
            label_two_pin_device(schematic, decoupler, rail, "GND")
        reset_pullup = add_network_passive("Device:R", f"R{reference[1:]}3", "10k", (x + 75, 205))
        label_two_pin_device(schematic, reset_pullup, "NET_3V3", f"{prefix}_RESET_LOCAL_N")
        for index, source_net, load_net, cap_x in (
            (12, f"{prefix}_SW_TX_P", f"{prefix}_PCIE_RX_P", x - 55),
            (13, f"{prefix}_SW_TX_N", f"{prefix}_PCIE_RX_N", x - 20),
            (14, f"{prefix}_EP_TX_P", f"{prefix}_PCIE_TX_P", x + 15),
            (15, f"{prefix}_EP_TX_N", f"{prefix}_PCIE_TX_N", x + 50),
        ):
            coupling = add_network_passive("Device:C", f"C{reference[1:]}{index}", "100nF", (cap_x, 265))
            label_two_pin_device(schematic, coupling, source_net, load_net)
        note(schematic, f"{prefix}: pin 35 tied low disables D3cold PME; pin 36 floats on its internal pull-down so advanced PCIe PM remains enabled.", (x - 82, 278), 0.59)
        note(schematic, f"{prefix}: RESET_N pulled high; PCIe TX pairs AC-coupled; analog 1.2 V/2.5 V rails ferrite-isolated.", (x - 82, 288), 0.62)
        suffix = {"WAN2": 2, "LAN1": 4, "LAN2": 6}[prefix]
        place_magjack(
            f"J{reference[1:]}",
            prefix,
            x,
            245,
            (
                f"{prefix}_MDI_A_P", f"{prefix}_MDI_A_N",
                f"{prefix}_MDI_B_P", f"{prefix}_MDI_B_N",
                f"{prefix}_MDI_C_P", f"{prefix}_MDI_C_N",
                f"{prefix}_MDI_D_P", f"{prefix}_MDI_D_N",
            ),
            (f"{prefix}_LED0", f"{prefix}_LED1"),
            (f"U63{suffix}", f"U63{suffix + 1}"),
            (f"R65{suffix}", f"R65{suffix + 1}"),
        )

    place_lan("U611", "WAN2", 350)
    place_lan("U612", "LAN1", 555)
    place_lan("U613", "LAN2", 760)

    heading(schematic, "3. WI-FI AP MINI PCIE 4T4R", (650, 315), 1.7)
    wifi = add_symbol(
        schematic,
        "CM5Carrier:Mini_PCIe_52",
        "J620",
        "AW7915-NP1_WIFI6_4T4R",
        (720, 420),
        footprint="CM5Carrier:Molex_0679101002_Mini_PCIe",
        manufacturer="Molex",
        mpn="0679101002",
    )
    wifi.set_property_effects(
        "Reference",
        {"position": (700, 378), "font_size": (0.9, 0.9), "justify_h": "left"},
    )
    wifi.set_property_effects(
        "Value",
        {"position": (700, 381), "font_size": (0.9, 0.9), "justify_h": "left"},
    )
    wifi_nets = {
        1: "PCIE_UP_WAKE_N",
        2: "WIFI_3V3", 24: "WIFI_3V3", 39: "WIFI_3V3",
        41: "WIFI_3V3", 52: "WIFI_3V3",
        7: "WIFI_PCIE_CLKREQ_N",
        11: "WIFI_PCIE_REFCLK_N", 13: "WIFI_PCIE_REFCLK_P",
        20: "WIFI_DISABLE1_N", 22: "WIFI_PCIE_PERST_N",
        23: "WIFI_PCIE_TX_N", 25: "WIFI_PCIE_TX_P",
        31: "WIFI_PCIE_RX_N", 33: "WIFI_PCIE_RX_P",
        30: "SYS_I2C7_SCL", 32: "SYS_I2C7_SDA",
    }
    wifi_ground_pins = {4, 9, 15, 18, 21, 26, 27, 29, 34, 35, 37, 40, 43, 50}
    for pin in range(1, 53):
        if wifi.get_pin_position(str(pin)) is None:
            continue
        if pin in wifi_nets:
            label_pin_auto(schematic, wifi, pin, wifi_nets[pin], 0.6)
        elif pin in wifi_ground_pins:
            label_pin_auto(schematic, wifi, pin, "GND", 0.58)
        else:
            schematic.no_connects.add(pin_xy(wifi, pin))
    if wifi.get_pin_position("MP") is not None:
        schematic.no_connects.add(pin_xy(wifi, "MP"))
    wifi_disable_pullup = add_network_passive("Device:R", "R620", "10k", (620, 490))
    label_two_pin_device(schematic, wifi_disable_pullup, "WIFI_DISABLE1_N", "WIFI_3V3")
    note(schematic, "AW7915-NP1: Wi-Fi 6 4T4R Mini PCIe AP; vendor maximum 9.1 W and 3.5 A recommended; dedicated rail is 3.3 V / 4 A.", (590, 540), 0.76)
    note(schematic, "J620 uses the controlled SD-67910-001 C2 land pattern; 3D/first-article fit remains a release gate.", (590, 550), 0.78)

    note(schematic, "Switch port 5 is disabled and unrouted to avoid PCIe Gen2 stubs and an unsafe generic test header.", (610, 345), 0.82)

    for reference, net_name, x in (
        ("#FLG0601", "NET_3V3", 320),
        ("#FLG0602", "PCIE_1V0", 360),
        ("#FLG0603", "WIFI_3V3", 400),
        ("#FLG0606", "CHASSIS_GND", 520),
    ):
        flag = add_symbol(schematic, "power:PWR_FLAG", reference, "PWR_FLAG", (x, 550))
        label_pin_auto(schematic, flag, 1, net_name, 0.7)
    note(schematic, "PWR_FLAG symbols declare off-sheet regulator outputs for ERC; they are not physical parts.", (300, 570), 0.8)

    output = folder / f"{name}.kicad_sch"
    save_generated_schematic(schematic, output)
    return output


def build_wwan_sim_sheet() -> Path:
    """Capture the universal M.2 B-key modem, USB links, controls, and dual SIM."""
    folder = ROOT / "CM5-CARRIER"
    name = "WWAN-SIM"
    create_project(folder, name)
    schematic = create_schematic(name)
    schematic.set_paper_size("A1")
    schematic.set_title_block(
        title="Radxa CM5 ProComm Carrier - WWAN and Dual SIM",
        date="2026-08-14",
        rev="A1",
        company="ProComm",
        comments={
            1: "M.2 B-key modem socket supports USB 3.0 and USB 2.0 fallback",
            2: "Dedicated 3.8 V / 6 A rail; verify selected modem peak-current waveform",
            3: "FSA2567 defaults to physical SIM 1; open-drain low selects physical SIM 2",
            4: "RF coax leaves the module directly for four right-side bulkhead antennas",
        },
    )

    def add_wwan_passive(
        lib_id: str,
        reference: str,
        value: str,
        position: tuple[float, float],
    ):
        try:
            manufacturer, mpn = WWAN_PASSIVE_PARTS[value]
        except KeyError as error:
            raise RuntimeError(
                f"WWAN-SIM {reference} has no locked production passive for {value}"
            ) from error
        return add_symbol(
            schematic,
            lib_id,
            reference,
            value,
            position,
            manufacturer=manufacturer,
            mpn=mpn,
        )

    heading(schematic, "1. UNIVERSAL M.2 B-KEY WWAN SOCKET", (45, 20), 1.9)
    modem = add_symbol(
        schematic,
        "Connector:Bus_M.2_Socket_B",
        "J701",
        "WWAN_M2_B_KEY_USB3_USB2",
        (155, 145),
        "CM5Carrier:TE_2199230-3_M2_Key_B_4.2mm",
        "TE Connectivity",
        "2199230-3",
    )
    modem_nets = {
        2: "MODEM_3V8", 4: "MODEM_3V8", 70: "MODEM_3V8",
        72: "MODEM_3V8", 74: "MODEM_3V8",
        6: "MODEM_FULL_CARD_POWER_OFF_N", 7: "WWAN_USB2_DP",
        8: "MODEM_W_DISABLE1_N", 9: "WWAN_USB2_DM",
        10: "MODEM_STATUS_LED_N",
        29: "WWAN_USB3_TX_N", 31: "WWAN_USB3_TX_P",
        30: "MODEM_UIM_RESET", 32: "MODEM_UIM_CLK",
        34: "MODEM_UIM_DATA", 35: "WWAN_USB3_RX_N",
        36: "MODEM_UIM_PWR", 37: "WWAN_USB3_RX_P",
        54: "MODEM_WAKE_N", 67: "MODEM_RESET_N",
    }
    modem_ground_pins = {3, 5, 11, 27, 33, 39, 45, 51, 57, 71, 73}
    for pin in range(1, 76):
        if modem.get_pin_position(str(pin)) is None:
            continue
        if pin in modem_nets:
            label_pin_auto(schematic, modem, pin, modem_nets[pin], 0.58)
        elif pin in modem_ground_pins:
            label_pin_auto(schematic, modem, pin, "GND", 0.58)
        else:
            schematic.no_connects.add(pin_xy(modem, pin))
    note(schematic, "Socket power pins retain the stock symbol's 3.3 V names but are intentionally fed from MODEM_3V8.", (45, 260), 0.82)
    note(schematic, "Qualification window: 3.135-4.4 V module input, >=5 A transient capability, keyed B socket, 42 mm standoff.", (45, 270), 0.82)

    heading(schematic, "2. USB ESD AND MODEM BULK CAPACITANCE", (300, 20), 1.75)
    usb_esd_1 = add_symbol(
        schematic, "Power_Protection:TPD4E05U06DQA", "U702",
        "TPD4E05U06DQA", (365, 80), manufacturer="Texas Instruments", mpn="TPD4E05U06DQAR",
    )
    for pin, net_name in {
        1: "WWAN_USB2_DP", 2: "WWAN_USB2_DM",
        4: "WWAN_USB3_TX_P", 5: "WWAN_USB3_TX_N", 3: "GND", 8: "GND",
    }.items():
        label_pin_auto(schematic, usb_esd_1, pin, net_name, 0.64)
    for pin in (6, 7, 9, 10):
        schematic.no_connects.add(pin_xy(usb_esd_1, pin))
    usb_esd_2 = add_symbol(
        schematic, "Power_Protection:TPD4E05U06DQA", "U703",
        "TPD4E05U06DQA", (365, 145), manufacturer="Texas Instruments", mpn="TPD4E05U06DQAR",
    )
    for pin, net_name in {
        1: "WWAN_USB3_RX_P", 2: "WWAN_USB3_RX_N", 3: "GND", 8: "GND",
    }.items():
        label_pin_auto(schematic, usb_esd_2, pin, net_name, 0.64)
    for pin in (4, 5, 6, 7, 9, 10):
        schematic.no_connects.add(pin_xy(usb_esd_2, pin))
    note(schematic, "Place USB ESD at the modem socket; route USB 3 pairs as short, continuous 90 ohm differential channels.", (300, 190), 0.8)

    local_modem_caps = (
        ("C701", "220uF 6.3V polymer", 300, 225),
        ("C702", "220uF 6.3V polymer", 350, 225),
        ("C703", "100nF", 400, 225),
        ("C704", "6.8nF C0G", 440, 225),
        ("C705", "220pF C0G", 480, 225),
        ("C706", "68pF C0G", 520, 225),
        ("C707", "100nF", 320, 260),
        ("C708", "220pF C0G", 360, 260),
        ("C709", "68pF C0G", 400, 260),
        ("C710", "15pF C0G", 440, 260),
        ("C711", "9.1pF C0G", 480, 260),
        ("C712", "4.7pF C0G", 520, 260),
    )
    for reference, value, x, y in local_modem_caps:
        capacitor = add_wwan_passive("Device:C", reference, value, (x, y))
        label_two_pin_device(schematic, capacitor, "MODEM_3V8", "GND")
    modem_tvs = add_symbol(
        schematic,
        "Device:D_TVS",
        "D701",
        "5.0V TVS",
        (555, 242),
        manufacturer="Littelfuse",
        mpn="SMF4L5.0AT1G",
    )
    label_two_pin_device(schematic, modem_tvs, "MODEM_3V8", "GND")
    flag = add_symbol(schematic, "power:PWR_FLAG", "#FLG0701", "PWR_FLAG", (310, 270))
    label_pin_auto(schematic, flag, 1, "MODEM_3V8", 0.7)
    note(schematic, "C701-C712 and D701 implement the RM520N-GL two-bank VCC reference network directly at M.2 pins 2/4 and 70/72/74.", (300, 290), 0.72)
    note(schematic, "Power-Regulators C1189-C1192 retain 1320 uF upstream bulk; they do not replace the socket-local 2 x 220 uF and HF ladder.", (300, 300), 0.72)

    heading(schematic, "3. FSA2567 DUAL-SIM MUX", (470, 20), 1.75)
    mux = add_symbol(
        schematic,
        "CM5Carrier:FSA2567MPX",
        "U704",
        "FSA2567MPX",
        (530, 115),
        "Package_DFN_QFN:WQFN-16-1EP_3x3mm_P0.5mm_EP1.68x1.68mm",
        "onsemi",
        "FSA2567MPX",
    )
    mux_nets = {
        # FSA2567 SEL high selects channel 2. Physical SIM 1 is therefore
        # cross-mapped to channel 2 so the UIM-domain pull-up is the safe default.
        1: "SIM2_VCC", 2: "MODEM_SIM_MUX_SEL_OD", 3: "SIM1_RESET_RAW",
        4: "MODEM_UIM_RESET", 5: "SIM2_RESET_RAW", 6: "GND",
        7: "SIM1_CLK_RAW", 8: "MODEM_UIM_CLK", 9: "SIM2_CLK_RAW",
        11: "SIM1_DATA_RAW", 12: "MODEM_UIM_DATA", 13: "SIM2_DATA_RAW",
        14: "MODEM_UIM_PWR", 15: "SIM1_VCC", 16: "MODEM_UIM_PWR", 17: "GND",
    }
    for pin, net_name in mux_nets.items():
        label_pin_auto(schematic, mux, pin, net_name, 0.61)
    schematic.no_connects.add(pin_xy(mux, 10))
    pullup = add_wwan_passive("Device:R", "R701", "100k", (530, 185))
    label_two_pin_device(schematic, pullup, "MODEM_UIM_PWR", "MODEM_SIM_MUX_SEL_OD")
    note(schematic, "R701 keeps SEL high: SIM 1 is selected by default. GPIO control must be open drain; low selects SIM 2.", (470, 210), 0.82)
    note(schematic, "Physical SIM 1 uses FSA2567 channel 2 and physical SIM 2 uses channel 1 to implement that default.", (470, 230), 0.76)
    note(schematic, "The control domain tracks MODEM_UIM_PWR, preventing a fixed 3.3 V overdrive when the SIM rail is 1.8 V.", (470, 220), 0.82)

    def place_sim(reference: str, index: int, x: float) -> None:
        prefix = f"SIM{index}"
        holder = add_symbol(
            schematic,
            "CM5Carrier:Wurth_Nano_SIM_693043020611",
            reference,
            "693043020611",
            (x, 105),
            "CM5Carrier:J_Wurth_WR-CRD_693043020611",
            "Wurth Elektronik",
            "693043020611",
        )
        holder_nets = {
            "C1": f"{prefix}_VCC", "C2": f"{prefix}_RESET",
            "C3": f"{prefix}_CLK", "C5": "GND", "C7": f"{prefix}_DATA",
            "S1": "CHASSIS_GND", "S2": "CHASSIS_GND", "S3": "CHASSIS_GND",
            "S4": "CHASSIS_GND", "S5": "CHASSIS_GND", "S6": "CHASSIS_GND",
        }
        for pin, net_name in holder_nets.items():
            label_pin_auto(schematic, holder, pin, net_name, 0.59)
        schematic.no_connects.add(pin_xy(holder, "C6"))
        sim_filter = add_symbol(
            schematic, "Power_Protection:TPD3F303DPV", f"U{704 + index}",
            "TPD3F303DPV", (x, 210), manufacturer="Texas Instruments", mpn="TPD3F303DPVR",
        )
        for pin, net_name in {
            1: f"{prefix}_DATA", 2: f"{prefix}_CLK", 3: f"{prefix}_RESET",
            5: f"{prefix}_VCC", 6: f"{prefix}_RESET_RAW",
            7: f"{prefix}_CLK_RAW", 8: f"{prefix}_DATA_RAW", 9: "GND",
        }.items():
            label_pin_auto(schematic, sim_filter, pin, net_name, 0.6)
        schematic.no_connects.add(pin_xy(sim_filter, 4))
        note(schematic, f"{prefix}: TPD3F303 filters RESET/CLK/DATA and clamps VCC beside the holder; shell bonds to CHASSIS_GND.", (x - 65, 265), 0.75)

    heading(schematic, "4. TWO FIELD-ACCESSIBLE NANO-SIM HOLDERS", (610, 20), 1.65)
    place_sim("J702", 1, 650)
    place_sim("J703", 2, 755)
    for reference, net_name, x in (
        ("#FLG0705", "SIM1_VCC", 650),
        ("#FLG0706", "SIM2_VCC", 755),
    ):
        sim_power_flag = add_symbol(schematic, "power:PWR_FLAG", reference, "PWR_FLAG", (x, 285))
        label_pin_auto(schematic, sim_power_flag, 1, net_name, 0.60)

    heading(schematic, "5. RF BULKHEAD HARNESS CONTRACT", (525, 315), 1.65)
    for index, (label, x) in enumerate(
        (("CELL_1_RF", 550), ("CELL_2_RF", 625), ("CELL_3_RF", 700), ("CELL_4_GNSS_RF", 775)),
        start=1,
    ):
        connector = add_symbol(
            schematic, "Connector:Conn_Coaxial", f"J{710 + index}",
            label, (x, 365), manufacturer="ECT", mpn="818033349",
        )
        connector.on_board = False
        schematic.no_connects.add(pin_xy(connector, 1))
        schematic.no_connects.add(pin_xy(connector, 2))
        note(schematic, label.replace("_RF", ""), (x - 22, 395), 0.75)
    note(schematic, "Four off-board ECT 818033349 USS RF IV-to-SMA bulkhead pigtails; final cable length remains a routed sample gate.", (525, 425), 0.82)

    heading(schematic, "6. MODEM CONTROL CONTRACT", (45, 325), 1.65)
    for index, (net_name, x, y) in enumerate(
        (
            ("MODEM_FULL_CARD_POWER_OFF_N", 80, 380),
            ("MODEM_W_DISABLE1_N", 120, 380),
            ("MODEM_RESET_N", 160, 380),
            ("MODEM_WAKE_N", 200, 380),
            ("MODEM_STATUS_LED_N", 80, 420),
            ("MODEM_SIM_MUX_SEL_OD", 120, 420),
            ("MODEM_3V8", 160, 420),
            ("GND", 200, 420),
        ),
        start=1,
    ):
        test_point = add_symbol(
            schematic,
            "Connector:TestPoint",
            f"TP72{index:02d}",
            net_name,
            (x, y),
            footprint=TEST_POINT_2MM,
        )
        label_pin_auto(schematic, test_point, 1, net_name, 0.52)
    note(schematic, "Control outputs originate at the system GPIO expander. FULL_CARD_POWER_OFF and RESET require final modem-specific timing validation.", (45, 445), 0.82)
    note(schematic, "TP7201-TP7208 are copper factory probes. USB 2/3 has no branch or generic test header.", (45, 455), 0.82)
    output = folder / f"{name}.kicad_sch"
    save_generated_schematic(schematic, output)
    return output


def build_display_harness_sheet() -> Path:
    """Capture the underside-facing HDMI, USB touch, and 12 V lid harness."""
    folder = ROOT / "CM5-CARRIER"
    name = "Display-Harness"
    create_project(folder, name)
    schematic = create_schematic(name)
    schematic.set_paper_size("A1")
    schematic.set_title_block(
        title="Radxa CM5 ProComm Carrier - Lid Display Harness",
        date="2026-08-14",
        rev="A1",
        company="ProComm",
        comments={
            1: "15.6-inch lid monitor: HDMI video, USB 2 touch, 12 V / 2.5 A",
            2: "Carrier connectors face downward at the hinge-edge harness notch",
            3: "USB-A is the CM5 host end; cable terminates in USB-B SuperSpeed at the monitor",
            4: "No display eFuse; separate fused DISPLAY_12V and DISPLAY_IO_12V branches feed this sheet",
        },
    )

    heading(schematic, "1. HDMI TYPE-A SOURCE CONNECTOR", (45, 20), 1.85)
    hdmi = add_symbol(
        schematic,
        "Connector:HDMI_A",
        "J801",
        "HDMI_TO_LID_DISPLAY",
        (135, 120),
        "Connector_Video:HDMI_A_Molex_208658-1001_Horizontal",
        "Molex",
        "208658-1001",
    )
    hdmi_nets = {
        1: "HDMI_D2_P", 2: "GND", 3: "HDMI_D2_N",
        4: "HDMI_D1_P", 5: "GND", 6: "HDMI_D1_N",
        7: "HDMI_D0_P", 8: "GND", 9: "HDMI_D0_N",
        10: "HDMI_CLK_P", 11: "GND", 12: "HDMI_CLK_N",
        13: "HDMI_CEC", 14: "HDMI_HEAC_P", 15: "HDMI_DDC_SCL",
        16: "HDMI_DDC_SDA", 17: "GND", 18: "HDMI_5V_OUT",
        19: "HDMI_HPD", "SH": "CHASSIS_GND",
    }
    for pin, net_name in hdmi_nets.items():
        label_pin_auto(schematic, hdmi, pin, net_name, 0.62)
    note(schematic, "HDMI_HEAC_N remains unused in Rev A; the selected monitor uses video/DDC/CEC/HPD only.", (45, 225), 0.82)
    note(schematic, "Shell bonds to chassis at the connector. TMDS shield returns use the local digital ground plane.", (45, 235), 0.82)

    heading(schematic, "2. HDMI ESD / LOW-CAPACITANCE SHUNTS", (270, 20), 1.65)
    hdmi_esd_maps = (
        ("U801", (1, "HDMI_D2_P"), (2, "HDMI_D2_N"), (4, "HDMI_D1_P"), (5, "HDMI_D1_N"), 80),
        ("U802", (1, "HDMI_D0_P"), (2, "HDMI_D0_N"), (4, "HDMI_CLK_P"), (5, "HDMI_CLK_N"), 145),
        ("U803", (1, "HDMI_CEC"), (2, "HDMI_DDC_SCL"), (4, "HDMI_DDC_SDA"), (5, "HDMI_HPD"), 210),
    )
    for reference, pin1, pin2, pin4, pin5, y in hdmi_esd_maps:
        protector = add_symbol(
            schematic, "Power_Protection:TPD4E05U06DQA", reference,
            "TPD4E05U06DQA", (365, y), manufacturer="Texas Instruments", mpn="TPD4E05U06DQAR",
        )
        for pin, net_name in (pin1, pin2, pin4, pin5, (3, "GND"), (8, "GND")):
            label_pin_auto(schematic, protector, pin, net_name, 0.61)
        for pin in (6, 7, 9, 10):
            schematic.no_connects.add(pin_xy(protector, pin))
    note(schematic, "Place all HDMI ESD arrays immediately behind J801 with no stubs on TMDS pairs.", (270, 255), 0.82)

    hdmi_fuse = add_symbol(
        schematic,
        "Device:Polyfuse",
        "F801",
        "0.10A hold / 15V",
        (465, 80),
        manufacturer="Littelfuse",
        mpn="0603L010YR",
    )
    label_two_pin_device(schematic, hdmi_fuse, "IO_5V0", "HDMI_5V_OUT")
    note(schematic, "HDMI DDC connects directly to the CM5 5 V DDC pins, matching Radxa CM5 IO V2.2; do not add duplicate carrier pull-ups.", (425, 170), 0.78)

    with isolated_uuid_namespace("display-io-5v"):
        heading(schematic, "3. DEDICATED IO_5V0 / 2 A BUCK", (535, 20), 1.55)
        io_buck = add_symbol(
            schematic,
            "CM5Carrier:TPS62913RPU",
            "U805",
            "TPS62913RPU",
            (610, 100),
            footprint="Package_DFN_QFN:Texas_RPU0010A_VQFN-HR-10_2x2mm_P0.5mm",
            manufacturer="Texas Instruments",
            mpn="TPS62913RPUT",
        )
        for pin, net_name in {
            1: "DISPLAY_IO_12V", 2: "IO5V0_SW", 3: "IO_5V0", 4: "GND",
            5: "IO_5V0_PG", 6: "DISPLAY_IO_12V", 7: "GND",
            8: "IO5V0_NRSS", 9: "IO5V0_FB", 10: "IO5V0_SCONF",
        }.items():
            label_pin_auto(schematic, io_buck, pin, net_name, 0.42)
        lock_component_pin_uuids(io_buck)
        io_inductor = add_symbol(
            schematic, "Device:L", "L805", "2.2uH / 3.7A", (690, 100),
            manufacturer="Coilcraft", mpn="XGL4030-222MEC",
        )
        label_two_pin_device(schematic, io_inductor, "IO5V0_SW", "IO_5V0")
        lock_component_pin_uuids(io_inductor)
        for reference, value, net_1, net_2, x, y in (
            ("R805", "26.1k 0.1%", "IO_5V0", "IO5V0_FB", 545, 170),
            ("R806", "4.99k 0.1%", "IO5V0_FB", "GND", 585, 170),
            ("C805", "470nF", "IO5V0_NRSS", "GND", 625, 170),
            ("R807", "7.5k S-CONF", "IO5V0_SCONF", "GND", 665, 170),
            ("C806", "10uF 25V X7R", "DISPLAY_IO_12V", "GND", 545, 215),
            ("C807", "10uF 25V X7R", "DISPLAY_IO_12V", "GND", 585, 215),
            ("C808", "22uF 10V X7R", "IO_5V0", "GND", 645, 215),
            ("C809", "22uF 10V X7R", "IO_5V0", "GND", 685, 215),
            ("C810", "22uF 10V X7R", "IO_5V0", "GND", 725, 215),
        ):
            manufacturer, mpn = DISPLAY_PASSIVE_PARTS[value]
            part = add_symbol(
                schematic, "Device:R" if reference.startswith("R") else "Device:C",
                reference, value, (x, y), manufacturer=manufacturer, mpn=mpn,
            )
            label_two_pin_device(schematic, part, net_1, net_2)
            lock_component_pin_uuids(part)
        note(schematic, "IO_5V0 feeds CM5 U13-B pin 106, HDMI source 5 V and USB-touch VBUS; no monitor-power eFuse.", (535, 255), 0.68)

    heading(schematic, "4. USB TOUCH HOST PORT", (45, 285), 1.8)
    touch = add_symbol(
        schematic,
        "Connector:USB3_A",
        "J802",
        "TOUCH_USB_HOST",
        (135, 350),
        "Connector_USB:USB3_A_Receptacle_Wuerth_692122030100",
        "Wurth Elektronik",
        "692122030100",
    )
    for pin, net_name in {
        1: "TOUCH_USB_5V", 2: "TOUCH_USB_DM", 3: "TOUCH_USB_DP",
        4: "GND", "SH": "CHASSIS_GND",
    }.items():
        label_pin_auto(schematic, touch, pin, net_name, 0.64)
    for pin in (5, 6, 7, 8, 9):
        schematic.no_connects.add(pin_xy(touch, pin))
    note(schematic, "Only USB 2 is allocated from the CM5. A standard USB 3 A-to-B cable still carries USB 2 touch data.", (45, 425), 0.82)
    note(schematic, "The monitor-side SuperSpeed Type-B receptacle is not duplicated as a device connector on this host carrier.", (45, 435), 0.82)

    touch_esd = add_symbol(
        schematic, "Power_Protection:TPD4E05U06DQA", "U804",
        "TPD4E05U06DQA", (290, 350), manufacturer="Texas Instruments", mpn="TPD4E05U06DQAR",
    )
    for pin, net_name in {
        1: "TOUCH_USB_DP", 2: "TOUCH_USB_DM", 3: "GND", 8: "GND",
    }.items():
        label_pin_auto(schematic, touch_esd, pin, net_name, 0.62)
    for pin in (4, 5, 6, 7, 9, 10):
        schematic.no_connects.add(pin_xy(touch_esd, pin))
    touch_fuse = add_symbol(
        schematic,
        "Device:Polyfuse",
        "F802",
        "1.10A hold / 16V",
        (360, 350),
        manufacturer="Littelfuse",
        mpn="1206L110/16WR",
    )
    label_two_pin_device(schematic, touch_fuse, "IO_5V0", "TOUCH_USB_5V")
    for reference, value, x in (("C801", "10uF 25V X5R", 420), ("C802", "100nF", 465)):
        manufacturer, mpn = DISPLAY_PASSIVE_PARTS[value]
        capacitor = add_symbol(
            schematic,
            "Device:C",
            reference,
            value,
            (x, 350),
            manufacturer=manufacturer,
            mpn=mpn,
        )
        label_two_pin_device(schematic, capacitor, "TOUCH_USB_5V", "GND")

    heading(schematic, "5. 12 V / 2.5 A DISPLAY POWER", (420, 285), 1.75)
    add_connector(
        schematic,
        "Connector_Generic:Conn_02x02_Odd_Even",
        "J803",
        "LID_DISPLAY_12V",
        (505, 350),
        {1: "DISPLAY_12V", 2: "DISPLAY_12V", 3: "GND", 4: "GND"},
        MF_2X2,
        "Molex",
        "43045-0412",
        two_row=True,
    )
    display_flag = add_symbol(schematic, "power:PWR_FLAG", "#FLG0801", "PWR_FLAG", (475, 410))
    label_pin_auto(schematic, display_flag, 1, "DISPLAY_12V", 0.66)
    note(schematic, "DISPLAY_12V design load: 2.5 A continuous / 30 W branch, covering the locked 25 W monitor.", (400, 445), 0.82)
    note(schematic, "No local display eFuse. Use the upstream AUX_12V branch fuse and H03C keyed 18 AWG harness.", (400, 455), 0.82)

    heading(schematic, "6. HINGE-EDGE HARNESS BUILD", (45, 490), 1.65)
    note(schematic, "H03A HDMI, H03B USB touch, H03C 12 V: 1000 +/- 25 mm each, with separate retained hinge and service loops.", (45, 510), 0.76)
    note(schematic, "Acceptance: full lid travel plus 300 mm panel lift / 45 degree tilt; no connector tension or hinge-line bend.", (45, 520), 0.76)
    note(schematic, "Carrier connectors face down. Store released loops clear of fans, PSU airflow, board keepouts and sharp edges.", (45, 530), 0.76)

    for reference, net_name, x in (
        ("#FLG0802", "DISPLAY_IO_12V", 270),
        ("#FLG0803", "IO_5V0", 310),
        ("#FLG0806", "HDMI_5V_OUT", 430),
    ):
        flag = add_symbol(schematic, "power:PWR_FLAG", reference, "PWR_FLAG", (x, 560))
        label_pin_auto(schematic, flag, 1, net_name, 0.64)
    with isolated_uuid_namespace("display-touch-power-flag"):
        touch_power_flag = add_symbol(
            schematic, "power:PWR_FLAG", "#FLG0807", "PWR_FLAG", (470, 560),
        )
        label_pin_auto(schematic, touch_power_flag, 1, "TOUCH_USB_5V", 0.64)
        lock_component_pin_uuids(touch_power_flag)

    output = folder / f"{name}.kicad_sch"
    save_generated_schematic(schematic, output)
    return output


def build_audio_control_sheet() -> Path:
    """Capture both CM5 audio interfaces, the TDM line drivers, and CTIA headset."""
    folder = ROOT / "CM5-CARRIER"
    name = "Audio-Control"
    create_project(folder, name)
    schematic = create_schematic(name)
    schematic.set_paper_size("A0")
    schematic.set_title_block(
        title="Radxa CM5 ProComm Carrier - TDM and Headset Audio",
        date="2026-08-16",
        rev="A1",
        company="ProComm",
        comments={
            1: "I2S0/TDM program audio is buffered over LVDS to the separate AUDIO-8X8 board",
            2: "I2S1 uses explicit 3.3 V to 1.8 V translation before the ES8316 codec",
            3: "TPA6132A2 provides 0 dB dedicated stereo headphone amplification",
            4: "Kycon STX-353K7A is CTIA-mapped; its preliminary land pattern requires a coupon",
        },
    )

    def add_audio_passive(
        lib_id: str,
        reference: str,
        value: str,
        position: tuple[float, float],
    ):
        manufacturer, mpn = AUDIO_PASSIVE_PARTS[value]
        return add_symbol(
            schematic,
            lib_id,
            reference,
            value,
            position,
            manufacturer=manufacturer,
            mpn=mpn,
        )

    heading(schematic, "1. I2S0/TDM PROGRAM-AUDIO LVDS LINK", (45, 20), 1.9)
    cm5_audio = add_symbol(
        schematic,
        "CM5Carrier:CM5_Audio_Port",
        "U900",
        "CM5_AUDIO_OFFSHEET_PORT",
        (65, 135),
    )
    cm5_audio.in_bom = False
    cm5_audio.on_board = False
    for pin, net_name in {
        1: "AUD_MCLK", 2: "AUD_BCLK", 3: "AUD_FSYNC", 4: "AUD_DAC_SDIN",
        5: "AUD_ADC_SDOUT", 6: "SYS_I2C7_SCL", 7: "SYS_I2C7_SDA",
        8: "AUD_IRQ_N", 9: "HS_MCLK", 10: "HS_BCLK", 11: "HS_LRCK",
        12: "HS_SDOUT_TO_CODEC", 13: "HS_SDIN_FROM_CODEC",
        14: "HS_I2C_SCL", 15: "HS_I2C_SDA", 16: "HS_JACK_DET_N",
    }.items():
        label_pin_auto(schematic, cm5_audio, pin, net_name, 0.58)
    driver = add_symbol(
        schematic,
        "Interface:SN65LVDS047PW",
        "U901",
        "SN65LVDS047PW",
        (145, 105),
        manufacturer="Texas Instruments",
        mpn="SN65LVDS047PWR",
    )
    for pin, net_name in {
        1: "AUDIO_ENABLE", 2: "AUD_MCLK", 3: "AUD_BCLK",
        4: "LOGIC_3V3", 5: "GND", 6: "AUD_FSYNC", 7: "AUD_DAC_SDIN",
        8: "GND", 9: "AUD_DAC_SDIN_N", 10: "AUD_DAC_SDIN_P",
        11: "AUD_FSYNC_P", 12: "AUD_FSYNC_N", 13: "AUD_BCLK_N",
        14: "AUD_BCLK_P", 15: "AUD_MCLK_P", 16: "AUD_MCLK_N",
    }.items():
        label_pin_auto(schematic, driver, pin, net_name, 0.62)
    receiver = add_symbol(
        schematic,
        "Interface:SN65LVDT2D",
        "U902",
        "SN65LVDT2D",
        (145, 220),
        manufacturer="Texas Instruments",
        mpn="SN65LVDT2DR",
    )
    for pin, net_name in {
        1: "AUD_ADC_SDOUT_N", 2: "AUD_ADC_SDOUT_P", 5: "GND",
        7: "AUD_ADC_SDOUT", 8: "LOGIC_3V3",
    }.items():
        label_pin_auto(schematic, receiver, pin, net_name, 0.62)
    for pin in (3, 4, 6):
        schematic.no_connects.add(pin_xy(receiver, pin))
    for reference, value, rail, x in (
        ("C901", "100nF", "LOGIC_3V3", 95),
        ("C902", "1uF", "LOGIC_3V3", 135),
    ):
        capacitor = add_audio_passive("Device:C", reference, value, (x, 285))
        label_two_pin_device(schematic, capacitor, rail, "GND")

    translator = add_symbol(
        schematic,
        "CM5Carrier:PCA9517ADP_A1",
        "U903",
        "PCA9517ADP",
        (300, 105),
        manufacturer="NXP",
        mpn="PCA9517ADP,118",
    )
    for pin, net_name in {
        1: "LOGIC_3V3", 2: "SYS_I2C7_SCL", 3: "SYS_I2C7_SDA", 4: "GND",
        5: "AUDIO_ENABLE", 6: "AUD_I2C_SDA", 7: "AUD_I2C_SCL", 8: "LOGIC_3V3",
    }.items():
        label_pin_auto(schematic, translator, pin, net_name, 0.62)
    for index, (rail, net_name, x) in enumerate(
        (("LOGIC_3V3", "SYS_I2C7_SCL", 250), ("LOGIC_3V3", "SYS_I2C7_SDA", 285),
         ("LOGIC_3V3", "AUD_I2C_SCL", 320), ("LOGIC_3V3", "AUD_I2C_SDA", 355)),
        start=1,
    ):
        resistor = add_audio_passive("Device:R", f"R90{index}", "2.2k", (x, 180))
        label_two_pin_device(schematic, resistor, rail, net_name)

    tdm_map = {
        1: "AUD_MCLK_P", 2: "AUD_MCLK_N", 3: "GND", 4: "GND",
        5: "AUD_BCLK_P", 6: "AUD_BCLK_N", 7: "AUD_FSYNC_P", 8: "AUD_FSYNC_N",
        9: "GND", 10: "GND", 11: "AUD_DAC_SDIN_P", 12: "AUD_DAC_SDIN_N",
        13: "AUD_ADC_SDOUT_P", 14: "AUD_ADC_SDOUT_N", 15: "GND", 16: "GND",
        17: "AUD_I2C_SCL", 18: "AUD_I2C_SDA", 19: "AUD_ADC_RST_N",
        20: "AUD_DAC_RST_N", 21: "AUD_DAC_MUTE_CMD_N", 22: "AUD_IRQ_N",
        23: "AUDIO_PRESENT_N", 24: "AUDIO_ENABLE", 25: "LOGIC_3V3", 26: "GND",
        27: "TDM_SPARE_1", 28: "TDM_SPARE_2", 29: "GND", 30: "GND",
    }
    add_connector(
        schematic,
        "Connector_Generic:Conn_02x15_Odd_Even",
        "J901",
        "AUDIO_8X8_TDM_CONTROL",
        (455, 135),
        tdm_map,
        MILLIGRID_2X15,
        "Molex",
        "87832-6423",
        two_row=True,
    )
    note(schematic, "J901 mates with AUDIO-8X8 J101. Route five 100 ohm differential pairs with the assigned interleaved returns.", (380, 260), 0.82)
    note(schematic, "SN65LVDT2D includes the ADC-data receiver termination; do not add a duplicate far-end resistor.", (380, 270), 0.82)

    heading(schematic, "2. I2S1 HEADSET LEVEL TRANSLATION", (45, 340), 1.8)
    headset_i2c = add_symbol(
        schematic,
        "Interface:PCA9306DP",
        "U904",
        "PCA9306DP",
        (130, 410),
        manufacturer="NXP",
        mpn="PCA9306DP,118",
    )
    headset_i2c.hidden_properties.add("Value")
    for pin, net_name in {
        1: "HEADSET_AGND", 2: "HEADSET_1V8", 3: "HS_CODEC_I2C_SCL",
        4: "HS_CODEC_I2C_SDA", 5: "HS_I2C_SDA", 6: "HS_I2C_SCL",
        7: "HS_I2C_BIAS", 8: "HS_I2C_BIAS",
    }.items():
        label_pin_auto(schematic, headset_i2c, pin, net_name, 0.58)
    for reference, rail, net_name, x in (
        ("R905", "HEADSET_1V8", "HS_CODEC_I2C_SCL", 65),
        ("R906", "HEADSET_1V8", "HS_CODEC_I2C_SDA", 110),
        ("R907", "LOGIC_3V3", "HS_I2C_SCL", 155),
        ("R908", "LOGIC_3V3", "HS_I2C_SDA", 200),
    ):
        resistor = add_audio_passive("Device:R", reference, "2.2k", (x, 485))
        label_two_pin_device(schematic, resistor, rail, net_name)
    i2c_bias = add_audio_passive("Device:R", "R909", "200k", (235, 450))
    label_two_pin_device(schematic, i2c_bias, "LOGIC_3V3", "HS_I2C_BIAS")
    note(schematic, "PCA9306 VREF2 and EN share HS_I2C_BIAS through the required 200 k pull-up to 3.3 V.", (45, 505), 0.72)

    i2s_to_codec = add_symbol(
        schematic,
        "Logic_LevelTranslator:SN74AVC4T245PW",
        "U905",
        "SN74AVC4T245PW",
        (330, 410),
        manufacturer="Texas Instruments",
        mpn="SN74AVC4T245PWR",
    )
    i2s_to_codec.hidden_properties.add("Value")
    for pin, net_name in {
        1: "LOGIC_3V3", 2: "LOGIC_3V3", 3: "LOGIC_3V3",
        4: "HS_MCLK", 5: "HS_BCLK", 6: "HS_LRCK", 7: "HS_SDOUT_TO_CODEC",
        8: "HEADSET_AGND", 9: "HEADSET_AGND",
        10: "HS_CODEC_SDOUT", 11: "HS_CODEC_LRCK", 12: "HS_CODEC_BCLK",
        13: "HS_CODEC_MCLK", 14: "HEADSET_AGND", 15: "HEADSET_AGND",
        16: "HEADSET_1V8",
    }.items():
        label_pin_auto(schematic, i2s_to_codec, pin, net_name, 0.54)

    i2s_from_codec = add_symbol(
        schematic,
        "Logic_LevelTranslator:SN74LVC1T45DCK",
        "U906",
        "SN74LVC1T45DCK",
        (330, 520),
        manufacturer="Texas Instruments",
        mpn="SN74LVC1T45DCKR",
    )
    i2s_from_codec.hidden_properties.add("Value")
    for pin, net_name in {
        1: "LOGIC_3V3", 2: "HEADSET_AGND", 3: "HS_SDIN_FROM_CODEC",
        4: "HS_CODEC_SDIN", 5: "HEADSET_AGND", 6: "HEADSET_1V8",
    }.items():
        label_pin_auto(schematic, i2s_from_codec, pin, net_name, 0.56)
    for reference, rail, x in (
        ("C903", "LOGIC_3V3", 45), ("C904", "HEADSET_1V8", 90),
        ("C905", "LOGIC_3V3", 270), ("C906", "HEADSET_1V8", 315),
        ("C907", "LOGIC_3V3", 360), ("C908", "HEADSET_1V8", 405),
    ):
        capacitor = add_audio_passive("Device:C", reference, "100nF", (x, 585))
        label_two_pin_device(schematic, capacitor, rail, "HEADSET_AGND")
    note(schematic, "CM5 audio GPIO is 3.3 V. ES8316 digital pins remain entirely in the translated 1.8 V domain.", (45, 615), 0.80)

    heading(schematic, "3. DEDICATED LOW-NOISE HEADSET RAILS", (45, 650), 1.8)
    ldo_3v3 = add_symbol(
        schematic, "Regulator_Linear:LP5907MFX-3.3", "U910",
        "LP5907MFX-3.3", (145, 700), manufacturer="Texas Instruments", mpn="LP5907MFX-3.3/NOPB",
    )
    for pin, net_name in {1: "SYS_4V0", 2: "HEADSET_AGND", 3: "HEADSET_POWER_EN", 5: "HEADSET_3V3"}.items():
        label_pin_auto(schematic, ldo_3v3, pin, net_name, 0.63)
    schematic.no_connects.add(pin_xy(ldo_3v3, 4))
    ldo_1v8 = add_symbol(
        schematic, "Regulator_Linear:LP5907MFX-1.8", "U911",
        "LP5907MFX-1.8", (300, 700), manufacturer="Texas Instruments", mpn="LP5907MFX-1.8/NOPB",
    )
    for pin, net_name in {1: "HEADSET_3V3", 2: "HEADSET_AGND", 3: "HEADSET_POWER_EN", 5: "HEADSET_1V8"}.items():
        label_pin_auto(schematic, ldo_1v8, pin, net_name, 0.63)
    schematic.no_connects.add(pin_xy(ldo_1v8, 4))
    for reference, value, rail, x in (
        ("C910", "1uF", "SYS_4V0", 80), ("C911", "1uF", "HEADSET_3V3", 205),
        ("C912", "1uF", "HEADSET_3V3", 250), ("C913", "1uF", "HEADSET_1V8", 360),
    ):
        capacitor = add_audio_passive("Device:C", reference, value, (x, 765))
        label_two_pin_device(schematic, capacitor, rail, "HEADSET_AGND")
    ground_bond = add_audio_passive("Device:R", "R900", "0R star bond", (430, 700))
    label_two_pin_device(schematic, ground_bond, "GND", "HEADSET_AGND")
    note(schematic, "R900 is the only headset analog-ground bond. Do not create a second PCB or chassis bond.", (45, 805), 0.80)

    heading(schematic, "4. ES8316 I2S1 HEADSET CODEC", (500, 340), 1.8)
    codec = add_symbol(
        schematic,
        "CM5Carrier:ES8316",
        "U912",
        "ES8316",
        (610, 445),
        "Package_DFN_QFN:QFN-32-1EP_4x4mm_P0.4mm_EP2.9x2.9mm_ThermalVias",
        "Everest Semiconductor",
        "ES8316",
    )
    codec.hidden_properties.add("Value")
    codec_nets = {
        1: "HS_CODEC_I2C_SCL", 2: "HS_CODEC_MCLK", 3: "HEADSET_1V8", 4: "HEADSET_1V8",
        5: "HEADSET_AGND", 6: "HS_CODEC_BCLK", 7: "HS_CODEC_SDOUT", 8: "HS_CODEC_LRCK",
        9: "HS_CODEC_SDIN", 13: "ES8316_CPVSS", 14: "HEADSET_1V8",
        15: "ES8316_CPTOP", 16: "ES8316_CPBOT", 17: "HEADSET_AGND",
        18: "HEADSET_AGND", 19: "HS_CODEC_HP_R", 20: "HS_CODEC_HP_L",
        21: "ES8316_DACVREF", 22: "HEADSET_3V3", 23: "HEADSET_AGND",
        24: "ES8316_ADCVREF", 25: "ES8316_VMID", 26: "ES8316_MICBIAS",
        27: "HS_MIC_INPUT", 31: "ES8316_CE", 32: "HS_CODEC_I2C_SDA",
        33: "HEADSET_AGND",
    }
    for pin, net_name in codec_nets.items():
        label_pin_auto(schematic, codec, pin, net_name, 0.58)
    for pin in (10, 11, 12, 28, 29, 30):
        schematic.no_connects.add(pin_xy(codec, pin))
    for reference, value, net_1, net_2, x, y in (
        ("C914", "10uF", "ES8316_CPVSS", "HEADSET_AGND", 500, 540),
        ("C915", "10uF", "HEADSET_1V8", "HEADSET_AGND", 545, 540),
        ("C916", "10uF", "ES8316_CPTOP", "ES8316_CPBOT", 590, 540),
        ("C917", "10uF", "ES8316_DACVREF", "HEADSET_AGND", 635, 540),
        ("C918", "10uF", "ES8316_ADCVREF", "HEADSET_AGND", 680, 540),
        ("C919", "4.7uF", "ES8316_VMID", "HEADSET_AGND", 725, 540),
    ):
        capacitor = add_audio_passive("Device:C", reference, value, (x, y))
        label_two_pin_device(schematic, capacitor, net_1, net_2)
    for reference, rail, x in (
        ("C926", "HEADSET_1V8", 520), ("C927", "HEADSET_1V8", 565),
        ("C928", "HEADSET_1V8", 610), ("C929", "HEADSET_3V3", 655),
    ):
        capacitor = add_audio_passive("Device:C", reference, "100nF", (x, 595))
        label_two_pin_device(schematic, capacitor, rail, "HEADSET_AGND")
    codec_enable = add_audio_passive("Device:R", "R914", "10k", (720, 595))
    label_two_pin_device(schematic, codec_enable, "HEADSET_1V8", "ES8316_CE")
    note(schematic, "CE is pulled to 1.8 V; all ES8316 digital pins are translated. Decoupling follows the Radxa reference topology.", (500, 630), 0.76)

    heading(schematic, "5. TPA6132A2 HEADPHONE AMPLIFIER", (790, 340), 1.8)
    amplifier = add_symbol(
        schematic,
        "Amplifier_Audio:TPA6132A2RTE",
        "U913",
        "TPA6132A2RTE",
        (875, 445),
        manufacturer="Texas Instruments",
        mpn="TPA6132A2RTER",
    )
    amplifier.hidden_properties.add("Value")
    amplifier_nets = {
        1: "HPAMP_IN_L", 2: "HEADSET_AGND", 3: "HEADSET_AGND", 4: "HPAMP_IN_R",
        5: "HEADSET_HP_R", 6: "HEADSET_3V3", 7: "HEADSET_AGND", 8: "HPAMP_HPVSS",
        9: "HPAMP_CPN", 10: "HEADSET_AGND", 11: "HPAMP_CPP", 12: "HEADSET_3V3",
        13: "HEADSET_AMP_EN", 14: "HEADSET_3V3", 15: "HEADSET_AGND",
        16: "HEADSET_HP_L", 17: "HEADSET_AGND",
    }
    for pin, net_name in amplifier_nets.items():
        label_pin_auto(schematic, amplifier, pin, net_name, 0.59)
    for reference, value, net_1, net_2, x, y in (
        ("C920", "1uF", "HS_CODEC_HP_L", "HPAMP_IN_L", 790, 515),
        ("C921", "1uF", "HS_CODEC_HP_R", "HPAMP_IN_R", 835, 515),
        ("C922", "1uF", "HPAMP_CPP", "HPAMP_CPN", 880, 515),
        ("C923", "2.2uF", "HEADSET_3V3", "HEADSET_AGND", 925, 515),
        ("C924", "2.2uF", "HPAMP_HPVSS", "HEADSET_AGND", 970, 515),
        ("C925", "2.2uF", "HEADSET_3V3", "HEADSET_AGND", 1015, 515),
    ):
        capacitor = add_audio_passive("Device:C", reference, value, (x, y))
        label_two_pin_device(schematic, capacitor, net_1, net_2)
    note(schematic, "G0=1 and G1=0 select 0 dB gain. EN remains software-controlled for pop-free mute.", (790, 575), 0.8)

    heading(schematic, "6. CTIA HEADSET JACK, MIC BIAS, DETECT, AND ESD", (745, 20), 1.75)
    jack = add_symbol(
        schematic,
        "CM5Carrier:Kycon_STX_353K7A_6N",
        "J910",
        "STX-353K7A-6N-KTTR",
        (870, 115),
        manufacturer="Kycon",
        mpn="STX-353K7A-6N-KTTR",
    )
    jack.hidden_properties.add("Value")
    for pin, net_name in {
        1: "HS_MIC_JACK", 2: "HEADSET_AGND", 3: "HEADSET_HP_R",
        4: "HEADSET_HP_L", 5: "HEADSET_AGND", 6: "HS_JACK_DET_N",
    }.items():
        label_pin_auto(schematic, jack, pin, net_name, 0.62)
    mic_bias = add_audio_passive("Device:R", "R910", "2.2k", (790, 205))
    label_two_pin_device(schematic, mic_bias, "ES8316_MICBIAS", "HS_MIC_JACK")
    mic_coupling = add_audio_passive("Device:C", "C930", "100nF", (870, 205))
    label_two_pin_device(schematic, mic_coupling, "HS_MIC_JACK", "HS_MIC_INPUT")
    detect_pullup = add_audio_passive("Device:R", "R911", "330k", (950, 205))
    label_two_pin_device(schematic, detect_pullup, "LOGIC_3V3", "HS_JACK_DET_N")
    jack_esd = add_symbol(
        schematic, "Power_Protection:TPD4E05U06DQA", "U914",
        "TPD4E05U06DQA", (1040, 175), manufacturer="Texas Instruments", mpn="TPD4E05U06DQAR",
    )
    for pin, net_name in {
        1: "HEADSET_HP_L", 2: "HEADSET_HP_R", 4: "HS_MIC_JACK",
        5: "HS_JACK_DET_N", 3: "HEADSET_AGND", 8: "HEADSET_AGND",
    }.items():
        label_pin_auto(schematic, jack_esd, pin, net_name, 0.6)
    for pin in (6, 7, 9, 10):
        schematic.no_connects.add(pin_xy(jack_esd, pin))
    note(schematic, "CTIA contact map: tip=L, ring1=R, ring2=GND, sleeve=MIC. Verify isolated switch polarity on a Kycon sample.", (745, 285), 0.76)
    note(schematic, "J910 is routable for prototype capture, but its drawing-derived footprint is blocked from production until sample/coupon sign-off.", (745, 295), 0.76)

    heading(schematic, "7. HEADSET CONTROL POLICY", (745, 650), 1.65)
    note(schematic, "HEADSET_POWER_EN and HEADSET_AMP_EN default low so the codec and amplifier remain muted during boot.", (745, 670), 0.80)
    for reference, net_name, x in (
        ("R912", "HEADSET_POWER_EN", 820),
        ("R913", "HEADSET_AMP_EN", 900),
    ):
        resistor = add_audio_passive("Device:R", reference, "100k", (x, 715))
        label_two_pin_device(schematic, resistor, net_name, "HEADSET_AGND")
    note(schematic, "Sequence: enable rails, configure ES8316, enable TPA6132A2, then ramp software volume.", (745, 760), 0.80)
    note(schematic, "Firmware must debounce HS_JACK_DET_N and perform pop-free mute before any route change.", (745, 770), 0.80)

    for reference, net_name, x in (
        ("#FLG0904", "HEADSET_AGND", 170),
        ("#FLG0905", "HPAMP_HPVSS", 215),
        ("#FLG0906", "HS_I2C_BIAS", 260),
    ):
        flag = add_symbol(schematic, "power:PWR_FLAG", reference, "PWR_FLAG", (x, 825))
        label_pin_auto(schematic, flag, 1, net_name, 0.64)

    output = folder / f"{name}.kicad_sch"
    save_generated_schematic(schematic, output)
    return output


def build_power_regulators_sheet() -> Path:
    """Capture the protected raw bus, production regulator tree, and rail supervision."""
    folder = ROOT / "CM5-CARRIER"
    name = "Power-Regulators-A1"
    create_project(folder, name)
    schematic = create_schematic(name)
    schematic.set_paper_size("A0")
    schematic.set_title_block(
        title="Radxa CM5 ProComm Carrier - Power Regulators and Sequencing",
        date="2026-08-14",
        rev="A1",
        company="ProComm",
        comments={
            1: "RAW_OUT_LOAD accepts 10.5 V to 30 V after source selection and load telemetry",
            2: "The 27.2 mF / 50 V no-blink bank resides on floor-mounted PWR-SELECT",
            3: "DISPLAY_12V is a fused 12 V / 2.5 A harness branch with no dedicated eFuse",
            4: "Component values are Rev A1 starting values; bench loop, thermal, and transfer validation remain mandatory",
        },
    )

    def add_power_passive(
        lib_id: str,
        reference: str,
        value: str,
        position: tuple[float, float],
    ):
        try:
            manufacturer, mpn = POWER_PASSIVE_PARTS[value]
        except KeyError as error:
            raise RuntimeError(
                f"Power-Regulators-A1 {reference} has no locked production passive for {value}"
            ) from error
        return add_symbol(
            schematic,
            lib_id,
            reference,
            value,
            position,
            manufacturer=manufacturer,
            mpn=mpn,
        )

    heading(schematic, "1. PROTECTED RAW INPUT FROM PWR-SELECT", (35, 18), 1.85)
    add_connector(
        schematic,
        "Connector_Generic:Conn_02x02_Odd_Even",
        "J1100",
        "RAW_POWER_FROM_PWR_SELECT",
        (65, 70),
        {1: "RAW_OUT_LOAD", 2: "RAW_OUT_LOAD", 3: "GND", 4: "GND"},
        MF_2X2,
        "Molex",
        "43045-0412",
        two_row=True,
    )
    note(schematic, "Mates with PWR-SELECT J301; 15 A path; do not mate live.", (35, 105), 0.72)
    raw_hf = add_power_passive("Device:C", "C1109", "1uF 50V X7R", (135, 70))
    label_two_pin_device(schematic, raw_hf, "RAW_OUT_LOAD", "GND")
    note(schematic, "C1109 is local high-frequency input decoupling; the clamped bulk bank is not carried by this suspended PCB.", (95, 125), 0.72)
    note(schematic, "PWR-SELECT C306-C309 provide 27.2 mF plus 660 uF local storage upstream of its load shunt.", (95, 135), 0.72)

    heading(schematic, "2. SYS_4V0 / 12 A SYNCHRONOUS BUCK", (320, 18), 1.85)
    sys_controller = add_symbol(
        schematic,
        "CM5Carrier:LM5146RGY",
        "U1110",
        "LM5146RGYR",
        (405, 100),
        footprint="Package_DFN_QFN:Texas_RGY_R-PVQFN-N20_EP2.05x3.05mm_ThermalVias",
        manufacturer="Texas Instruments",
        mpn="LM5146RGYR",
    )
    sys_pin_map = {
        1: "SYS_EN_UVLO", 2: "SYS_RT", 3: "SYS_SS", 4: "SYS_COMP",
        5: "SYS_FB", 6: "GND", 7: "SYS_SYNCOUT_TP", 8: "GND",
        10: "SYS_4V0_PG", 11: "SYS_ILIM", 12: "GND", 13: "SYS_LO",
        14: "SYS_VCC", 15: "GND", 17: "SYS_BST", 18: "SYS_HO",
        19: "SYS_SW", 20: "RAW_OUT_LOAD",
    }
    for pin, net_name in sys_pin_map.items():
        label_pin_auto(schematic, sys_controller, pin, net_name, 0.52)
    for pin in (9, 16):
        schematic.no_connects.add(pin_xy(sys_controller, pin))
    sys_parts = (
        ("R1110", "Device:R", "40.2k", "SYS_RT", "GND", "Yageo", "RC0603FR-0740K2L", 330, 190),
        ("C1110", "Device:C", "47nF", "SYS_SS", "GND", "Murata", "GRM188R71H473KA61D", 375, 190),
        ("R1111", "Device:R", "20.0k 0.1%", "SYS_4V0", "SYS_FB", "Vishay", "TNPW060320K0BEEA", 420, 190),
        ("R1112", "Device:R", "4.99k 0.1%", "SYS_FB", "GND", "Vishay", "TNPW06034K99BEEA", 465, 190),
        ("R1113", "Device:R", "7.5k", "SYS_COMP", "SYS_COMP_C", "Yageo", "RC0603FR-077K5L", 510, 190),
        ("C1111", "Device:C", "6.8nF", "SYS_COMP_C", "GND", "Murata", "GRM1885C1H682JA01D", 555, 190),
        ("C1112", "Device:C", "150pF C0G", "SYS_COMP", "SYS_FB", "Murata", "GRM1885C1H151JA01D", 510, 230),
        ("R1114", "Device:R", "499R", "SYS_ILIM", "SYS_SW", "Yageo", "RC0603FR-07499RL", 555, 230),
        ("R1115", "Device:R", "69.8k 1%", "RAW_OUT_LOAD", "SYS_EN_UVLO", "Yageo", "RC0603FR-0769K8L", 330, 235),
        ("R1116", "Device:R", "10k 1%", "SYS_EN_UVLO", "GND", "Yageo", "RC0603FR-0710KL", 375, 235),
    )
    for reference, lib_id, value, net_1, net_2, manufacturer, mpn, x, y in sys_parts:
        part = add_symbol(schematic, lib_id, reference, value, (x, y), manufacturer=manufacturer, mpn=mpn)
        label_two_pin_device(schematic, part, net_1, net_2)
    q_sys_high = add_symbol(schematic, "Transistor_FET:Q_NMOS_GSD", "Q1110", "NVMFS6B25NL", (590, 85), manufacturer="onsemi", mpn="NVMFS6B25NLT1G")
    for pin, net_name in {1: "SYS_HO", 2: "SYS_SW", 3: "RAW_OUT_LOAD"}.items():
        label_pin_auto(schematic, q_sys_high, pin, net_name, 0.55)
    q_sys_low = add_symbol(schematic, "Transistor_FET:Q_NMOS_GSD", "Q1111", "FDWS86068-F085", (590, 135), manufacturer="onsemi", mpn="FDWS86068-F085")
    for pin, net_name in {1: "SYS_LO", 2: "GND", 3: "SYS_SW"}.items():
        label_pin_auto(schematic, q_sys_low, pin, net_name, 0.55)
    sys_inductor = add_symbol(schematic, "Device:L", "L1110", "3.3uH / 29.2A Isat", (650, 110), manufacturer="TDK", mpn="SPM10065VC-3R3M-D")
    label_two_pin_device(schematic, sys_inductor, "SYS_SW", "SYS_4V0")
    for index, x in enumerate((620, 650, 680, 710, 740), start=3):
        cap = add_symbol(schematic, "Device:C", f"C111{index}", "47uF 6.3V X7R", (x, 190), manufacturer="Murata", mpn="GCM32ER70J476KE19L")
        label_two_pin_device(schematic, cap, "SYS_4V0", "GND")
    sys_boot = add_symbol(schematic, "Device:C", "C1118", "100nF 25V X7R", (600, 230), manufacturer="Murata", mpn="GRM188R71E104KA01D")
    label_two_pin_device(schematic, sys_boot, "SYS_BST", "SYS_SW")
    sys_vcc = add_symbol(schematic, "Device:C", "C1119", "4.7uF 10V X7R", (645, 230), manufacturer="Murata", mpn="GRM21BR71A475KA73L")
    label_two_pin_device(schematic, sys_vcc, "SYS_VCC", "GND")
    for index, x in enumerate((680, 710, 740, 770), start=4):
        cap = add_symbol(schematic, "Device:C", f"C119{index}", "4.7uF 50V X7R", (x, 230), manufacturer="Murata", mpn="GRM31CR71H475KA12L")
        label_two_pin_device(schematic, cap, "RAW_OUT_LOAD", "GND")
    sys_pg_pullup = add_power_passive("Device:R", "R1117", "10k 1%", (770, 190))
    label_two_pin_device(schematic, sys_pg_pullup, "LOGIC_3V3", "SYS_4V0_PG")
    note(schematic, "4.006 V nominal follows Radxa RK806 guidance. 300 kHz; target loop crossover about 20 kHz.", (320, 280), 0.72)

    heading(schematic, "3. AUX_12V / 8 A FOUR-SWITCH BUCK-BOOST", (775, 18), 1.85)
    aux_controller = add_symbol(
        schematic,
        "CM5Carrier:LM5176PWP_A1",
        "U1120",
        "LM5176PWP",
        (865, 105),
        footprint="Package_SO:HTSSOP-28-1EP_4.4x9.7mm_P0.65mm_EP3.4x9.5mm_ThermalVias",
        manufacturer="Texas Instruments",
        mpn="LM5176PWP",
    )
    aux_pin_map = {
        1: "AUX_EN_UVLO", 2: "RAW_OUT_LOAD", 3: "RAW_OUT_LOAD", 4: "AUX_MODE",
        5: "AUX_DITH", 6: "AUX_RT", 7: "AUX_SLOPE", 8: "AUX_SS",
        9: "AUX_COMP", 10: "GND", 11: "AUX_FB", 12: "AUX_12V",
        13: "AUX_12V", 14: "AUX_12V_PRE", 15: "GND", 16: "AUX_CS",
        17: "AUX_12V_PG", 18: "AUX_SW2", 19: "AUX_HDRV2", 20: "AUX_BOOT2",
        21: "AUX_LDRV2", 22: "GND", 23: "AUX_VCC", 24: "AUX_12V",
        25: "AUX_LDRV1", 26: "AUX_BOOT1", 27: "AUX_HDRV1", 28: "AUX_SW1",
        29: "GND",
    }
    for pin, net_name in aux_pin_map.items():
        label_pin_auto(schematic, aux_controller, pin, net_name, 0.48)
    for index, (reference, gate, source, drain, mpn, x, y) in enumerate(
        (
            ("Q1120", "AUX_HDRV1", "AUX_SW1", "RAW_OUT_LOAD", "CSD18532Q5B", 1010, 65),
            ("Q1121", "AUX_LDRV1", "AUX_CS", "AUX_SW1", "CSD18532Q5B", 1010, 110),
            ("Q1122", "AUX_HDRV2", "AUX_SW2", "AUX_12V_PRE", "CSD17573Q5B", 1070, 65),
            ("Q1123", "AUX_LDRV2", "AUX_CS", "AUX_SW2", "CSD17573Q5B", 1070, 110),
        )
    ):
        mosfet = add_symbol(schematic, "Transistor_FET:Q_NMOS_GSD", reference, mpn, (x, y), manufacturer="Texas Instruments", mpn=mpn)
        for pin, net_name in {1: gate, 2: source, 3: drain}.items():
            label_pin_auto(schematic, mosfet, pin, net_name, 0.48)
    aux_inductor = add_symbol(schematic, "Device:L", "L1120", "4.7uH / 20.9A Isat", (1040, 165), manufacturer="Wurth Elektronik", mpn="74439370047")
    label_two_pin_device(schematic, aux_inductor, "AUX_SW1", "AUX_SW2")
    sense = add_symbol(schematic, "Device:R", "R1120", "6mR 2W", (1100, 165), manufacturer="Susumu", mpn="KRL6432E-M-R006-F-T1")
    label_two_pin_device(schematic, sense, "AUX_12V_PRE", "AUX_12V")
    main_sense = add_symbol(schematic, "Device:R", "R1126", "4mR 6W", (1145, 165), manufacturer="Susumu", mpn="KRL11050-C-R004-F-T1")
    label_two_pin_device(schematic, main_sense, "AUX_CS", "GND")
    aux_parts = (
        ("R1121", "Device:R", "27.4k", "AUX_RT", "GND", 785, 200),
        ("R1122", "Device:R", "93.1k", "AUX_MODE", "GND", 825, 200),
        ("C1120", "Device:C", "220pF", "AUX_SLOPE", "GND", 865, 200),
        ("C1121", "Device:C", "100nF", "AUX_SS", "GND", 905, 200),
        ("R1123", "Device:R", "280k 0.1%", "AUX_12V", "AUX_FB", 945, 200),
        ("R1124", "Device:R", "20k 0.1%", "AUX_FB", "GND", 985, 200),
        ("R1125", "Device:R", "10k", "AUX_COMP", "AUX_COMP_C", 1025, 220),
        ("C1122", "Device:C", "33nF", "AUX_COMP_C", "GND", 1065, 220),
        ("C1123", "Device:C", "560pF", "AUX_COMP", "AUX_FB", 1105, 220),
        ("R1127", "Device:R", "73.2k 1%", "RAW_OUT_LOAD", "AUX_EN_UVLO", 785, 285),
        ("R1128", "Device:R", "10k 1%", "AUX_EN_UVLO", "GND", 825, 285),
        ("C1128", "Device:C", "100nF 25V", "AUX_BOOT1", "AUX_SW1", 980, 285),
        ("C1129", "Device:C", "100nF 25V", "AUX_BOOT2", "AUX_SW2", 1020, 285),
        ("C1200", "Device:C", "4.7uF 10V", "AUX_VCC", "GND", 1060, 285),
        ("C1201", "Device:C", "2.2nF", "AUX_DITH", "GND", 1100, 285),
    )
    for reference, lib_id, value, net_1, net_2, x, y in aux_parts:
        part = add_power_passive(lib_id, reference, value, (x, y))
        label_two_pin_device(schematic, part, net_1, net_2)
    for index, x in enumerate((790, 825, 860, 895), start=4):
        cap = add_symbol(schematic, "Device:C_Polarized", f"C112{index}", "330uF 16V polymer", (x, 250), manufacturer="Panasonic", mpn="16SVP330M")
        label_two_pin_device(schematic, cap, "AUX_12V", "GND")
    for reference, boot_net, x in (("D1120", "AUX_BOOT1", 900), ("D1121", "AUX_BOOT2", 940)):
        diode = add_symbol(schematic, "Device:D_Schottky", reference, "BAT54WS", (x, 285), manufacturer="Diodes Incorporated", mpn="BAT54WS-7-F")
        label_two_pin_device(schematic, diode, "AUX_VCC", boot_net)
    for index, x in enumerate((1075, 1100, 1125, 1150), start=2):
        cap = add_symbol(schematic, "Device:C", f"C120{index}", "4.7uF 50V X7R", (x, 250), manufacturer="Murata", mpn="GRM31CR71H475KA12L")
        label_two_pin_device(schematic, cap, "RAW_OUT_LOAD", "GND")
    aux_pg_pullup = add_power_passive("Device:R", "R1129", "10k 1%", (1145, 285))
    label_two_pin_device(schematic, aux_pg_pullup, "LOGIC_3V3", "AUX_12V_PG")
    note(schematic, "Revised 12 V / 8 A starting design. Recalculate losses and compensation, then verify by Bode plot and full thermal load before routing.", (775, 305), 0.72)

    def place_lm614_stage(
        reference: str,
        part: str,
        rail: str,
        current: str,
        fb_bottom: str,
        pgood: str,
        x: float,
        y: float,
    ) -> None:
        base = int(reference[1:])
        controller = add_symbol(
            schematic,
            f"CM5Carrier:{part}",
            reference,
            part,
            (x, y),
            manufacturer="Texas Instruments",
            mpn=part,
        )
        prefix = rail.replace("_", "")
        for pin, net_name in {
            1: rail, 2: f"{prefix}_VCC", 3: "GND", 4: f"{prefix}_FB",
            5: pgood, 6: f"{prefix}_RT", 7: "RAW_OUT_LOAD", 8: "RAW_OUT_LOAD",
            9: "GND", 10: f"{prefix}_SW", 11: "GND", 12: "RAW_OUT_LOAD",
            13: f"{prefix}_RBOOT", 14: f"{prefix}_CBOOT",
        }.items():
            label_pin_auto(schematic, controller, pin, net_name, 0.46)
        inductor = add_symbol(schematic, "Device:L", f"L{base}", f"4.7uH / {current}", (x + 65, y), manufacturer="Coilcraft", mpn="XAL7070-472MEC")
        label_two_pin_device(schematic, inductor, f"{prefix}_SW", rail)
        rt = add_power_passive("Device:R", f"R{base}", "33.2k / 400kHz", (x - 55, y + 70))
        label_two_pin_device(schematic, rt, f"{prefix}_RT", "GND")
        rtop = add_power_passive("Device:R", f"R{base + 1}", "100k 0.1%", (x, y + 70))
        label_two_pin_device(schematic, rtop, rail, f"{prefix}_FB")
        rbottom = add_power_passive("Device:R", f"R{base + 2}", fb_bottom, (x + 45, y + 70))
        label_two_pin_device(schematic, rbottom, f"{prefix}_FB", "GND")
        boot = add_power_passive("Device:C", f"C{base}", "100nF", (x + 90, y + 70))
        label_two_pin_device(schematic, boot, f"{prefix}_CBOOT", f"{prefix}_SW")
        rboot = add_symbol(schematic, "Device:R", f"R{base + 3}", "0R RBOOT", (x + 135, y + 70), manufacturer="Yageo", mpn="RC0603JR-070RL")
        label_two_pin_device(schematic, rboot, f"{prefix}_RBOOT", f"{prefix}_CBOOT")
        for index, dx in enumerate((-60, -25), start=1):
            cap = add_power_passive("Device:C", f"C{base + index}", "4.7uF 50V", (x + dx, y + 120))
            label_two_pin_device(schematic, cap, "RAW_OUT_LOAD", "GND")
        for index, dx in enumerate((20, 55, 90), start=1):
            cap = add_symbol(schematic, "Device:C", f"C{base + 2 + index}", "47uF X7R", (x + dx, y + 120), manufacturer="Murata", mpn="GRM32ER71A476KE15L")
            label_two_pin_device(schematic, cap, rail, "GND")
        vcc_cap = add_symbol(schematic, "Device:C", f"C{base + 6}", "1uF 10V X7R", (x + 135, y + 120), manufacturer="Murata", mpn="GRM188R71A105KA61D")
        label_two_pin_device(schematic, vcc_cap, f"{prefix}_VCC", "GND")

    heading(schematic, "4. DEDICATED RADIO, NETWORK, AND LOGIC BUCKS", (35, 315), 1.85)
    place_lm614_stage("U1130", "LM61460RJR", "MODEM_3V8_PRE", "15.2A Isat", "35.7k 0.1%", "MODEM_3V8_PRE_PG", 125, 385)
    place_lm614_stage("U1140", "LM61440RJR", "WIFI_3V3_PRE", "15.2A Isat", "43.2k 0.1%", "WIFI_3V3_PRE_PG", 405, 385)
    place_lm614_stage("U1150", "LM61440RJR", "NET_3V3", "15.2A Isat", "43.2k 0.1%", "NET_3V3_PG", 685, 385)
    place_lm614_stage("U1160", "LM61440RJR", "LOGIC_3V3", "15.2A Isat", "43.2k 0.1%", "LOGIC_3V3_PG", 965, 385)
    for reference, pgood, x in (
        ("R1134", "MODEM_3V8_PRE_PG", 105),
        ("R1144", "WIFI_3V3_PRE_PG", 385),
        ("R1154", "NET_3V3_PG", 665),
        ("R1164", "LOGIC_3V3_PG", 945),
    ):
        pullup = add_power_passive("Device:R", reference, "10k 1%", (x, 535))
        label_two_pin_device(schematic, pullup, "LOGIC_3V3", pgood)

    modem_efuse = add_symbol(
        schematic, "CM5Carrier:TPS259827LNRGE", "U1131", "TPS259827LNRGER", (155, 580),
        footprint="Package_DFN_QFN:Texas_RGE0024H_VQFN-24-1EP_4x4mm_P0.5mm_EP2.7x2.7mm_ThermalVias",
        manufacturer="Texas Instruments", mpn="TPS259827LNRGER",
    )
    modem_map = {
        1: "MODEM_3V8_PRE", 2: "MODEM_3V8_PRE", 3: "MODEM_3V8_PRE", 4: "GND", 5: "GND",
        6: "MODEM_POWER_EN", 7: "MODEM_ITIMER", 8: "MODEM_ILIM", 9: "MODEM_IMON",
        10: "GND", 11: "GND", 12: "GND", 13: "MODEM_3V8_PG", 14: "GND",
        15: "MODEM_DVDT", 16: "MODEM_3V8_PRE", 17: "MODEM_3V8", 18: "MODEM_3V8",
        19: "MODEM_3V8", 20: "MODEM_3V8", 21: "MODEM_3V8", 22: "MODEM_3V8",
        23: "MODEM_3V8", 24: "MODEM_3V8",
    }
    for pin, net_name in modem_map.items():
        label_pin_auto(schematic, modem_efuse, pin, net_name, 0.40)
    modem_ilim = add_power_passive("Device:R", "R1138", "255R 1% / 6A ILIM", (65, 650))
    label_two_pin_device(schematic, modem_ilim, "MODEM_ILIM", "GND")
    modem_timer = add_power_passive("Device:C", "C1138", "4.7nF", (110, 650))
    label_two_pin_device(schematic, modem_timer, "MODEM_ITIMER", "GND")
    modem_dvdt = add_power_passive("Device:C", "C1210", "10nF", (155, 650))
    label_two_pin_device(schematic, modem_dvdt, "MODEM_DVDT", "GND")
    modem_enable_pulldown = add_power_passive("Device:R", "R1139", "100k 1%", (155, 690))
    label_two_pin_device(schematic, modem_enable_pulldown, "MODEM_POWER_EN", "GND")
    modem_pg_pullup = add_power_passive("Device:R", "R1137", "10k 1%", (200, 690))
    label_two_pin_device(schematic, modem_pg_pullup, "LOGIC_3V3", "MODEM_3V8_PG")
    for index, x in enumerate((200, 235, 270, 305), start=1):
        cap = add_symbol(schematic, "Device:C_Polarized", f"C{1188 + index}", "330uF 6.3V polymer", (x, 650), manufacturer="Panasonic", mpn="6SVP330M")
        label_two_pin_device(schematic, cap, "MODEM_3V8", "GND")
    note(schematic, "TPS259827L active current limiter; 255 ohm ILIM; 1320 uF local modem bulk; validate weak-signal TX bursts.", (35, 705), 0.70)

    wifi_switch = add_symbol(schematic, "CM5Carrier:TPS22990DML", "U1141", "TPS22990DMLR", (430, 580), manufacturer="Texas Instruments", mpn="TPS22990DMLR")
    for pin, net_name in {
        1: "WIFI_DVDT", 3: "WIFI_3V3_PRE", 4: "LOGIC_3V3", 5: "WIFI_POWER_EN",
        6: "GND", 7: "WIFI_3V3_PG", 8: "WIFI_3V3", 9: "WIFI_3V3", 10: "WIFI_3V3",
    }.items():
        label_pin_auto(schematic, wifi_switch, pin, net_name, 0.44)
    schematic.no_connects.add(pin_xy(wifi_switch, 2))
    wifi_ct = add_power_passive("Device:C", "C1149", "10nF", (375, 650))
    label_two_pin_device(schematic, wifi_ct, "WIFI_DVDT", "GND")
    wifi_enable_pulldown = add_power_passive("Device:R", "R1148", "100k 1%", (420, 690))
    label_two_pin_device(schematic, wifi_enable_pulldown, "WIFI_POWER_EN", "GND")
    wifi_pg_pullup = add_power_passive("Device:R", "R1149", "10k 1%", (465, 690))
    label_two_pin_device(schematic, wifi_pg_pullup, "LOGIC_3V3", "WIFI_3V3_PG")
    wifi_bulk = add_symbol(schematic, "Device:C_Polarized", "C1193", "470uF 6.3V polymer", (475, 650), manufacturer="Panasonic", mpn="6SVP470M")
    label_two_pin_device(schematic, wifi_bulk, "WIFI_3V3", "GND")
    note(schematic, "TPS22990 provides controlled Wi-Fi startup and PG; it is not a current limiter.", (350, 705), 0.70)

    heading(schematic, "5. LOW-NOISE POINT-OF-LOAD RAILS", (555, 555), 1.75)
    pol_specs = (
        ("U1170", "PCIE_1V0", "SYS_4V0", "1.24k", 2.0, 570, 620),
        ("U1171", "LOGIC_1V8", "SYS_4V0", "6.19k", 1.5, 800, 620),
    )
    for reference, rail, input_rail, rtop_value, rating, x, y in pol_specs:
        passive_base = int(reference[1:]) * 10
        pol = add_symbol(
            schematic, "CM5Carrier:TPS62913RPU", reference, "TPS62913RPU", (x, y),
            footprint="Package_DFN_QFN:Texas_RPU0010A_VQFN-HR-10_2x2mm_P0.5mm",
            manufacturer="Texas Instruments", mpn="TPS62913RPUT",
        )
        prefix = rail.replace("_", "")
        for pin, net_name in {
            1: "SYS_4V0_PG", 2: f"{prefix}_SW", 3: rail, 4: "GND", 5: f"{rail}_PG",
            6: input_rail, 7: "GND", 8: f"{prefix}_NRSS", 9: f"{prefix}_FB", 10: f"{prefix}_SCONF",
        }.items():
            label_pin_auto(schematic, pol, pin, net_name, 0.38)
        inductor = add_symbol(schematic, "Device:L", f"L{passive_base}", f"2.2uH / {rating:.1f}A rail", (x + 60, y), manufacturer="Coilcraft", mpn="XGL4030-222MEC")
        label_two_pin_device(schematic, inductor, f"{prefix}_SW", rail)
        rtop = add_power_passive("Device:R", f"R{passive_base}", f"{rtop_value} 0.1%", (x - 45, y + 55))
        label_two_pin_device(schematic, rtop, rail, f"{prefix}_FB")
        rbottom = add_power_passive("Device:R", f"R{passive_base + 1}", "4.99k 0.1%", (x, y + 55))
        label_two_pin_device(schematic, rbottom, f"{prefix}_FB", "GND")
        nrss = add_power_passive("Device:C", f"C{passive_base}", "470nF", (x + 45, y + 55))
        label_two_pin_device(schematic, nrss, f"{prefix}_NRSS", "GND")
        for index, dx in enumerate((-90, -65), start=4):
            cap = add_symbol(schematic, "Device:C", f"C{passive_base + index}", "10uF 25V X7R", (x + dx, y + 55), manufacturer="Murata", mpn="GRM21BR71E106KA73L")
            label_two_pin_device(schematic, cap, input_rail, "GND")
        for index, dx in enumerate((65, 90, 115), start=1):
            cap = add_symbol(schematic, "Device:C", f"C{passive_base + index}", "22uF 10V X7R", (x + dx, y + 55), manufacturer="Murata", mpn="GRM21BR71A226ME44L")
            label_two_pin_device(schematic, cap, rail, "GND")
        sconf = add_symbol(schematic, "Device:R", f"R{passive_base + 2}", "7.5k S-CONF", (x + 135, y - 35), manufacturer="Yageo", mpn="RC0603FR-077K5L")
        label_two_pin_device(schematic, sconf, f"{prefix}_SCONF", "GND")
        pg_pullup = add_power_passive("Device:R", f"R{passive_base + 3}", "10k 1%", (x + 175, y - 35))
        label_two_pin_device(schematic, pg_pullup, "LOGIC_3V3", f"{rail}_PG")
    note(schematic, "TPS62913 uses 2.2 MHz, 470 nF NR/SS and 2.2 uH. Headset LDOs live on Audio-Control; all AKM rails live on AUDIO-8X8.", (555, 820), 0.68)

    heading(schematic, "6. AUX_12V BRANCHES AND AUDIO HANDOFF", (930, 555), 1.75)
    branch_specs = (
        ("F1180", "3A time-lag", "AUX_12V", "DISPLAY_12V", "0453003.MR", 925, 620),
        ("F1181", "3A time-lag", "AUX_12V", "FAN_CPU_12V", "0453003.MR", 975, 620),
        ("F1182", "2A time-lag", "AUX_12V", "AUDIO_12V", "0453002.MR", 1025, 620),
        ("F1183", "0.25A fast", "AUX_12V", "NIGHT_LIGHT_12V", "0453.250MR", 1075, 620),
        ("F1184", "3A time-lag", "AUX_12V", "FAN_AUX_12V", "0453003.MR", 1125, 620),
    )
    for reference, value, net_1, net_2, mpn, x, y in branch_specs:
        fuse = add_symbol(schematic, "Device:Fuse", reference, value, (x, y), manufacturer="Littelfuse", mpn=mpn)
        label_two_pin_device(schematic, fuse, net_1, net_2)
    with isolated_uuid_namespace("display-io-branch-fuse"):
        display_io_fuse = add_symbol(
            schematic, "Device:Fuse", "F1185", "2A time-lag", (1170, 620),
            manufacturer="Littelfuse", mpn="0453002.MR",
        )
        label_two_pin_device(schematic, display_io_fuse, "AUX_12V", "DISPLAY_IO_12V")
        lock_component_pin_uuids(display_io_fuse)
    note(schematic, "DISPLAY_12V: 12 V / 2.5 A harness, simple fuse only. No dedicated monitor eFuse by requirement.", (925, 660), 0.70)
    note(schematic, "NIGHT_LIGHT_12V: independent 0.25 A fused panel branch; no CM5 or software dependency.", (925, 670), 0.70)
    note(schematic, "FAN_CPU_12V and FAN_AUX_12V are separate 3 A branches; every fan also has local resettable protection.", (925, 680), 0.70)
    note(schematic, "AUDIO_12V leaves this board through the controlled inter-board harness. AUDIO-8X8 locally creates +/-15 V, the quiet 5.5 V pre-rail, separate LT3045 ADC/DAC 5 V rails and AKM 3.3 V.", (900, 825), 0.64)

    heading(schematic, "7. SEQUENCING, POWER-GOOD, AND TEST ACCESS", (35, 735), 1.65)
    sequence_notes = (
        "RAW valid -> SYS_4V0, AUX_12V, LOGIC_3V3 and radio pre-regulators start from local UVLO.",
        "SYS_4V0_PG directly enables LOGIC_1V8 and PCIE_1V0; all open-drain PGOOD outputs have 10k LOGIC_3V3 pull-ups.",
        "Wi-Fi and modem final rails default off through 100k pulls; Thermal-IO U1003 asserts them only after checks.",
        "AUDIO_ENABLE controls the AUDIO-8X8 converters; DAC mute stays asserted through rail and clock qualification.",
        "All PGOOD outputs are open drain; control-sheet 10k pull-ups to LOGIC_3V3 and adjacent rail test pads are required.",
    )
    for index, text_line in enumerate(sequence_notes):
        note(schematic, text_line, (35, 760 + index * 11), 0.67)
    for index, (net_name, x) in enumerate(
        (("RAW_OUT_LOAD", 390), ("SYS_4V0", 440), ("AUX_12V", 490), ("MODEM_3V8", 540)),
        start=0,
    ):
        tp = add_symbol(schematic, "Connector:TestPoint", f"TP119{index}", net_name, (x, 780), footprint=TEST_POINT_2MM)
        tp.in_bom = False
        label_pin_auto(schematic, tp, 1, net_name, 0.52)
    for reference, net_name, x in (
        ("TP1194", "MODEM_IMON", 590),
        ("TP1195", "SYS_SYNCOUT_TP", 640),
    ):
        tp = add_symbol(schematic, "Connector:TestPoint", reference, net_name, (x, 780), footprint=TEST_POINT_2MM)
        tp.in_bom = False
        label_pin_auto(schematic, tp, 1, net_name, 0.52)
    for reference, net_name, x in (
        ("#FLG1101", "RAW_OUT_LOAD", 415), ("#FLG1102", "SYS_4V0", 465),
        ("#FLG1103", "AUX_12V", 515), ("#FLG1104", "GND", 565),
    ):
        flag = add_symbol(schematic, "power:PWR_FLAG", reference, "PWR_FLAG", (x, 835))
        label_pin_auto(schematic, flag, 1, net_name, 0.52)

    output = folder / f"{name}.kicad_sch"
    save_generated_schematic(schematic, output)
    return output


def build_thermal_io_sheet() -> Path:
    """Capture the system I2C control plane, four fans, sensors, expanders, and LEDs."""
    folder = ROOT / "CM5-CARRIER"
    name = "Thermal-IO"
    create_project(folder, name)
    schematic = create_schematic(name)
    schematic.set_paper_size("A0")
    schematic.set_title_block(
        title="Radxa CM5 ProComm Carrier - Thermal and GPIO Control",
        date="2026-08-14",
        rev="A1",
        company="ProComm",
        comments={
            1: "EMC2305 controls CPU, modem, intake, and exhaust fans independently",
            2: "PWM pull-ups provide full-speed hardware fallback if the controller or firmware fails",
            3: "TMP117 sensors monitor CM5, modem, and board/power zones at 0x48/0x49/0x4A",
            4: "TCA9535 devices at 0x20/0x21/0x22 capture status and drive safe control outputs",
            5: "Two 12 V diffused warm-white panel lamps use an independent latching capacitive touch switch",
        },
    )

    def add_thermal_passive(
        lib_id: str,
        reference: str,
        value: str,
        position: tuple[float, float],
    ):
        try:
            manufacturer, mpn = THERMAL_PASSIVE_PARTS[value]
        except KeyError as error:
            raise RuntimeError(
                f"Thermal-IO {reference} has no locked production passive for {value}"
            ) from error
        return add_symbol(
            schematic,
            lib_id,
            reference,
            value,
            position,
            manufacturer=manufacturer,
            mpn=mpn,
        )

    heading(schematic, "1. 1.8 V TO 3.3 V SYSTEM I2C CONTROL BUS", (45, 20), 1.85)
    bus_level = add_symbol(
        schematic,
        "Interface:PCA9306D",
        "U1000",
        "PCA9306DP",
        (145, 100),
        manufacturer="NXP",
        mpn="PCA9306DP,118",
    )
    for pin, net_name in {
        1: "GND", 2: "LOGIC_1V8", 3: "SYS_I2C7_SCL", 4: "SYS_I2C7_SDA",
        5: "CTRL_I2C_SDA", 6: "CTRL_I2C_SCL", 7: "CTRL_I2C_BIAS", 8: "CTRL_I2C_BIAS",
    }.items():
        label_pin_auto(schematic, bus_level, pin, net_name, 0.62)
    for index, (rail, net_name, x) in enumerate(
        (("LOGIC_1V8", "SYS_I2C7_SCL", 75), ("LOGIC_1V8", "SYS_I2C7_SDA", 115),
         ("LOGIC_3V3", "CTRL_I2C_SCL", 175), ("LOGIC_3V3", "CTRL_I2C_SDA", 215)),
        start=1,
    ):
        resistor = add_thermal_passive("Device:R", f"R100{index}", "2.2k", (x, 175))
        label_two_pin_device(schematic, resistor, rail, net_name)
    power_alert_pullup = add_thermal_passive("Device:R", "R1005", "4.7k", (255, 175))
    label_two_pin_device(schematic, power_alert_pullup, "LOGIC_3V3", "PWR_MON_ALERT_N")
    i2c_bias = add_thermal_passive("Device:R", "R1006", "200k", (255, 130))
    label_two_pin_device(schematic, i2c_bias, "LOGIC_3V3", "CTRL_I2C_BIAS")
    note(schematic, "The control branch is isolated by its own PCA9306. Recheck total bus capacitance with the separate AUDIO-8X8 branch populated.", (45, 215), 0.8)

    def place_expander(reference: str, address: int, x: float, pin_map: dict[int, str]) -> None:
        address_bits = address - 0x20
        if not 0 <= address_bits <= 7:
            raise ValueError(f"Unsupported TCA9535 address 0x{address:02X}")
        expander = add_symbol(
            schematic,
            "Interface_Expansion:TCA9535PWR",
            reference,
            f"TCA9535PWR_0x{address:02X}",
            (x, 115),
            manufacturer="Texas Instruments",
            mpn="TCA9535PWR",
        )
        common = {
            1: f"{reference}_INT_N",
            2: "LOGIC_3V3" if address_bits & 0b010 else "GND",
            3: "LOGIC_3V3" if address_bits & 0b100 else "GND",
            12: "GND",
            21: "LOGIC_3V3" if address_bits & 0b001 else "GND",
            22: "CTRL_I2C_SCL", 23: "CTRL_I2C_SDA", 24: "LOGIC_3V3",
        }
        for pin, net_name in {**common, **pin_map}.items():
            label_pin_auto(schematic, expander, pin, net_name, 0.56)
        for pin in (*range(4, 12), *range(13, 21)):
            if pin not in pin_map:
                schematic.no_connects.add(pin_xy(expander, pin))
        bypass = add_thermal_passive(
            "Device:C", f"C101{int(reference[-1])}", "100nF", (x, 245)
        )
        label_two_pin_device(schematic, bypass, "LOGIC_3V3", "GND")
        interrupt_tp = add_symbol(
            schematic, "Connector:TestPoint", f"TP{reference[1:]}",
            f"{reference}_INT_N", (x + 45, 245), footprint=TEST_POINT_2MM,
        )
        interrupt_tp.in_bom = False
        label_pin_auto(schematic, interrupt_tp, 1, f"{reference}_INT_N", 0.52)
        note(schematic, f"{reference}: 7-bit address 0x{address:02X}; INT is test-point only and firmware polls status.", (x - 75, 275), 0.70)

    heading(schematic, "2. THREE 16-BIT GPIO EXPANDERS", (290, 20), 1.85)
    place_expander(
        "U1001",
        0x20,
        390,
        {
            4: "CH_24V_N", 5: "CH_BAT_N", 6: "VALID_24V_N", 7: "VALID_BAT_N",
            8: "BAT_LOW_N", 9: "FAN_ALERT_N", 10: "TEMP_ALERT_N", 11: "PWR_MON_ALERT_N",
            13: "LED_PWR_GATE", 14: "LED_BACKUP_GATE", 15: "LED_WIFI_GATE", 16: "LED_CELL_GATE",
            17: "LED_TEMP_GATE", 18: "LED_AUDIO_GATE", 19: "AUDIO_PRESENT_N", 20: "AUD_IRQ_N",
        },
    )
    place_expander(
        "U1002",
        0x21,
        625,
        {
            4: "AUDIO_ENABLE", 5: "AUD_ADC_RST_N", 6: "AUD_DAC_RST_N", 7: "AUD_DAC_MUTE_CMD_N",
            8: "HEADSET_POWER_EN", 9: "HEADSET_AMP_EN", 10: "MODEM_FULL_CARD_POWER_OFF_N",
            11: "MODEM_W_DISABLE1_N", 13: "MODEM_RESET_N", 14: "MODEM_SIM_MUX_SEL_OD",
            15: "WIFI_DISABLE1_N", 16: "WAN2_RESET_LOCAL_N", 17: "LAN1_RESET_LOCAL_N",
            18: "LAN2_RESET_LOCAL_N", 19: "MODEM_WAKE_N", 20: "MODEM_STATUS_LED_N",
        },
    )
    place_expander(
        "U1003",
        0x22,
        860,
        {
            4: "MODEM_POWER_EN", 5: "WIFI_POWER_EN", 6: "SYS_4V0_PG",
            7: "MODEM_3V8_PRE_PG", 8: "MODEM_3V8_PG", 9: "WIFI_3V3_PRE_PG",
            10: "WIFI_3V3_PG", 11: "NET_3V3_PG", 13: "PCIE_1V0_PG",
            14: "LOGIC_1V8_PG", 15: "AUX_12V_PG", 16: "LOGIC_3V3_PG",
            17: "IO_5V0_PG", 18: "VALID_DTAP_N", 19: "VALID_GOLD_N",
        },
    )
    note(schematic, "TCA9535 outputs power up as inputs. External straps define safe reset, mute, radio-power, and amplifier states.", (290, 295), 0.78)
    safe_straps = (
        ("R1010", "100k", "AUDIO_ENABLE", "GND"),
        ("R1011", "100k", "AUD_DAC_MUTE_CMD_N", "GND"),
        ("R1012", "100k", "MODEM_FULL_CARD_POWER_OFF_N", "LOGIC_3V3"),
        ("R1013", "100k", "MODEM_W_DISABLE1_N", "GND"),
        ("R1014", "100k", "MODEM_RESET_N", "LOGIC_3V3"),
        ("R1015", "100k", "WIFI_DISABLE1_N", "GND"),
        ("R1016", "100k", "WAN2_RESET_LOCAL_N", "LOGIC_3V3"),
        ("R1017", "100k", "LAN1_RESET_LOCAL_N", "LOGIC_3V3"),
        ("R1018", "100k", "LAN2_RESET_LOCAL_N", "LOGIC_3V3"),
        ("R1019", "100k", "MODEM_POWER_EN", "GND"),
        ("R1022", "100k", "WIFI_POWER_EN", "GND"),
    )
    for index, (reference, value, net_1, net_2) in enumerate(safe_straps):
        resistor = add_thermal_passive(
            "Device:R", reference, value, (1000 + (index % 3) * 65, 55 + (index // 3) * 45)
        )
        label_two_pin_device(schematic, resistor, net_1, net_2)
    note(schematic, "Radio final rails default off; W_DISABLE/WIFI_DISABLE remain asserted until firmware verifies power and thermal state.", (790, 305), 0.74)

    heading(schematic, "3. EMC2305 FOUR-FAN CONTROL WITH FAILSAFE", (45, 315), 1.9)
    fan_controller = add_symbol(
        schematic,
        "Driver_Motor:EMC2305-x-AP",
        "U1020",
        "EMC2305-1-AP-TR",
        (160, 415),
        manufacturer="Microchip",
        mpn="EMC2305-1-AP-TR",
    )
    for pin, net_name in {
        1: "CTRL_I2C_SDA", 2: "CTRL_I2C_SCL", 3: "LOGIC_3V3",
        4: "EMC2305_ADDR", 5: "CPU_FAN_PWM", 6: "CPU_FAN_TACH",
        7: "MODEM_FAN_PWM", 8: "MODEM_FAN_TACH", 9: "INTAKE_FAN_PWM",
        10: "INTAKE_FAN_TACH", 12: "FAN_ALERT_N", 13: "EXHAUST_FAN_PWM",
        14: "EXHAUST_FAN_TACH", 17: "GND",
    }.items():
        label_pin_auto(schematic, fan_controller, pin, net_name, 0.58)
    for pin in (11, 15, 16):
        schematic.no_connects.add(pin_xy(fan_controller, pin))
    address_resistor = add_thermal_passive("Device:R", "R1020", "4.7k 5%", (160, 510))
    label_two_pin_device(schematic, address_resistor, "LOGIC_3V3", "EMC2305_ADDR")
    alert_pullup = add_thermal_passive("Device:R", "R1021", "4.7k", (225, 510))
    label_two_pin_device(schematic, alert_pullup, "LOGIC_3V3", "FAN_ALERT_N")
    fan_decoupling = add_thermal_passive("Device:C", "C1020", "100nF", (290, 510))
    label_two_pin_device(schematic, fan_decoupling, "LOGIC_3V3", "GND")
    note(schematic, "ADDR_SEL 4.7k to 3.3 V selects 7-bit SMBus address 0x2E; internal clock is used.", (45, 550), 0.8)
    note(schematic, "Enclosure channels: 26.00 kHz base / divide 26 = 1.000 kHz PWM; direct duty with tach monitoring, 100% start and 30% minimum running duty.", (360, 550), 0.72)

    fan_specs = (
        ("J1021", "FFB0412EN-00Y2E CPU FAN", "FAN_CPU_12V", "CPU_FAN_12V", "3.0A hold / 15V", "2920L300/15DR", "CPU_FAN_TACH", "CPU_FAN_PWM", 395),
        ("J1022", "AFB0412SHB-SP04 MODEM FAN", "FAN_AUX_12V", "MODEM_FAN_12V", "1.1A hold / 33V", "1812L110/33DR", "MODEM_FAN_TACH", "MODEM_FAN_PWM", 575),
        ("J1023", "THA0412AD-TZW3 INTAKE", "FAN_AUX_12V", "INTAKE_FAN_12V", "1.1A hold / 33V", "1812L110/33DR", "INTAKE_FAN_TACH", "INTAKE_FAN_PWM", 755),
        ("J1024", "THA0412AD-TZW3 EXHAUST", "FAN_AUX_12V", "EXHAUST_FAN_12V", "1.1A hold / 33V", "1812L110/33DR", "EXHAUST_FAN_TACH", "EXHAUST_FAN_PWM", 935),
    )
    for index, (reference, value, source_rail, fan_rail, fuse_value, fuse_mpn, tach, pwm, x) in enumerate(fan_specs, start=1):
        add_connector(
            schematic, "Connector_Generic:Conn_02x02_Odd_Even", reference, value, (x, 400),
            {1: "GND", 2: fan_rail, 3: tach, 4: pwm},
            MF_2X2,
            "Molex",
            "43045-0412",
            two_row=True,
        )
        branch_fuse = add_symbol(
            schematic,
            "Device:Polyfuse",
            f"F102{index}",
            fuse_value,
            (x - 25, 475),
            manufacturer="Littelfuse",
            mpn=fuse_mpn,
        )
        label_two_pin_device(schematic, branch_fuse, source_rail, fan_rail)
        pwm_pullup = add_thermal_passive("Device:R", f"R103{index}", "10k", (x + 20, 475))
        label_two_pin_device(schematic, pwm_pullup, "LOGIC_3V3", pwm)
        tach_pullup = add_thermal_passive("Device:R", f"R104{index}", "10k", (x + 65, 475))
        label_two_pin_device(schematic, tach_pullup, "LOGIC_3V3", tach)
        note(schematic, value.replace("_", " "), (x - 45, 525), 0.74)
    note(schematic, "Fan 1 is panel intake and fan 2 is panel exhaust. CPU and modem fans circulate locally beneath their dedicated mesh/keepout areas.", (360, 560), 0.82)
    note(schematic, "All four PWM nets pull high when U1020 is absent or unpowered, commanding the selected 4-wire fans to full speed.", (360, 570), 0.82)

    heading(schematic, "4. THREE HIGH-ACCURACY TEMPERATURE SENSORS", (45, 610), 1.85)
    sensor_specs = (
        ("U1050", "TMP117_CM5", "GND", "TEMP_ALERT_N", "0x48", 170),
        ("U1051", "TMP117_MODEM", "LOGIC_3V3", "TEMP_ALERT_N", "0x49", 390),
        ("U1052", "TMP117_BOARD_POWER", "CTRL_I2C_SDA", "TEMP_ALERT_N", "0x4A", 610),
    )
    for index, (reference, value, addr_net, alert_net, address, x) in enumerate(sensor_specs):
        sensor = add_symbol(
            schematic,
            "Sensor_Temperature:TMP117xxDRV",
            reference,
            value,
            (x, 680),
            manufacturer="Texas Instruments",
            mpn="TMP117AIDRVR",
        )
        for pin, net_name in {
            1: "CTRL_I2C_SCL", 2: "GND", 3: alert_net,
            4: addr_net, 5: "LOGIC_3V3", 6: "CTRL_I2C_SDA",
        }.items():
            label_pin_auto(schematic, sensor, pin, net_name, 0.59)
        bypass = add_thermal_passive("Device:C", f"C105{index}", "100nF", (x - 30, 750))
        label_two_pin_device(schematic, bypass, "LOGIC_3V3", "GND")
        if index == 0:
            pullup = add_thermal_passive("Device:R", "R1050", "5.1k", (x + 35, 750))
            label_two_pin_device(schematic, pullup, "LOGIC_3V3", alert_net)
        note(schematic, f"{value.replace('_', ' ')} / {address}", (x - 60, 790), 0.75)
    note(schematic, "The three open-drain TMP117 alerts share TEMP_ALERT_N; firmware reads 0x48/0x49/0x4A to identify the source.", (45, 810), 0.8)

    heading(schematic, "5. SIX 12 V TOP-PANEL STATUS-LAMP DRIVERS", (770, 610), 1.60)
    led_specs = (
        ("PWR", "DX06WG012B", "LED_PWR_GATE", 785),
        ("BACKUP", "DX06WY012B", "LED_BACKUP_GATE", 840),
        ("WIFI", "DX06WB012B", "LED_WIFI_GATE", 895),
        ("CELL", "DX06WW012B", "LED_CELL_GATE", 950),
        ("TEMP", "DX06WR012B", "LED_TEMP_GATE", 1005),
        ("AUDIO", "DX06WG012B", "LED_AUDIO_GATE", 1060),
    )
    for index, (label, panel_mpn, control_net, x) in enumerate(led_specs):
        gate_net = f"STATUS_{label}_GATE"
        sink_net = f"STATUS_{label}_SINK"
        gate_resistor = add_thermal_passive("Device:R", f"R106{index}", "1k", (x, 675))
        label_two_pin_device(schematic, gate_resistor, control_net, gate_net)
        gate_pulldown = add_thermal_passive("Device:R", f"R108{index}", "100k", (x, 720))
        label_two_pin_device(schematic, gate_pulldown, gate_net, "GND")
        driver = add_symbol(
            schematic,
            "Transistor_FET:Q_NMOS_GSD",
            f"Q106{index}",
            "2N7002K-7",
            (x, 770),
            manufacturer="Diodes Incorporated",
            mpn="2N7002K-7",
        )
        for pin, net_name in {1: gate_net, 2: "GND", 3: sink_net}.items():
            label_pin_auto(schematic, driver, pin, net_name, 0.48)
        note(schematic, f"{label} / {panel_mpn}", (x - 22, 812), 0.54)
    add_connector(
        schematic,
        "Connector_Generic:Conn_02x04_Odd_Even",
        "J1060",
        "STATUS_LAMP_PANEL_HARNESS",
        (1135, 735),
        {
            1: "STATUS_LED_12V", 2: "STATUS_LED_12V",
            3: "STATUS_PWR_SINK", 4: "STATUS_BACKUP_SINK",
            5: "STATUS_WIFI_SINK", 6: "STATUS_CELL_SINK",
            7: "STATUS_TEMP_SINK", 8: "STATUS_AUDIO_SINK",
        },
        MF_2X4,
        "Molex",
        "43045-0812",
        two_row=True,
    )
    status_fuse = add_symbol(
        schematic,
        "Device:Fuse",
        "F1060",
        "0.25A fast",
        (1135, 655),
        manufacturer="Littelfuse",
        mpn="0453.250MR",
    )
    label_two_pin_device(schematic, status_fuse, "AUX_12V", "STATUS_LED_12V")
    note(schematic, "J1060 feeds six external Bulgin DX06 12 V wire-lead indicators. Two supply contacts plus six protected low-side returns.", (770, 825), 0.62)

    heading(schematic, "6. INDEPENDENT WARM-WHITE PANEL LIGHTS", (1015, 315), 1.55)
    add_connector(
        schematic,
        "Connector_Generic:Conn_02x04_Odd_Even",
        "J1070",
        "NIGHT_LIGHT_PANEL_HARNESS",
        (1045, 390),
        {
            1: "GND",
            2: "NIGHT_LIGHT_12V",
            3: "CS_GREEN_RING_K",
            4: "NIGHT_LIGHT_SINK",
            5: "NIGHT_LIGHT_12V",
            6: "NIGHT_LIGHT_SINK",
            7: "NIGHT_LIGHT_12V",
            8: "NIGHT_LIGHT_SINK",
        },
        MF_2X4,
        "Molex",
        "43045-0812",
        two_row=True,
    )
    ring_resistor = add_thermal_passive("Device:R", "R1070", "1k ring limit", (1110, 390))
    label_two_pin_device(schematic, ring_resistor, "CS_GREEN_RING_K", "NIGHT_LIGHT_SINK")
    note(schematic, "J1070 is the keyed eight-circuit boundary to external CS7L2FR and two LS102W panel lights. Unused blue/red switch-ring leads are insulated in the harness.", (1010, 475), 0.67)
    note(schematic, "Pins 1-4 serve switch GND, 12 V, green ring, and latching sink. Pins 5-8 are the two lamp +/- pairs; the PCB joins both lamps to the switched sink.", (1010, 505), 0.67)
    note(schematic, "Hardware-only control: lights remain usable if the CM5 is booting, crashed, or powered down while AUX_12V is available.", (1010, 555), 0.70)

    for reference, net_name, x in (
        ("#FLG1001", "LOGIC_3V3", 55), ("#FLG1002", "LOGIC_1V8", 100),
        ("#FLG1003", "FAN_CPU_12V", 145), ("#FLG1004", "FAN_AUX_12V", 190),
        ("#FLG1005", "CTRL_I2C_BIAS", 235),
    ):
        flag = add_symbol(schematic, "power:PWR_FLAG", reference, "PWR_FLAG", (x, 835))
        label_pin_auto(schematic, flag, 1, net_name, 0.64)

    output = folder / f"{name}.kicad_sch"
    save_generated_schematic(schematic, output)
    return output


def build_carrier() -> Path:
    folder = ROOT / "CM5-CARRIER"
    name = "CM5-Carrier"
    create_project(folder, name)

    schematic = create_schematic(name)
    schematic.set_paper_size("A1")
    schematic.set_title_block(
        title="Radxa CM5 ProComm Carrier - Interface Contract",
        date="2026-08-14",
        rev="A1",
        company="ProComm",
        comments={
            1: "Physical CM5-CARRIER root; nested detailed pages form one connected board netlist",
            2: "Fan 1 intake, fan 2 exhaust; independent PWM/tach channels",
            3: "Program audio crosses a buffered differential TDM link",
            4: "Global labels are intentional inter-page nets; prototype qualification gates remain",
        },
    )

    heading(schematic, "A. PWR-SELECT INTERFACE", (60, 22))
    raw_power_representation = add_connector(
        schematic,
        "Connector_Generic:Conn_02x02_Odd_Even",
        "J101",
        "RAW_POWER_FROM_PWR_SELECT",
        (45, 48),
        {1: "RAW_OUT_LOAD", 2: "RAW_OUT_LOAD", 3: "GND", 4: "GND"},
        MF_2X2,
        "Molex",
        "43045-0412",
        two_row=True,
    )
    raw_power_representation.in_bom = False
    raw_power_representation.on_board = False
    note(schematic, "Mates with PWR-SELECT J301; two contacts per polarity; validate 15 A derating.", (105, 67), 0.9)

    add_connector(
        schematic,
        "Connector_Generic:Conn_01x08",
        "J102",
        "PWR_STATUS_FROM_SELECTOR",
        (120, 50),
        {
            1: "GND",
            2: "CH_24V_N",
            3: "CH_BAT_N",
            4: "VALID_24V_N",
            5: "VALID_BAT_N",
            6: "BAT_LOW_N",
            7: "VALID_DTAP_N",
            8: "VALID_GOLD_N",
        },
        PICO_8,
        "Molex",
        "53047-0810",
    )
    note(schematic, "Mates with PWR-SELECT J401; add 3.3 V pull-ups on the carrier.", (93, 78), 0.9)

    add_connector(
        schematic,
        "Connector_Generic:Conn_01x06",
        "J103",
        "PWR_TELEMETRY_FROM_SELECTOR",
        (200, 50),
        {
            1: "LOGIC_3V3",
            2: "GND",
            3: "CTRL_I2C_SDA",
            4: "CTRL_I2C_SCL",
            5: "PWR_MON_ALERT_N",
            6: "GND",
        },
        PICO_6,
        "Molex",
        "53047-0610",
    )
    note(schematic, "Mates with PWR-SELECT J402; carrier supplies 3.3 V and owns I2C pull-ups.", (170, 78), 0.82)

    heading(schematic, "B. BUFFERED DIFFERENTIAL TDM + AUDIO CONTROL", (105, 105))
    tdm_map = {
        1: "AUD_MCLK_P",
        2: "AUD_MCLK_N",
        3: "GND",
        4: "GND",
        5: "AUD_BCLK_P",
        6: "AUD_BCLK_N",
        7: "AUD_FSYNC_P",
        8: "AUD_FSYNC_N",
        9: "GND",
        10: "GND",
        11: "AUD_DAC_SDIN_P",
        12: "AUD_DAC_SDIN_N",
        13: "AUD_ADC_SDOUT_P",
        14: "AUD_ADC_SDOUT_N",
        15: "GND",
        16: "GND",
        17: "AUD_I2C_SCL",
        18: "AUD_I2C_SDA",
        19: "AUD_ADC_RST_N",
        20: "AUD_DAC_RST_N",
        21: "AUD_DAC_MUTE_CMD_N",
        22: "AUD_IRQ_N",
        23: "AUDIO_PRESENT_N",
        24: "AUDIO_ENABLE",
        25: "LOGIC_3V3",
        26: "GND",
        27: "TDM_SPARE_1",
        28: "TDM_SPARE_2",
        29: "GND",
        30: "GND",
    }
    tdm_representation = add_connector(
        schematic,
        "Connector_Generic:Conn_02x15_Odd_Even",
        "J201",
        "AUDIO_TDM_CONTROL",
        (70, 160),
        tdm_map,
        MILLIGRID_2X15,
        "Molex",
        "87832-6423",
        two_row=True,
    )
    tdm_representation.in_bom = False
    tdm_representation.on_board = False
    note(schematic, "Mates with AUDIO-8X8 J101 using Molex 87832-6423 headers; harness assembly remains vibration-qualified.", (115, 208), 0.9)
    note(schematic, "Carrier drives MCLK/BCLK/FSYNC/DAC data; AUDIO-8X8 drives ADC data back.", (105, 213), 0.9)

    add_connector(
        schematic,
        "Connector_Generic:Conn_02x02_Odd_Even",
        "J202",
        "AUDIO_12V_POWER",
        (165, 155),
        {1: "AUDIO_12V", 2: "AUDIO_12V", 3: "GND", 4: "GND"},
        MF_2X2,
        "Molex",
        "43045-0412",
        two_row=True,
    )
    note(schematic, "Separate filtered audio-power harness; no fan/radio loads.", (140, 180), 0.9)

    heading(schematic, "C. FOUR INDEPENDENT PWM / TACH FANS", (90, 235))
    fans = (
        ("J401", "CPU_FAN", 45, "CPU_FAN_12V", "CPU_FAN_TACH", "CPU_FAN_PWM"),
        ("J402", "MODEM_FAN", 115, "MODEM_FAN_12V", "MODEM_FAN_TACH", "MODEM_FAN_PWM"),
        ("J403", "THA0412AD-TZW3_INTAKE", 185, "INTAKE_FAN_12V", "INTAKE_FAN_TACH", "INTAKE_FAN_PWM"),
        ("J404", "THA0412AD-TZW3_EXHAUST", 255, "EXHAUST_FAN_12V", "EXHAUST_FAN_TACH", "EXHAUST_FAN_PWM"),
    )
    for reference, value, x, power, tach, pwm in fans:
        fan_representation = add_connector(
            schematic,
            "Connector_Generic:Conn_02x02_Odd_Even",
            reference,
            value,
            (x, 275),
            {1: "GND", 2: power, 3: tach, 4: pwm},
            MF_2X2,
            "Molex",
            "43045-0412",
            two_row=True,
        )
        fan_representation.in_bom = False
        fan_representation.on_board = False
        note(schematic, value, (x, 294), 0.78)
    note(schematic, "J403 is physically keyed/labeled INTAKE; J404 is keyed/labeled EXHAUST.", (105, 315), 0.95)

    heading(schematic, "D. DETAILED CAPTURE SUITE - REV A1", (320, 22))
    capture_items = (
        "01 CM5-Core-Allocated: exact 300-contact CM5 symbol, 76 owned contacts (74 connected, 2 assigned NC)",
        "02 CM5-Core-Allocated: VCC_SYSIN, reset, recovery, power-on and debug UART",
        "03 Network-PCIe: PI7C9X2G608GP, 3x LAN7430 and native WAN1 front end",
        "04 Network-PCIe: 4x Wurth MagJack and Molex 0679101002 Mini PCIe 4T4R Wi-Fi",
        "05 WWAN-SIM: TE 2199230-3 B-key, USB3/USB2, controls and dual-SIM mux",
        "06 Display-Harness: HDMI, USB touch and dedicated 12 V / 2.5 A branch",
        "07 Audio-Control: I2S0 TDM, I2S1 ES8316 headset and CTIA amplifier path",
        "08 Thermal-IO: I2C translation, GPIO expansion, sensors and four fan channels",
        "09 Power-Regulators-A1: protected raw hold-up, all main rails, sequencing and test points",
        "10 Audio-8x8: all XLR shields bonded to chassis with one controlled AGND bond",
        "11 Review: all three board roots have zero ERC violations; child context is allowlisted",
    )
    for index, item in enumerate(capture_items):
        note(schematic, item, (320, 38 + index * 9), 0.95)
    note(schematic, "DC/DC compensation and magnetics remain a separate power-design calculation milestone.", (320, 150), 1.0)
    note(schematic, "High-speed routing starts only after stackup and connector Z datums are released.", (320, 160), 1.0)
    note(schematic, "The two enclosure fans require an underside baffle to prevent direct intake/exhaust short flow.", (320, 170), 1.0)

    heading(schematic, "E. CONNECTED CM5-CARRIER HIERARCHY", (610, 22), 1.65)
    add_board_child_sheets(
        schematic,
        name,
        (
            ("CM5 Core / Allocated Pins", "CM5-Core-Allocated.kicad_sch"),
            ("Network / PCIe / Wi-Fi", "Network-PCIe.kicad_sch"),
            ("WWAN / Dual SIM", "WWAN-SIM.kicad_sch"),
            ("Display Harness", "Display-Harness.kicad_sch"),
            ("Audio Control / Headset", "Audio-Control.kicad_sch"),
            ("Power Regulators", "Power-Regulators-A1.kicad_sch"),
            ("Thermal / Fans / IO", "Thermal-IO.kicad_sch"),
        ),
        (605, 45),
    )
    note(schematic, "J101/J201/J401-J404 are interface-only duplicates; the detailed-page connectors own the board footprints.", (585, 390), 0.82)

    output = folder / f"{name}.kicad_sch"
    save_generated_schematic(schematic, output)
    return output


def build_audio_tdm_clock_sheet() -> Path:
    """Capture the differential TDM harness termination and local logic."""
    folder = ROOT / "AUDIO-8X8"
    name = "Audio-TDM-Clock"
    create_project(folder, name)
    schematic = create_schematic(name)
    schematic.set_paper_size("A1")
    schematic.set_title_block(
        title="ProComm AUDIO-8X8 - TDM, Clock, and Control Interface",
        date="2026-08-16",
        rev="A1",
        company="ProComm",
        comments={
            1: "Four terminated LVDS receivers carry MCLK, BCLK, FSYNC, and DAC TDM data",
            2: "One LVDS driver returns the ADC TDM data stream to the CM5 carrier",
            3: "87832-6423 pin assignment is locked and matches CM5-CARRIER J901",
            4: "100 ohm differential; interleaved returns; no branch stubs",
        },
    )

    heading(schematic, "1. LOCKED CARRIER HARNESS", (35, 22), 1.8)
    tdm_map = {
        1: "AUD_MCLK_P", 2: "AUD_MCLK_N", 3: "GND", 4: "GND",
        5: "AUD_BCLK_P", 6: "AUD_BCLK_N", 7: "AUD_FSYNC_P", 8: "AUD_FSYNC_N",
        9: "GND", 10: "GND", 11: "AUD_DAC_SDIN_P", 12: "AUD_DAC_SDIN_N",
        13: "AUD_ADC_SDOUT_P", 14: "AUD_ADC_SDOUT_N", 15: "GND", 16: "GND",
        17: "AUD_I2C_SCL", 18: "AUD_I2C_SDA", 19: "AUD_ADC_RST_N",
        20: "AUD_DAC_RST_N", 21: "AUD_DAC_MUTE_CMD_N", 22: "AUD_IRQ_N",
        23: "AUDIO_PRESENT_N", 24: "AUDIO_ENABLE", 25: "LOGIC_3V3", 26: "GND",
        27: "TDM_SPARE_1", 28: "TDM_SPARE_2", 29: "GND", 30: "GND",
    }
    add_connector(
        schematic,
        "Connector_Generic:Conn_02x15_Odd_Even",
        "J101",
        "AUDIO_TDM_CONTROL",
        (115, 115),
        tdm_map,
        MILLIGRID_2X15,
        "Molex",
        "87832-6423",
        two_row=True,
    )
    note(schematic, "Harness length and pair assignment are controlled by docs/audio_tdm_harness_a1.csv.", (35, 205), 0.85)

    heading(schematic, "2. TERMINATED LVDS RECEIVERS", (260, 22), 1.8)
    receiver_specs = (
        ("U101", "MCLK", "AUD_MCLK_N", "AUD_MCLK_P", "AKM_MCLK", 250),
        ("U102", "BCLK", "AUD_BCLK_N", "AUD_BCLK_P", "AKM_BCLK", 390),
        ("U103", "FSYNC", "AUD_FSYNC_N", "AUD_FSYNC_P", "AKM_FSYNC", 530),
        ("U104", "DAC_SDIN", "AUD_DAC_SDIN_N", "AUD_DAC_SDIN_P", "DAC_TDM_IN", 670),
    )
    for reference, signal, net_n, net_p, local_net, x in receiver_specs:
        receiver = add_symbol(
            schematic,
            "Interface:SN65LVDT2D",
            reference,
            "SN65LVDT2D",
            (x, 105),
            manufacturer="Texas Instruments",
            mpn="SN65LVDT2DR",
        )
        for pin, net_name in {1: net_n, 2: net_p, 5: "AGND", 7: local_net, 8: "AKM_3V3_D"}.items():
            label_pin_auto(schematic, receiver, pin, net_name, 0.58)
        for pin in (3, 4, 6):
            schematic.no_connects.add(pin_xy(receiver, pin))
        capacitor = add_audio8_passive(schematic, "Device:C", f"C{reference[1:]}", "100nF", (x, 165))
        label_two_pin_device(schematic, capacitor, "AKM_3V3_D", "AGND")
        note(schematic, f"{signal}: integrated 110 ohm termination", (x - 45, 190), 0.78)

    heading(schematic, "3. ADC RETURN LVDS DRIVER", (260, 245), 1.8)
    driver = add_symbol(
        schematic,
        "Interface:SN65LVDS1D",
        "U105",
        "SN65LVDS1D",
        (350, 310),
        manufacturer="Texas Instruments",
        mpn="SN65LVDS1DR",
    )
    for pin, net_name in {
        1: "AKM_3V3_D", 2: "ADC_TDM_OUT", 4: "AGND",
        7: "AUD_ADC_SDOUT_P", 8: "AUD_ADC_SDOUT_N",
    }.items():
        label_pin_auto(schematic, driver, pin, net_name, 0.60)
    for pin in (3, 5, 6):
        schematic.no_connects.add(pin_xy(driver, pin))
    for reference, value, x in (("C105", "100nF", 300), ("C106", "4.7uF", 405)):
        capacitor = add_audio8_passive(schematic, "Device:C", reference, value, (x, 380))
        label_two_pin_device(schematic, capacitor, "AKM_3V3_D", "AGND")

    heading(schematic, "4. CONTROL DEFAULTS AND POWER STATE", (525, 245), 1.8)
    controls = (
        ("R111", "AUD_ADC_RST_N", 500),
        ("R112", "AUD_DAC_RST_N", 590),
        ("R113", "AUD_DAC_MUTE_CMD_N", 680),
        ("R114", "AUDIO_PRESENT_N", 770),
    )
    for reference, net_name, x in controls:
        resistor = add_audio8_passive(schematic, "Device:R", reference, "10k 0.1%", (x, 310))
        label_two_pin_device(schematic, resistor, net_name, "AGND")
    for reference, net_name, x in (("R115", "AUD_I2C_SCL", 580), ("R116", "AUD_I2C_SDA", 700)):
        resistor = add_audio8_passive(schematic, "Device:R", reference, "4.7k 0.1%", (x, 390))
        label_two_pin_device(schematic, resistor, "AKM_3V3_D", net_name)
    heading(schematic, "5. HARDWARE FAIL-SILENT GATE", (525, 430), 1.6)
    safe_gate = add_symbol(
        schematic,
        "74xGxx:74LVC1G11",
        "U106",
        "SN74LVC1G11DBVR",
        (650, 515),
        manufacturer="Texas Instruments",
        mpn="SN74LVC1G11DBVR",
    )
    for pin, net_name in {
        1: "AUD_DAC_MUTE_CMD_N", 2: "AGND", 3: "ADC_5V_PG",
        4: "AUDIO_SAFE_UNMUTE_N", 5: "AKM_3V3_D", 6: "DAC_5V_PG",
    }.items():
        label_pin_auto(schematic, safe_gate, pin, net_name, 0.54)
    safe_pulldown = add_audio8_passive(
        schematic, "Device:R", "R117", "100k 1%", (745, 500)
    )
    label_two_pin_device(schematic, safe_pulldown, "AUDIO_SAFE_UNMUTE_N", "AGND")
    safe_bypass = add_audio8_passive(
        schematic, "Device:C", "C107", "100nF", (745, 545)
    )
    label_two_pin_device(schematic, safe_bypass, "AKM_3V3_D", "AGND")
    external_command_flag = add_symbol(
        schematic, "power:PWR_FLAG", "#FLG0101", "PWR_FLAG", (535, 525)
    )
    label_pin_auto(schematic, external_command_flag, 1, "AUD_DAC_MUTE_CMD_N", 0.54)
    note(schematic, "U106 permits DAC unmute and relay closure only when the carrier command, ADC_5V_PG, and DAC_5V_PG are all high.", (525, 585), 0.78)
    note(schematic, "Any power-good loss forces AUDIO_SAFE_UNMUTE_N low in hardware; R117 keeps mute/relays asserted if AKM_3V3_D is absent.", (525, 598), 0.78)
    note(schematic, "SN74LVC1G11 inputs tolerate the 5 V PG pull-ups and Ioff prevents back-powering while its 3.3 V supply is off.", (525, 611), 0.78)
    note(schematic, "Route AKM_MCLK, AKM_BCLK, AKM_FSYNC, DAC_TDM_IN, and ADC_TDM_OUT as short referenced single-ended traces after the translators.", (525, 624), 0.78)

    output = folder / f"{name}.kicad_sch"
    save_generated_schematic(schematic, output)
    return output


def build_ak5558_adc_sheet() -> Path:
    """Capture every AK5558 pin, mode strap, reference, and supply bypass."""
    folder = ROOT / "AUDIO-8X8"
    name = "AK5558-ADC"
    create_project(folder, name)
    schematic = create_schematic(name)
    schematic.set_paper_size("A1")
    schematic.set_title_block(
        title="ProComm AUDIO-8X8 - AK5558VN Eight-Channel ADC",
        date="2026-08-16",
        rev="A1",
        company="ProComm",
        comments={
            1: "TDM256, 32-bit I2S, slave mode: TDM1:0=10, DIF1:0=11, MSN=0",
            2: "AVDD=5.0 V, TVDD=3.3 V, internal 1.8 V LDO enabled",
            3: "Each VREFH uses 20 ohm plus 100 nF and 100 uF to its VREFL/AGND",
            4: "QFN exposed-pad geometry and assembly coupon remain a routing release gate",
        },
    )
    heading(schematic, "1. AK5558VN EXACT PIN CAPTURE", (35, 22), 1.8)
    adc = add_symbol(
        schematic,
        "CM5Carrier:AK5558VN",
        "U201",
        "AK5558VN",
        (250, 180),
        manufacturer="Asahi Kasei Microdevices",
        mpn="AK5558VN",
    )
    adc_pin_nets = {
        1: "AGND", 2: "ADC_5V_A", 3: "ADC_CH3P", 4: "ADC_CH3N",
        5: "AGND", 6: "ADC_VREFH2", 7: "ADC_CH4N", 8: "ADC_CH4P",
        9: "ADC_CH5P", 10: "ADC_CH5N", 11: "ADC_VREFH3", 12: "AGND",
        13: "ADC_CH6N", 14: "ADC_CH6P", 15: "ADC_5V_A", 16: "AGND",
        17: "ADC_CH7P", 18: "ADC_CH7N", 19: "ADC_VREFH4", 20: "AGND",
        21: "ADC_CH8N", 22: "ADC_CH8P", 23: "AGND", 24: "AKM_MCLK",
        25: "AKM_3V3_D", 26: "AGND", 27: "ADC_VDD18", 28: "AUD_ADC_RST_N",
        29: "ADC_PW0", 30: "ADC_PW1", 31: "ADC_PW2", 32: "ADC_MSN",
        33: "AKM_BCLK", 34: "AKM_FSYNC", 35: "AGND", 36: "ADC_TDM_OUT",
        42: "AUD_IRQ_N",
        43: "AUD_I2C_SDA", 44: "ADC_CAD0", 45: "AUD_I2C_SCL", 46: "ADC_CAD1",
        47: "ADC_SLOW", 48: "ADC_SD_PMOD", 49: "ADC_DIF0", 50: "ADC_DIF1",
        51: "ADC_TDM0", 52: "ADC_TDM1", 53: "ADC_PSN", 54: "ADC_I2C_MODE",
        55: "ADC_DP", 56: "ADC_HPFE", 57: "ADC_LDOE", 58: "ADC_ODP",
        59: "ADC_CH1P", 60: "ADC_CH1N", 61: "AGND", 62: "ADC_VREFH1",
        63: "ADC_CH2N", 64: "ADC_CH2P", 65: "AGND",
    }
    for pin, net_name in adc_pin_nets.items():
        label_pin_auto(schematic, adc, pin, net_name, 0.48)
    for pin in (37, 38, 39, 40, 41):
        schematic.no_connects.add(pin_xy(adc, pin))

    heading(schematic, "2. MODE STRAPS", (520, 22), 1.8)
    high_straps = ("ADC_PW0", "ADC_PW1", "ADC_PW2", "ADC_DIF0", "ADC_DIF1", "ADC_TDM1", "ADC_I2C_MODE", "ADC_LDOE")
    low_straps = ("ADC_MSN", "ADC_CAD0", "ADC_CAD1", "ADC_SLOW", "ADC_SD_PMOD", "ADC_TDM0", "ADC_PSN", "ADC_DP", "ADC_HPFE", "ADC_ODP")
    for index, net_name in enumerate(high_straps, start=1):
        x = 515 + ((index - 1) % 4) * 90
        y = 80 + ((index - 1) // 4) * 70
        resistor = add_audio8_passive(schematic, "Device:R", f"R{220 + index}", "10k 0.1%", (x, y))
        label_two_pin_device(schematic, resistor, "AKM_3V3_D", net_name)
    for index, net_name in enumerate(low_straps, start=1):
        x = 515 + ((index - 1) % 5) * 70
        y = 235 + ((index - 1) // 5) * 70
        resistor = add_audio8_passive(schematic, "Device:R", f"R{240 + index}", "10k 0.1%", (x, y))
        label_two_pin_device(schematic, resistor, net_name, "AGND")
    note(schematic, "Straps establish the safe hardware default; firmware still writes and verifies the serial-control registers after reset.", (520, 390), 0.82)

    heading(schematic, "3. REFERENCES AND LOCAL BYPASS", (35, 390), 1.8)
    for index in range(1, 5):
        x = 45 + (index - 1) * 190
        resistor = add_audio8_passive(schematic, "Device:R", f"R{260 + index}", "20R", (x, 450))
        label_two_pin_device(schematic, resistor, "ADC_5V_A", f"ADC_VREFH{index}")
        small = add_audio8_passive(schematic, "Device:C", f"C{260 + index}", "100nF", (x + 55, 450))
        label_two_pin_device(schematic, small, f"ADC_VREFH{index}", "AGND")
        bulk = add_audio8_passive(schematic, "Device:C_Polarized", f"C{270 + index}", "100uF 6.3V polymer", (x + 105, 450))
        label_two_pin_device(schematic, bulk, f"ADC_VREFH{index}", "AGND")
        note(schematic, f"VREF pair {index}", (x, 505), 0.78)
    for reference, value, rail, x in (
        ("C281", "4.7uF", "ADC_VDD18", 55),
        ("C282", "100nF", "ADC_5V_A", 165),
        ("C283", "10uF", "ADC_5V_A", 265),
        ("C284", "100nF", "AKM_3V3_D", 365),
        ("C285", "10uF", "AKM_3V3_D", 465),
        ("C286", "100nF", "ADC_5V_A", 565),
        ("C287", "10uF", "ADC_5V_A", 665),
    ):
        capacitor = add_audio8_passive(schematic, "Device:C", reference, value, (x, 570))
        label_two_pin_device(schematic, capacitor, rail, "AGND")
    note(schematic, "Place C282/C283 at AVDD1, C286/C287 at AVDD2, and C284/C285 at TVDD. C281 is the required 4.7 uF VDD18 stabilizer.", (35, 650), 0.86)
    note(schematic, "Input full-scale qualification target: +24 dBu at XLR maps to 2.49 Vpp differential at the ADC, approximately 1.0 dB below 2.8 Vpp typical full scale.", (35, 665), 0.86)

    output = folder / f"{name}.kicad_sch"
    save_generated_schematic(schematic, output)
    return output


def build_ak4458_dac_sheet() -> Path:
    """Capture every AK4458 pin, serial-control mode, references, and bypass."""
    folder = ROOT / "AUDIO-8X8"
    name = "AK4458-DAC"
    create_project(folder, name)
    schematic = create_schematic(name)
    schematic.set_paper_size("A1")
    schematic.set_title_block(
        title="ProComm AUDIO-8X8 - AK4458VN Eight-Channel DAC",
        date="2026-08-16",
        rev="A1",
        company="ProComm",
        comments={
            1: "TDM256, 32-bit I2S, slave mode is programmed and read back over I2C",
            2: "AVDD=5.0 V, TVDD=3.3 V, internal 1.8 V LDO enabled",
            3: "DAC mute remains asserted until rails, clocks, reset, and register readback pass",
            4: "QFN exposed-pad geometry and assembly coupon remain a routing release gate",
        },
    )
    heading(schematic, "1. AK4458VN EXACT PIN CAPTURE", (35, 22), 1.8)
    dac = add_symbol(
        schematic,
        "CM5Carrier:AK4458VN",
        "U301",
        "AK4458VN",
        (250, 170),
        manufacturer="Asahi Kasei Microdevices",
        mpn="AK4458VN",
    )
    dac_pin_nets = {
        1: "AKM_MCLK", 2: "AKM_BCLK", 3: "AKM_FSYNC", 4: "DAC_TDM_IN",
        5: "AGND", 6: "AGND", 7: "AGND", 8: "AGND", 9: "AGND", 10: "AGND",
        11: "AUDIO_SAFE_UNMUTE_N", 12: "DAC_CAD1", 13: "AUD_I2C_SDA", 14: "AUD_I2C_SCL",
        15: "DAC_CAD0", 16: "DAC_PS", 17: "DAC_I2C_MODE",
        18: "DAC_CH1P", 19: "DAC_CH1N", 20: "AGND", 21: "DAC_VREFH1",
        22: "DAC_CH2N", 23: "DAC_CH2P", 24: "DAC_CH3P", 25: "DAC_CH3N",
        26: "AGND", 27: "DAC_VREFH2", 28: "DAC_CH4N", 29: "DAC_CH4P",
        30: "AGND", 31: "DAC_5V_A", 32: "DAC_CH5P", 33: "DAC_CH5N",
        34: "DAC_VREFH3", 35: "AGND", 36: "DAC_CH6N", 37: "DAC_CH6P",
        38: "DAC_CH7P", 39: "DAC_CH7N", 40: "DAC_VREFH4", 41: "AGND",
        42: "DAC_CH8N", 43: "DAC_CH8P", 44: "DAC_LDOE", 45: "AKM_3V3_D",
        46: "AGND", 47: "DAC_VDD18", 48: "AUD_DAC_RST_N", 49: "AGND",
    }
    for pin, net_name in dac_pin_nets.items():
        label_pin_auto(schematic, dac, pin, net_name, 0.48)

    heading(schematic, "2. SERIAL-CONTROL DEFAULTS", (520, 22), 1.8)
    for index, net_name in enumerate(("DAC_I2C_MODE", "DAC_LDOE", "DAC_CAD0"), start=1):
        resistor = add_audio8_passive(schematic, "Device:R", f"R{320 + index}", "10k 0.1%", (545 + index * 95, 90))
        label_two_pin_device(schematic, resistor, "AKM_3V3_D", net_name)
    for index, net_name in enumerate(("DAC_CAD1", "DAC_PS"), start=1):
        resistor = add_audio8_passive(schematic, "Device:R", f"R{330 + index}", "10k 0.1%", (510 + index * 90, 190))
        label_two_pin_device(schematic, resistor, net_name, "AGND")
    note(schematic, "CAD1:0=01 selects DAC address 0x11; ADC remains at 0x10. TDM256 and 32-bit I2S are written only while mute is asserted.", (520, 275), 0.84)
    note(schematic, "The unused SDTI2/3/4 and DSD pins are tied low in PCM TDM mode to prevent floating inputs.", (520, 290), 0.84)

    heading(schematic, "3. REFERENCES AND LOCAL BYPASS", (35, 365), 1.8)
    for index in range(1, 5):
        x = 45 + (index - 1) * 190
        resistor = add_audio8_passive(schematic, "Device:R", f"R{350 + index}", "10R", (x, 430))
        label_two_pin_device(schematic, resistor, "DAC_5V_A", f"DAC_VREFH{index}")
        small = add_audio8_passive(schematic, "Device:C", f"C{350 + index}", "100nF", (x + 55, 430))
        label_two_pin_device(schematic, small, f"DAC_VREFH{index}", "AGND")
        bulk = add_audio8_passive(schematic, "Device:C_Polarized", f"C{360 + index}", "220uF 6.3V polymer", (x + 105, 430))
        label_two_pin_device(schematic, bulk, f"DAC_VREFH{index}", "AGND")
        note(schematic, f"VREF pair {index}", (x, 485), 0.78)
    for reference, value, rail, x in (
        ("C371", "1uF", "DAC_VDD18", 70),
        ("C372", "100nF", "DAC_5V_A", 200),
        ("C373", "10uF", "DAC_5V_A", 330),
        ("C374", "100nF", "AKM_3V3_D", 460),
        ("C375", "10uF", "AKM_3V3_D", 590),
    ):
        capacitor = add_audio8_passive(schematic, "Device:C", reference, value, (x, 550))
        label_two_pin_device(schematic, capacitor, rail, "AGND")
    note(schematic, "Each VREFH branch uses the AK4458 optional 10 ohm / 220 uF low-impedance filter. Place AVDD, TVDD, and VDD18 bypass parts at their pins.", (35, 625), 0.86)
    note(schematic, "Output level budget reserves at least 0.22 dB at +24 dBu into 600 ohm using minimum AK4458/THAT gains and 0.1% resistors.", (35, 640), 0.86)

    output = folder / f"{name}.kicad_sch"
    save_generated_schematic(schematic, output)
    return output


def build_audio_inputs_sheet() -> Path:
    """Capture eight identical protected active-balanced ADC input channels."""
    folder = ROOT / "AUDIO-8X8"
    name = "Audio-Inputs"
    create_project(folder, name)
    schematic = create_schematic(name)
    schematic.set_paper_size("A0")
    schematic.set_title_block(
        title="ProComm AUDIO-8X8 - Eight Balanced Input Channels",
        date="2026-08-16",
        rev="A1",
        company="ProComm",
        comments={
            1: "THAT1206 -6 dB receiver; OPA1652 single-supply differential ADC driver",
            2: "+4 dBu nominal, +24 dBu maximum, no phantom power",
            3: "13 k/1 k network maps +24 dBu to 2.49 Vpp differential at AK5558",
            4: "Protection and filter values require simulation plus surge/RFI bench qualification",
        },
    )
    heading(schematic, "EIGHT IDENTICAL INPUT CHANNELS - SIGNAL FLOWS LEFT TO RIGHT", (300, 20), 1.8)
    note(schematic, "XLR -> 100R/RFI/fault clamps -> THAT1206 -> 10u bipolar -> 13k/1k level shift -> OPA1652 -> 220R/3.3nF -> AK5558", (300, 34), 0.86)
    note(schematic, "Every XLR pin 1 and shell bonds to XLR_CHASSIS at the connector. AGND and chassis meet only at the controlled board-entry bond.", (350, 45), 0.86)

    for channel in range(1, 9):
        y = 100 + (channel - 1) * 86
        heading(schematic, f"CH{channel} INPUT", (25, y - 31), 1.25)
        base = 4000 + channel * 30
        r_hot = add_audio8_passive(schematic, "Device:R", f"R{base}", "100R", (70, y - 14))
        label_two_pin_device(schematic, r_hot, f"AIN_CH{channel}_HOT", f"AIN{channel}_PROT_P")
        r_cold = add_audio8_passive(schematic, "Device:R", f"R{base + 1}", "100R", (70, y + 14))
        label_two_pin_device(schematic, r_cold, f"AIN_CH{channel}_COLD", f"AIN{channel}_PROT_N")

        c_hot = add_audio8_passive(schematic, "Device:C", f"C{base}", "470pF C0G", (125, y - 14))
        label_two_pin_device(schematic, c_hot, f"AIN{channel}_PROT_P", "XLR_CHASSIS")
        c_cold = add_audio8_passive(schematic, "Device:C", f"C{base + 1}", "470pF C0G", (125, y + 14))
        label_two_pin_device(schematic, c_cold, f"AIN{channel}_PROT_N", "XLR_CHASSIS")
        c_diff = add_audio8_passive(schematic, "Device:C", f"C{base + 2}", "100pF C0G", (175, y))
        label_two_pin_device(schematic, c_diff, f"AIN{channel}_PROT_P", f"AIN{channel}_PROT_N")

        clamp_nets = (
            (f"AIN{channel}_PROT_P", "AUDIO_P15V"),
            ("AUDIO_N15V", f"AIN{channel}_PROT_P"),
            (f"AIN{channel}_PROT_N", "AUDIO_P15V"),
            ("AUDIO_N15V", f"AIN{channel}_PROT_N"),
        )
        for index, (net_1, net_2) in enumerate(clamp_nets):
            diode = add_audio8_passive(
                schematic,
                "Device:D",
                f"D{base + index}",
                "1N4148W",
                (215 + (index % 2) * 38, y - 14 + (index // 2) * 28),
            )
            label_two_pin_device(schematic, diode, net_1, net_2)
        for index, net_name in enumerate((f"AIN{channel}_PROT_P", f"AIN{channel}_PROT_N")):
            zener = add_audio8_passive(schematic, "Device:D_Zener", f"D{base + 4 + index}", "12V zener", (295, y - 14 + index * 28))
            label_two_pin_device(schematic, zener, net_name, "AGND")

        receiver = add_symbol(
            schematic,
            "CM5Carrier:THAT1206",
            f"U{400 + channel}",
            "THAT1206S08-U",
            (365, y),
            manufacturer="THAT Corporation",
            mpn="THAT1206S08-U",
        )
        for pin, net_name in {
            1: "AGND", 2: f"AIN{channel}_PROT_N", 3: f"AIN{channel}_PROT_P",
            4: "AUDIO_N15V", 5: "AGND", 6: f"AIN{channel}_RX_SE", 7: "AUDIO_P15V",
        }.items():
            label_pin_auto(schematic, receiver, pin, net_name, 0.47)
        schematic.no_connects.add(pin_xy(receiver, 8))
        for offset, rail, x in (
            (6, "AUDIO_P15V", 335),
            (7, "AUDIO_N15V", 395),
        ):
            bypass = add_audio8_passive(
                schematic, "Device:C", f"C{base + offset}", "100nF", (x, y + 32)
            )
            label_two_pin_device(schematic, bypass, rail, "AGND")

        coupling = add_audio8_passive(schematic, "Device:C", f"C{base + 3}", "10uF bipolar", (430, y))
        label_two_pin_device(schematic, coupling, f"AIN{channel}_RX_SE", f"AIN{channel}_AC")
        r_div_in = add_audio8_passive(schematic, "Device:R", f"R{base + 6}", "13k 0.1%", (485, y - 18))
        label_two_pin_device(schematic, r_div_in, f"AIN{channel}_AC", f"AIN{channel}_DIVP")
        r_div_ref = add_audio8_passive(schematic, "Device:R", f"R{base + 7}", "1k 0.1%", (535, y - 18))
        label_two_pin_device(schematic, r_div_ref, f"AIN{channel}_DIVP", "VCM_2V5")
        r_inv_in = add_audio8_passive(schematic, "Device:R", f"R{base + 8}", "13k 0.1%", (485, y + 18))
        label_two_pin_device(schematic, r_inv_in, f"AIN{channel}_AC", f"AIN{channel}_INVN")
        r_inv_fb = add_audio8_passive(schematic, "Device:R", f"R{base + 9}", "1k 0.1%", (535, y + 18))
        label_two_pin_device(schematic, r_inv_fb, f"ADC_CH{channel}N_RAW", f"AIN{channel}_INVN")

        opamp = add_symbol(
            schematic,
            "CM5Carrier:OPA1652",
            f"U{420 + channel}",
            "OPA1652AIDR",
            (625, y),
            manufacturer="Texas Instruments",
            mpn="OPA1652AIDR",
        )
        for pin, net_name in {
            1: f"ADC_CH{channel}P_RAW", 2: f"ADC_CH{channel}P_RAW", 3: f"AIN{channel}_DIVP",
            4: "AGND", 5: "VCM_2V5", 6: f"AIN{channel}_INVN",
            7: f"ADC_CH{channel}N_RAW", 8: "ADC_5V_A",
        }.items():
            label_pin_auto(schematic, opamp, pin, net_name, 0.44)

        r_adc_p = add_audio8_passive(schematic, "Device:R", f"R{base + 10}", "220R", (715, y - 14))
        label_two_pin_device(schematic, r_adc_p, f"ADC_CH{channel}P_RAW", f"ADC_CH{channel}P")
        r_adc_n = add_audio8_passive(schematic, "Device:R", f"R{base + 11}", "220R", (715, y + 14))
        label_two_pin_device(schematic, r_adc_n, f"ADC_CH{channel}N_RAW", f"ADC_CH{channel}N")
        anti_alias = add_audio8_passive(schematic, "Device:C", f"C{base + 4}", "3.3nF C0G", (780, y))
        label_two_pin_device(schematic, anti_alias, f"ADC_CH{channel}P", f"ADC_CH{channel}N")
        decoupling = add_audio8_passive(schematic, "Device:C", f"C{base + 5}", "100nF", (850, y))
        label_two_pin_device(schematic, decoupling, "ADC_5V_A", "AGND")
        note(schematic, "+24 dBu -> 2.49 Vpp diff", (910, y - 10), 0.72)
        note(schematic, "2.5 V common mode", (910, y + 8), 0.72)

    note(schematic, "INPUT RELEASE GATES: SPICE stability/noise, THAT fault-clamp review, +24 dBu THD+N, RF injection, ESD/surge, and phantom-fault tests on every connector type.", (400, 790), 0.82)
    output = folder / f"{name}.kicad_sch"
    save_generated_schematic(schematic, output)
    return output


def build_audio_outputs_sheet() -> Path:
    """Capture eight reconstructed, protected, and fail-silent balanced outputs."""
    folder = ROOT / "AUDIO-8X8"
    name = "Audio-Outputs"
    create_project(folder, name)
    schematic = create_schematic(name)
    schematic.set_paper_size("A0")
    schematic.set_title_block(
        title="ProComm AUDIO-8X8 - Eight Balanced Output Channels",
        date="2026-08-16",
        rev="A1",
        company="ProComm",
        comments={
            1: "AK4458 datasheet PCM LPF: 3.9 k, 4.7 k, 150 R, 470 pF, 3.9 nF",
            2: "OPA1652 gain stage and THAT1646 active-balanced line driver",
            3: "TQ2-12V DPDT relays short both XLR legs to AGND while de-energized",
            4: "+24 dBu into 600 ohm is a qualification target; simulation and bench release required",
        },
    )
    heading(schematic, "EIGHT IDENTICAL OUTPUT CHANNELS - SIGNAL FLOWS LEFT TO RIGHT", (320, 20), 1.8)
    note(schematic, "AK4458 -> PCM reconstruction/difference amp -> x3 gain -> THAT1646 -> fail-silent relay -> ferrite/RFI/fault clamps -> XLR", (360, 34), 0.86)
    note(schematic, "Relays may energize only after all audio PG signals, TDM clocks, resets, register readback, and DAC unmute sequencing pass.", (350, 45), 0.86)

    for channel in range(1, 9):
        y = 100 + (channel - 1) * 86
        heading(schematic, f"CH{channel} OUTPUT", (28, y - 31), 1.25)
        base = 5000 + channel * 40

        r_in_p = add_audio8_passive(schematic, "Device:R", f"R{base}", "3.9k 0.1%", (70, y - 16))
        label_two_pin_device(schematic, r_in_p, f"DAC_CH{channel}P", f"AOUT{channel}_LPF_P")
        r_in_n = add_audio8_passive(schematic, "Device:R", f"R{base + 1}", "3.9k 0.1%", (70, y + 16))
        label_two_pin_device(schematic, r_in_n, f"DAC_CH{channel}N", f"AOUT{channel}_LPF_N")
        r_fb = add_audio8_passive(schematic, "Device:R", f"R{base + 2}", "4.7k 0.1%", (130, y - 16))
        label_two_pin_device(schematic, r_fb, f"AOUT{channel}_LPF_SE", f"AOUT{channel}_LPF_N")
        r_ref = add_audio8_passive(schematic, "Device:R", f"R{base + 3}", "4.7k 0.1%", (130, y + 16))
        label_two_pin_device(schematic, r_ref, f"AOUT{channel}_LPF_P", "AGND")
        c_fb = add_audio8_passive(schematic, "Device:C", f"C{base}", "470pF C0G", (190, y - 16))
        label_two_pin_device(schematic, c_fb, f"AOUT{channel}_LPF_SE", f"AOUT{channel}_LPF_N")
        c_ref = add_audio8_passive(schematic, "Device:C", f"C{base + 1}", "470pF C0G", (190, y + 16))
        label_two_pin_device(schematic, c_ref, f"AOUT{channel}_LPF_P", "AGND")

        opamp = add_symbol(
            schematic,
            "CM5Carrier:OPA1652",
            f"U{500 + channel}",
            "OPA1652AIDR",
            (285, y),
            manufacturer="Texas Instruments",
            mpn="OPA1652AIDR",
        )
        for pin, net_name in {
            1: f"AOUT{channel}_LPF_SE", 2: f"AOUT{channel}_LPF_N", 3: f"AOUT{channel}_LPF_P",
            4: "AUDIO_N15V", 5: f"AOUT{channel}_LPF_POST", 6: f"AOUT{channel}_GAIN_FB",
            7: f"AOUT{channel}_DRIVE", 8: "AUDIO_P15V",
        }.items():
            label_pin_auto(schematic, opamp, pin, net_name, 0.42)
        r_post = add_audio8_passive(schematic, "Device:R", f"R{base + 4}", "150R", (370, y - 16))
        label_two_pin_device(schematic, r_post, f"AOUT{channel}_LPF_SE", f"AOUT{channel}_LPF_POST")
        c_post = add_audio8_passive(schematic, "Device:C", f"C{base + 2}", "3.9nF C0G", (370, y + 16))
        label_two_pin_device(schematic, c_post, f"AOUT{channel}_LPF_POST", "AGND")
        r_gain_fb = add_audio8_passive(schematic, "Device:R", f"R{base + 5}", "21.5k 0.1%", (430, y - 16))
        label_two_pin_device(schematic, r_gain_fb, f"AOUT{channel}_DRIVE", f"AOUT{channel}_GAIN_FB")
        r_gain_ref = add_audio8_passive(schematic, "Device:R", f"R{base + 6}", "10k 0.1%", (430, y + 16))
        label_two_pin_device(schematic, r_gain_ref, f"AOUT{channel}_GAIN_FB", "AGND")

        driver = add_symbol(
            schematic,
            "CM5Carrier:THAT1646",
            f"U{520 + channel}",
            "THAT1646S08-U",
            (535, y),
            manufacturer="THAT Corporation",
            mpn="THAT1646S08-U",
        )
        for pin, net_name in {
            1: f"AOUT{channel}_DRV_N", 2: f"AOUT{channel}_SNS_N", 3: "AGND",
            4: f"AOUT{channel}_DRIVE", 5: "AUDIO_N15V", 6: "AUDIO_P15V",
            7: f"AOUT{channel}_SNS_P", 8: f"AOUT{channel}_DRV_P",
        }.items():
            label_pin_auto(schematic, driver, pin, net_name, 0.44)
        for offset, rail, x in (
            (7, "AUDIO_P15V", 265),
            (8, "AUDIO_N15V", 325),
            (9, "AUDIO_P15V", 505),
            (10, "AUDIO_N15V", 565),
        ):
            bypass = add_audio8_passive(
                schematic, "Device:C", f"C{base + offset}", "100nF", (x, y + 32)
            )
            label_two_pin_device(schematic, bypass, rail, "AGND")
        sense_n = add_audio8_passive(schematic, "Device:C", f"C{base + 3}", "10uF bipolar", (615, y - 16))
        label_two_pin_device(schematic, sense_n, f"AOUT{channel}_DRV_N", f"AOUT{channel}_SNS_N")
        sense_p = add_audio8_passive(schematic, "Device:C", f"C{base + 4}", "10uF bipolar", (615, y + 16))
        label_two_pin_device(schematic, sense_p, f"AOUT{channel}_DRV_P", f"AOUT{channel}_SNS_P")

        relay = add_symbol(
            schematic,
            "CM5Carrier:Panasonic_TQ2_12V",
            f"K{500 + channel}",
            "TQ2-12V",
            (715, y),
            manufacturer="Panasonic Industry",
            mpn="TQ2-12V",
        )
        for pin, net_name in {
            1: "AUDIO_12V", 2: "AGND", 3: f"AOUT{channel}_DRV_N", 4: f"AOUT{channel}_RELAY_N",
            7: f"AOUT{channel}_RELAY_P", 8: f"AOUT{channel}_DRV_P", 9: "AGND", 10: f"AOUT{channel}_COIL_LOW",
        }.items():
            label_pin_auto(schematic, relay, pin, net_name, 0.40)
        mosfet = add_symbol(
            schematic,
            "Transistor_FET:2N7002",
            f"Q{500 + channel}",
            "2N7002K",
            (800, y),
            manufacturer="Diodes Incorporated",
            mpn="2N7002K-7",
        )
        for pin, net_name in {1: "AUDIO_SAFE_UNMUTE_N", 2: "AGND", 3: f"AOUT{channel}_COIL_LOW"}.items():
            label_pin_auto(schematic, mosfet, pin, net_name, 0.42)
        flyback = add_audio8_passive(schematic, "Device:D", f"D{base}", "1N4148W", (855, y))
        label_two_pin_device(schematic, flyback, f"AOUT{channel}_COIL_LOW", "AUDIO_12V")

        bead_n = add_audio8_passive(schematic, "Device:FerriteBead", f"L{base}", "220R ferrite", (915, y - 16))
        label_two_pin_device(schematic, bead_n, f"AOUT{channel}_RELAY_N", f"AOUT_CH{channel}_COLD")
        bead_p = add_audio8_passive(schematic, "Device:FerriteBead", f"L{base + 1}", "220R ferrite", (915, y + 16))
        label_two_pin_device(schematic, bead_p, f"AOUT{channel}_RELAY_P", f"AOUT_CH{channel}_HOT")
        rf_n = add_audio8_passive(schematic, "Device:C", f"C{base + 5}", "100pF C0G", (980, y - 16))
        label_two_pin_device(schematic, rf_n, f"AOUT_CH{channel}_COLD", "XLR_CHASSIS")
        rf_p = add_audio8_passive(schematic, "Device:C", f"C{base + 6}", "100pF C0G", (980, y + 16))
        label_two_pin_device(schematic, rf_p, f"AOUT_CH{channel}_HOT", "XLR_CHASSIS")
        clamp_nets = (
            (f"AOUT_CH{channel}_HOT", "AUDIO_P15V"),
            ("AUDIO_N15V", f"AOUT_CH{channel}_HOT"),
            (f"AOUT_CH{channel}_COLD", "AUDIO_P15V"),
            ("AUDIO_N15V", f"AOUT_CH{channel}_COLD"),
        )
        for index, (net_1, net_2) in enumerate(clamp_nets):
            diode = add_audio8_passive(schematic, "Device:D", f"D{base + 1 + index}", "S1G", (1040 + (index % 2) * 45, y - 16 + (index // 2) * 32))
            label_two_pin_device(schematic, diode, net_1, net_2)

    note(schematic, "AUDIO_SAFE_UNMUTE_N is generated locally from carrier mute command AND ADC/DAC power-good; loss of any condition drops every relay.", (520, 780), 0.78)
    note(schematic, "OUTPUT RELEASE GATES: AK4458/OPA SPICE, relay sample/insertion coupon, +24 dBu THD+N into 600 ohm and 10 kohm, short/open recovery, RF injection, ESD/surge, and power-transfer pop/click tests.", (520, 790), 0.82)
    output = folder / f"{name}.kicad_sch"
    save_generated_schematic(schematic, output)
    return output


def build_audio_power_sheet() -> Path:
    """Capture the AUDIO-8X8 local isolated bipolar and low-noise rails."""
    folder = ROOT / "AUDIO-8X8"
    name = "Audio-Power"
    create_project(folder, name)
    schematic = create_schematic(name)
    schematic.set_paper_size("A1")
    schematic.set_title_block(
        title="ProComm AUDIO-8X8 - Local Power, Sequencing, and Ground Star",
        date="2026-08-16",
        rev="A1",
        company="ProComm",
        comments={
            1: "AUDIO_12V is the only carrier power input; local 3 A fuse and one GND-to-AGND star",
            2: "TRI 20-1223 creates +/-15 V; TPS62913 creates the quiet 5.5 V pre-rail",
            3: "Separate LT3045 5 V/300 mA rails feed ADC and DAC; TPS7A2033 feeds AKM digital I/O",
            4: "Power-good, sequencing, thermal rise, ripple, and conducted-emissions tests gate release",
        },
    )

    heading(schematic, "1. AUDIO_12V ENTRY AND SINGLE GROUND STAR", (185, 22), 1.8)
    add_connector(
        schematic,
        "Connector_Generic:Conn_02x02_Odd_Even",
        "J102",
        "AUDIO_12V_POWER",
        (70, 90),
        {1: "AUDIO_12V_IN", 2: "AUDIO_12V_IN", 3: "GND", 4: "GND"},
        MF_2X2,
        "Molex",
        "43045-0412",
        two_row=True,
    )
    input_fuse = add_audio8_passive(schematic, "Device:Fuse", "F601", "3A fuse", (145, 72))
    label_two_pin_device(schematic, input_fuse, "AUDIO_12V_IN", "AUDIO_12V")
    ground_star = add_audio8_passive(schematic, "Device:R", "R601", "0R power star", (145, 118))
    label_two_pin_device(schematic, ground_star, "GND", "AGND")
    for reference, value, net_1, net_2, x in (
        ("C601", "100nF", "AUDIO_12V", "AGND", 205),
        ("C602", "22uF 25V", "AUDIO_12V", "AGND", 255),
        ("C603", "22uF 25V", "AUDIO_12V", "AGND", 305),
    ):
        capacitor = add_audio8_passive(schematic, "Device:C", reference, value, (x, 90))
        label_two_pin_device(schematic, capacitor, net_1, net_2)
    note(schematic, "R601 is a wide 2512 zero-ohm star: no fan, RF, Ethernet, or chassis current may share the AGND side.", (195, 155), 0.82)
    note(schematic, "AUDIO_12V source is separately fused on CM5-CARRIER; F601 protects this removable board and harness locally.", (195, 170), 0.82)

    heading(schematic, "2. ISOLATED BIPOLAR LINE-STAGE RAILS", (165, 180), 1.8)
    bipolar = add_symbol(
        schematic,
        "CM5Carrier:TRI20_1223",
        "U601",
        "TRI 20-1223",
        (175, 240),
        manufacturer="TRACO Power",
        mpn="TRI 20-1223",
    )
    for pin, net_name in {
        1: "AUDIO_12V", 2: "AGND", 3: "AUDIO_P15V",
        4: "AGND", 5: "AUDIO_N15V",
    }.items():
        label_pin_auto(schematic, bipolar, pin, net_name, 0.58)
    for reference, rail, x in (
        ("C610", "AUDIO_P15V", 285),
        ("C611", "AUDIO_N15V", 345),
    ):
        capacitor = add_audio8_passive(schematic, "Device:C", reference, "4.7uF 50V", (x, 225))
        label_two_pin_device(schematic, capacitor, rail, "AGND")
    for reference, rail, x in (
        ("C612", "AUDIO_P15V", 285),
        ("C613", "AUDIO_N15V", 345),
    ):
        capacitor = add_audio8_passive(schematic, "Device:C", reference, "100nF", (x, 270))
        label_two_pin_device(schematic, capacitor, rail, "AGND")
    note(schematic, "TRI20: 9-18 V input, +/-15 V at 670 mA per rail, 89% typical; place 4.7 uF directly at each output pin.", (195, 310), 0.82)

    heading(schematic, "3. QUIET 5.5 V PRE-RAIL", (555, 22), 1.8)
    buck = add_symbol(
        schematic,
        "CM5Carrier:TPS62913RPU",
        "U602",
        "TPS62913RPU",
        (520, 100),
        manufacturer="Texas Instruments",
        mpn="TPS62913RPUT",
    )
    for pin, net_name in {
        1: "AUDIO_ENABLE", 2: "AUDIO_5V5_SW", 3: "AUDIO_5V5_PRE",
        4: "AGND", 5: "AUDIO_5V5_PG", 6: "AUDIO_12V",
        7: "AGND", 8: "AUDIO_5V5_NRSS",
        9: "AUDIO_5V5_FB", 10: "AUDIO_5V5_SCONF",
    }.items():
        label_pin_auto(schematic, buck, pin, net_name, 0.52)
    inductor = add_audio8_passive(schematic, "Device:L", "L601", "2.2uH", (610, 72))
    label_two_pin_device(schematic, inductor, "AUDIO_5V5_SW", "AUDIO_5V5_PRE")
    rtop = add_audio8_passive(schematic, "Device:R", "R610", "29.4k 0.1%", (610, 110))
    label_two_pin_device(schematic, rtop, "AUDIO_5V5_PRE", "AUDIO_5V5_FB")
    rbottom = add_audio8_passive(schematic, "Device:R", "R611", "4.99k 0.1%", (675, 110))
    label_two_pin_device(schematic, rbottom, "AUDIO_5V5_FB", "AGND")
    nrss = add_audio8_passive(schematic, "Device:C", "C620", "470nF", (610, 150))
    label_two_pin_device(schematic, nrss, "AUDIO_5V5_NRSS", "AGND")
    sconf = add_audio8_passive(schematic, "Device:R", "R612", "7.5k 0.1%", (675, 150))
    label_two_pin_device(schematic, sconf, "AUDIO_5V5_SCONF", "AGND")
    pg_pullup = add_audio8_passive(schematic, "Device:R", "R613", "10k 0.1%", (740, 72))
    label_two_pin_device(schematic, pg_pullup, "AUDIO_5V5_PRE", "AUDIO_5V5_PG")
    for reference, value, rail, x, y in (
        ("C621", "4.7uF 50V", "AUDIO_12V", 450, 175),
        ("C622", "4.7uF 50V", "AUDIO_12V", 500, 175),
        ("C623", "22uF 10V", "AUDIO_5V5_PRE", 610, 200),
        ("C624", "22uF 10V", "AUDIO_5V5_PRE", 670, 200),
    ):
        capacitor = add_audio8_passive(schematic, "Device:C", reference, value, (x, y))
        label_two_pin_device(schematic, capacitor, rail, "AGND")
    note(schematic, "TPS62913: 2.2 MHz low-noise mode; 29.4k/4.99k feedback; 470 nF NR/SS; verify ripple below the AKM rail budget.", (600, 235), 0.80)

    heading(schematic, "4. SEPARATE ADC AND DAC 5 V LT3045 RAILS", (595, 250), 1.8)

    def add_lt3045_branch(prefix: str, reference: str, y: float, passive_base: int) -> None:
        output_rail = f"{prefix}_5V_A"
        pg_net = f"{prefix}_5V_PG"
        ilim_net = f"{prefix}_5V_ILIM"
        pgfb_net = f"{prefix}_5V_PGFB"
        set_net = f"{prefix}_5V_SET"
        regulator = add_symbol(
            schematic,
            "Regulator_Linear:LT3045xMSE",
            reference,
            "LT3045IMSE",
            (535, y),
            manufacturer="Analog Devices",
            mpn="LT3045IMSE#TRPBF",
        )
        for pin, net_name in {
            1: "AUDIO_5V5_PRE", 2: "AUDIO_5V5_PRE", 3: "AUDIO_5V5_PRE",
            4: "AUDIO_5V5_PG", 5: pg_net, 6: ilim_net, 7: pgfb_net,
            8: set_net, 9: "AGND", 10: output_rail, 11: output_rail,
            12: output_rail, 13: "AGND",
        }.items():
            label_pin_auto(schematic, regulator, pin, net_name, 0.45)
        parts = (
            ("Device:R", f"R{passive_base}", "499R 1% / 300mA ILIM", ilim_net, "AGND", 625, y - 42),
            ("Device:R", f"R{passive_base + 1}", "49.9k 0.1%", set_net, "AGND", 690, y - 42),
            ("Device:C", f"C{passive_base}", "470nF", set_net, "AGND", 755, y - 42),
            ("Device:R", f"R{passive_base + 2}", "140k 1%", output_rail, pgfb_net, 625, y + 42),
            ("Device:R", f"R{passive_base + 3}", "10k 0.1%", pgfb_net, "AGND", 690, y + 42),
            ("Device:R", f"R{passive_base + 4}", "100k 1%", output_rail, pg_net, 755, y + 42),
            ("Device:C", f"C{passive_base + 1}", "10uF", "AUDIO_5V5_PRE", "AGND", 455, y - 42),
            ("Device:C", f"C{passive_base + 2}", "10uF", output_rail, "AGND", 455, y + 42),
        )
        for lib_id, part_ref, value, net_1, net_2, x, part_y in parts:
            part = add_audio8_passive(schematic, lib_id, part_ref, value, (x, part_y))
            label_two_pin_device(schematic, part, net_1, net_2)
        note(schematic, f"{output_rail}: 4.99 V nominal, 300 mA limit, PG threshold about 4.5 V", (620, y + 60), 0.70)

    add_lt3045_branch("ADC", "U603", 320, 630)
    add_lt3045_branch("DAC", "U604", 435, 650)

    heading(schematic, "5. AKM DIGITAL RAIL AND 2.5 V INPUT COMMON MODE", (190, 335), 1.8)
    digital_ldo = add_symbol(
        schematic,
        "CM5Carrier:TPS7A20DBV_A1",
        "U605",
        "TPS7A2033PDBVR",
        (175, 395),
        manufacturer="Texas Instruments",
        mpn="TPS7A2033PDBVR",
    )
    for pin, net_name in {1: "AUDIO_5V5_PRE", 2: "AGND", 3: "AUDIO_5V5_PG", 5: "AKM_3V3_D"}.items():
        label_pin_auto(schematic, digital_ldo, pin, net_name, 0.52)
    schematic.no_connects.add(pin_xy(digital_ldo, 4))
    for reference, value, rail, x in (
        ("C670", "1uF", "AUDIO_5V5_PRE", 90),
        ("C671", "1uF", "AKM_3V3_D", 260),
        ("C672", "4.7uF", "AKM_3V3_D", 320),
    ):
        capacitor = add_audio8_passive(schematic, "Device:C", reference, value, (x, 440))
        label_two_pin_device(schematic, capacitor, rail, "AGND")
    vcm_top = add_audio8_passive(schematic, "Device:R", "R670", "10k 0.1%", (90, 485))
    label_two_pin_device(schematic, vcm_top, "ADC_5V_A", "VCM_2V5")
    vcm_bottom = add_audio8_passive(schematic, "Device:R", "R671", "10k 0.1%", (160, 485))
    label_two_pin_device(schematic, vcm_bottom, "VCM_2V5", "AGND")
    vcm_bulk = add_audio8_passive(schematic, "Device:C_Polarized", "C673", "100uF 6.3V polymer", (230, 485))
    label_two_pin_device(schematic, vcm_bulk, "VCM_2V5", "AGND")
    note(schematic, "VCM_2V5 is a quiet divider reference for the eight ADC input op-amps only; no external load is permitted.", (205, 525), 0.76)

    heading(schematic, "6. ERC SOURCES AND SEQUENCING CONTRACT", (505, 515), 1.6)
    flag_nets = (
        "AUDIO_12V_IN", "AUDIO_12V", "GND", "AGND", "AUDIO_P15V",
        "AUDIO_N15V", "AUDIO_5V5_PRE",
        "AKM_3V3_D", "VCM_2V5", "AUDIO_ENABLE",
    )
    for index, net_name in enumerate(flag_nets):
        x = 365 + (index % 5) * 60
        y = 558 + (index // 5) * 20
        flag = add_symbol(schematic, "power:PWR_FLAG", f"#FLG{601 + index}", "PWR_FLAG", (x, y))
        label_pin_auto(schematic, flag, 1, net_name, 0.48)
    note(schematic, "Release order: AUDIO_ENABLE -> 5V5_PG -> LT3045 PGs + AKM 3V3 valid -> clocks -> resets -> carrier unmute command.", (515, 530), 0.72)
    note(schematic, "U106 on Audio-TDM-Clock hardware-gates DAC unmute and every output relay with ADC_5V_PG and DAC_5V_PG.", (505, 542), 0.72)

    output = folder / f"{name}.kicad_sch"
    save_generated_schematic(schematic, output)
    return output


def build_audio() -> Path:
    folder = ROOT / "AUDIO-8X8"
    name = "Audio-8x8"
    create_project(folder, name)

    schematic = create_schematic(name)
    schematic.set_paper_size("A1")
    schematic.set_title_block(
        title="Radxa CM5 ProComm Audio 8x8 - Interface Contract",
        date="2026-08-16",
        rev="A1",
        company="ProComm",
        comments={
            1: "AK5558VN ADC and AK4458VN DAC over buffered differential TDM",
            2: "+4 dBu nominal, +24 dBu maximum, active-balanced XLR stages",
            3: "Left bank male outputs; right bank female inputs",
            4: "Physical AUDIO-8X8 root; nested detailed pages form one connected board netlist",
        },
    )

    heading(schematic, "A. PANEL-SUPPORTED BALANCED XLR BANK", (92, 20))
    for channel in range(1, 9):
        y = 47 + (channel - 1) * 39
        output = add_symbol(
            schematic,
            "Connector_Audio:NC3MAV",
            f"J{200 + channel}",
            "NC3MAV",
            (55, y),
            footprint="Connector_Audio:Jack_XLR_Neutrik_NC3MAV_Vertical",
            manufacturer="Neutrik",
            mpn="NC3MAV",
        )
        output.hidden_properties.update({"Reference", "Value"})
        label_pin_with_stub(schematic, output, 1, "XLR_CHASSIS", ((-7.62, 0),), "right")
        label_pin_with_stub(schematic, output, "G", "XLR_CHASSIS", ((-7.62, 0),), "right")
        label_pin_with_stub(schematic, output, 2, f"AOUT_CH{channel}_HOT", ((7.62, 0),), "left")
        label_pin_with_stub(
            schematic,
            output,
            3,
            f"AOUT_CH{channel}_COLD",
            ((15.24, 0),),
            "left",
        )
        note(schematic, f"J{200 + channel} / NC3MAV", (55, y + 13), 0.78)
        note(schematic, f"CH{channel} OUT", (55, y + 20), 0.85)

        input_connector = add_symbol(
            schematic,
            "Connector_Audio:NC3FAV",
            f"J{300 + channel}",
            "NC3FAV",
            (150, y),
            footprint="Connector_Audio:Jack_XLR_Neutrik_NC3FAV_Vertical",
            manufacturer="Neutrik",
            mpn="NC3FAV",
        )
        input_connector.hidden_properties.update({"Reference", "Value"})
        label_pin_with_stub(schematic, input_connector, 1, "XLR_CHASSIS", ((-7.62, 0),), "right")
        label_pin_with_stub(schematic, input_connector, 2, f"AIN_CH{channel}_HOT", ((7.62, 0),), "left")
        label_pin_with_stub(
            schematic,
            input_connector,
            3,
            f"AIN_CH{channel}_COLD",
            ((-15.24, 0),),
            "right",
        )
        note(schematic, f"J{300 + channel} / NC3FAV", (150, y + 13), 0.78)
        note(schematic, f"CH{channel} IN", (150, y + 20), 0.85)

    heading(schematic, "B. CARRIER INTERFACES", (235, 20))
    tdm_map = {
        1: "AUD_MCLK_P",
        2: "AUD_MCLK_N",
        3: "GND",
        4: "GND",
        5: "AUD_BCLK_P",
        6: "AUD_BCLK_N",
        7: "AUD_FSYNC_P",
        8: "AUD_FSYNC_N",
        9: "GND",
        10: "GND",
        11: "AUD_DAC_SDIN_P",
        12: "AUD_DAC_SDIN_N",
        13: "AUD_ADC_SDOUT_P",
        14: "AUD_ADC_SDOUT_N",
        15: "GND",
        16: "GND",
        17: "AUD_I2C_SCL",
        18: "AUD_I2C_SDA",
        19: "AUD_ADC_RST_N",
        20: "AUD_DAC_RST_N",
        21: "AUD_DAC_MUTE_CMD_N",
        22: "AUD_IRQ_N",
        23: "AUDIO_PRESENT_N",
        24: "AUDIO_ENABLE",
        25: "LOGIC_3V3",
        26: "GND",
        27: "TDM_SPARE_1",
        28: "TDM_SPARE_2",
        29: "GND",
        30: "GND",
    }
    tdm_representation = add_connector(
        schematic,
        "Connector_Generic:Conn_02x15_Odd_Even",
        "J9901",
        "AUDIO_TDM_CONTROL",
        (285, 72),
        tdm_map,
        MILLIGRID_2X15,
        "Molex",
        "87832-6423",
        two_row=True,
    )
    tdm_representation.in_bom = False
    tdm_representation.on_board = False
    note(schematic, "Mates with CM5-CARRIER J201; 100 ohm differential pairs with interleaved grounds.", (235, 120), 0.9)

    power_representation = add_connector(
        schematic,
        "Connector_Generic:Conn_02x02_Odd_Even",
        "J9902",
        "AUDIO_12V_POWER",
        (380, 70),
        {1: "AUDIO_12V", 2: "AUDIO_12V", 3: "GND", 4: "GND"},
        MF_2X2,
        "Molex",
        "43045-0412",
        two_row=True,
    )
    power_representation.in_bom = False
    power_representation.on_board = False
    note(schematic, "Local filtered conversion produces +/-15 V, AKM 5 V analog and audio 3.3 V.", (345, 95), 0.9)

    heading(schematic, "C. INPUT SIGNAL CHAIN", (235, 150))
    note(schematic, "NC3FAV -> chassis/RFI/ESD -> THAT1206 -> level/filter network -> AK5558VN", (235, 165), 1.0)
    note(schematic, "Eight differential channels; no phantom power; +4 dBu nominal / +24 dBu max.", (235, 175), 1.0)
    note(schematic, "Detailed A1 sheet captures TDM256 straps, four VREFH feeds, low-noise rails, and the +24 dBu level map.", (235, 185), 0.9)

    heading(schematic, "D. OUTPUT SIGNAL CHAIN", (235, 215))
    note(schematic, "AK4458VN -> reconstruction/gain filter -> OPA165x -> THAT1646 -> RFI/fault network -> NC3MAV", (235, 230), 1.0)
    note(schematic, "Stable into 600 ohm test load; 10 kohm or higher is the normal operating load.", (235, 240), 1.0)
    note(schematic, "Hardware mute remains asserted through boot, clock loss, reset and power transfer.", (235, 250), 1.0)

    heading(schematic, "E. DETAILED A1 CAPTURE SHEETS", (235, 285))
    items = (
        "01 Audio-TDM-Clock: LVDS TDM, protected I2C/control, clock distribution",
        "02 AK5558-ADC: exact pins, TDM256/32-bit I2S straps, references and bypass",
        "03 AK4458-DAC: exact pins, serial control, references and reconstruction contract",
        "04 Audio-Inputs: 8x THAT1206, OPA1652, RFI/ESD and anti-alias networks",
        "05 Audio-Outputs: 8x OPA1652/THAT1646, fail-silent relays and protection",
        "06 Audio-Power: TRI20 +/-15 V, 5.5 V pre-rail, dual LT3045, 3.3 V and star",
        "Routing hold: AKM exposed-pad coupons and TQ2 insertion/land-pattern coupon",
    )
    for index, item in enumerate(items):
        note(schematic, item, (235, 300 + index * 10), 0.95)

    heading(schematic, "F. XLR SHIELD / CHASSIS BOND", (350, 150))
    shield_resistor = add_symbol(
        schematic,
        "Device:R",
        "R901",
        "1M",
        (390, 180),
        manufacturer="Yageo",
        mpn="RC0603FR-071ML",
    )
    label_pin(schematic, shield_resistor, 1, "XLR_CHASSIS")
    label_pin(schematic, shield_resistor, 2, "AGND")
    shield_capacitor = add_symbol(
        schematic,
        "Device:C",
        "C901",
        "4.7nF Y1 500Vac",
        (440, 180),
        manufacturer="Vishay",
        mpn="VY1472M63Y5UQ6TV0",
    )
    label_pin(schematic, shield_capacitor, 1, "XLR_CHASSIS")
    label_pin(schematic, shield_capacitor, 2, "AGND")
    add_connector(
        schematic,
        "Connector_Generic:Conn_01x02",
        "J901",
        "CHASSIS_BOND_TEST",
        (500, 180),
        {1: "XLR_CHASSIS", 2: "AGND"},
        "TestPoint:TestPoint_2Pads_Pitch2.54mm_Drill0.8mm",
    )
    note(schematic, "Every XLR pin 1 and NC3MAV shell G bonds directly to the local chassis plane.", (350, 205), 0.88)
    note(schematic, "R901 and Y1-rated C901 form the single controlled RF/static bond to AGND; verify in EMC test.", (350, 214), 0.88)
    note(schematic, "No audio return current may use the XLR_CHASSIS copper or panel fasteners.", (350, 223), 0.88)

    heading(schematic, "G. CONNECTED AUDIO-8X8 HIERARCHY", (650, 20), 1.65)
    add_board_child_sheets(
        schematic,
        name,
        (
            ("TDM / Clock / Control", "Audio-TDM-Clock.kicad_sch"),
            ("AK5558VN ADC", "AK5558-ADC.kicad_sch"),
            ("AK4458VN DAC", "AK4458-DAC.kicad_sch"),
            ("Balanced Inputs 1-8", "Audio-Inputs.kicad_sch"),
            ("Balanced Outputs 1-8", "Audio-Outputs.kicad_sch"),
            ("Audio Power", "Audio-Power.kicad_sch"),
        ),
        (645, 45),
    )
    note(schematic, "Root J9901/J9902 are interface-only duplicates; detailed J101/J102 own the production footprints.", (625, 345), 0.82)

    output = folder / f"{name}.kicad_sch"
    save_generated_schematic(schematic, output)
    return output


def main() -> None:
    pinout = write_cm5_local_library()
    write_cm5_local_footprints()
    symbol_cache = get_symbol_cache()
    symbol_cache.add_library_path(CM5_LOCAL_LIBRARY)
    symbol_cache.add_library_path(CM5_WURTH_LIBRARY)
    paths = (
        build_carrier(),
        build_cm5_core_sheet(pinout),
        build_network_pcie_sheet(),
        build_wwan_sim_sheet(),
        build_display_harness_sheet(),
        build_audio_control_sheet(),
        build_power_regulators_sheet(),
        build_thermal_io_sheet(),
        build_audio(),
        build_audio_tdm_clock_sheet(),
        build_ak5558_adc_sheet(),
        build_ak4458_dac_sheet(),
        build_audio_inputs_sheet(),
        build_audio_outputs_sheet(),
        build_audio_power_sheet(),
    )
    for path in paths:
        print(path)


if __name__ == "__main__":
    main()
