#!/usr/bin/env python3
"""Cross-check every controlled CM5 contact against source data and KiCad."""

from __future__ import annotations

import ast
import os
from pathlib import Path
import subprocess
import tempfile
import xml.etree.ElementTree as ET

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
WORKSPACE = ROOT.parent.parent.parent
GENERATOR = ROOT.parent / "generate_interface_schematics.py"
SCHEMATIC = ROOT / "CM5-Core-Allocated.kicad_sch"
ALLOCATION_WORKBOOK = (
    WORKSPACE
    / "outputs"
    / "cm5-pin-allocation-a0"
    / "radxa_cm5_pin_allocation_a0.xlsx"
)
VENDOR_WORKBOOK = WORKSPACE / "docs" / "radxa_cm5_v2210_pinout.xlsx"
KICAD_CLI = Path(
    os.environ.get("KICAD_CLI", "/Applications/KiCad/KiCad.app/Contents/MacOS/kicad-cli")
)

CONNECTOR_REFS = {"U13-A": "J501", "U13-B": "J502", "J1": "J503"}
VENDOR_COLUMNS = {
    "U13-A": (4, 5, 6),
    "U13-B": (2, 3, 4),
    "J1": (1, 2, 3),
}
GENERIC_MUX_NAMES = {"-", "Power", "Native PHY MDI", "Native PHY LED"}


def check(name: str, condition: bool, detail: str) -> bool:
    print(f"{'PASS' if condition else 'FAIL'}  {name}: {detail}")
    return condition


def generator_constants() -> tuple[tuple[tuple[str, int, str, str], ...], set[tuple[str, int]]]:
    tree = ast.parse(GENERATOR.read_text(), filename=str(GENERATOR))
    values: dict[str, object] = {}
    for node in tree.body:
        if not isinstance(node, (ast.Assign, ast.AnnAssign)):
            continue
        targets = node.targets if isinstance(node, ast.Assign) else [node.target]
        value = node.value
        for target in targets:
            if isinstance(target, ast.Name) and target.id in {
                "CM5_ALLOCATIONS",
                "CM5_ASSIGNED_NC",
            }:
                values[target.id] = ast.literal_eval(value)
    return tuple(values["CM5_ALLOCATIONS"]), set(values["CM5_ASSIGNED_NC"])


def vendor_rows() -> dict[tuple[str, int], tuple[str, str, tuple[str, ...], str]]:
    workbook = load_workbook(VENDOR_WORKBOOK, read_only=True, data_only=True)
    rows: dict[tuple[str, int], tuple[str, str, tuple[str, ...], str]] = {}
    for connector, (signal_column, ball_column, function_column) in VENDOR_COLUMNS.items():
        sheet = workbook[connector]
        for row in sheet.iter_rows(min_row=3, values_only=True):
            if row[0] is None:
                continue
            functions = tuple(
                str(value) for value in row[function_column:-1] if value not in (None, "")
            )
            rows[(connector, int(row[0]))] = (
                str(row[signal_column]),
                str(row[ball_column]),
                functions,
                "" if row[-1] is None else str(row[-1]),
            )
    return rows


def allocation_rows() -> tuple[list[dict[str, object]], dict[tuple[str, int], tuple[object, ...]]]:
    workbook = load_workbook(ALLOCATION_WORKBOOK, read_only=False, data_only=False)
    sheet = workbook["Assignments"]
    headings = [cell.value for cell in sheet[4]]
    assignments = [
        dict(zip(headings, row, strict=True))
        for row in sheet.iter_rows(min_row=5, max_row=80, values_only=True)
        if row[0] is not None
    ]

    source_sheet = workbook["Source Rows"]
    source_rows: dict[tuple[str, int], tuple[object, ...]] = {}
    for row in source_sheet.iter_rows(min_row=5, max_row=80, values_only=True):
        if row[0] is None:
            continue
        source_rows[(str(row[0]), int(row[1]))] = tuple(row[2:15])
    return assignments, source_rows


def net_map() -> dict[tuple[str, str], str]:
    with tempfile.TemporaryDirectory(prefix="radxa-cm5-pin-validation-") as temp:
        output = Path(temp) / "CM5-Core-Allocated.xml"
        result = subprocess.run(
            [
                str(KICAD_CLI),
                "sch",
                "export",
                "netlist",
                "--format",
                "kicadxml",
                "--output",
                str(output),
                str(SCHEMATIC),
            ],
            check=False,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
        )
        if result.returncode:
            raise RuntimeError(f"KiCad netlist export failed:\n{result.stderr.strip()}")
        root = ET.parse(output)
        return {
            (node.attrib["ref"], node.attrib["pin"]): net.attrib["name"].lstrip("/")
            for net in root.findall("./nets/net")
            for node in net.findall("node")
        }


def main() -> int:
    generator_allocations, assigned_nc = generator_constants()
    assignments, source_rows = allocation_rows()
    vendor = vendor_rows()
    nets = net_map()
    checks: list[bool] = []

    generator_by_pin = {
        (connector, pin): (direction, net)
        for connector, pin, direction, net in generator_allocations
    }
    workbook_by_pin = {
        (str(row["Connector"]), int(row["Pin"])): (
            str(row["Direction at CM5"]),
            str(row["Carrier net"]),
        )
        for row in assignments
    }
    checks.append(
        check(
            "controlled allocation cardinality",
            len(generator_allocations) == len(generator_by_pin) == len(assignments) == len(workbook_by_pin) == 76,
            "76 unique physical contacts in both generator and allocation workbook",
        )
    )
    checks.append(
        check(
            "generator/workbook electrical identity",
            generator_by_pin == workbook_by_pin,
            "connector, pin, direction, and carrier net agree for every controlled contact",
        )
    )

    status_counts = {
        status: sum(row["Status"] == status for row in assignments)
        for status in ("LOCKED", "CONDITIONAL", "RESERVED")
    }
    checks.append(
        check(
            "allocation status accounting",
            status_counts == {"LOCKED": 71, "CONDITIONAL": 2, "RESERVED": 3},
            f"status counts {status_counts}",
        )
    )

    workbook_nc = {
        (str(row["Connector"]), int(row["Pin"]))
        for row in assignments
        if str(row["Destination"]).startswith("Assigned no-connect")
    }
    checks.append(
        check(
            "assigned no-connect ownership",
            assigned_nc == workbook_nc == {("U13-A", 15), ("U13-B", 147)},
            f"assigned NC contacts {sorted(workbook_nc)}",
        )
    )

    vendor_identity_ok = True
    source_extract_ok = len(source_rows) == 76
    mux_ok = True
    for row in assignments:
        key = (str(row["Connector"]), int(row["Pin"]))
        vendor_signal, vendor_ball, functions, vendor_note = vendor[key]
        vendor_identity_ok &= (
            str(row["CM5 signal"]) == vendor_signal
            and str(row["SoC ball"]) == vendor_ball
        )
        mux = str(row["Selected mux"])
        mux_ok &= (
            mux in GENERIC_MUX_NAMES
            or mux == vendor_signal
            or mux in functions
        )
        expected_source = (
            vendor_signal,
            vendor_ball,
            *functions,
            *(None for _ in range(10 - len(functions))),
            vendor_note or None,
        )
        source_extract_ok &= source_rows.get(key) == expected_source

    checks.append(
        check(
            "Radxa V2.21 physical identity",
            vendor_identity_ok,
            "all connector signals and SoC balls match the controlled vendor workbook",
        )
    )
    checks.append(
        check(
            "Radxa V2.21 selected mux legality",
            mux_ok,
            "every selected mux is listed on the exact vendor pin or is a physical power/PHY role",
        )
    )
    checks.append(
        check(
            "controlled source-row extract",
            source_extract_ok,
            "all 76 normalized source rows remain byte-for-cell equivalent to the vendor workbook",
        )
    )

    connected_ok = True
    nc_ok = True
    for connector, pin, _direction, expected_net in generator_allocations:
        actual_net = nets.get((CONNECTOR_REFS[connector], str(pin)), "")
        if (connector, pin) in assigned_nc:
            nc_ok &= actual_net.startswith("unconnected-")
        else:
            connected_ok &= actual_net == expected_net
    checks.append(
        check(
            "KiCad connected allocation",
            connected_ok,
            "all 74 routed CM5 contacts carry the exact controlled net",
        )
    )
    checks.append(
        check(
            "KiCad deliberate no-connects",
            nc_ok,
            "WAN1 LED2 and HDMI SBDN are explicit no-connects, not dangling labels",
        )
    )

    failures = len(checks) - sum(checks)
    print(f"\nRESULT: {len(checks)} checks, {failures} failures")
    return 1 if failures else 0


if __name__ == "__main__":
    raise SystemExit(main())
