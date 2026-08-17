#!/usr/bin/env python3
"""Prove the carrier CM5 mating geometry matches Radxa's official PCB."""

from __future__ import annotations

import hashlib
from pathlib import Path
import re
import subprocess
import tempfile
import xml.etree.ElementTree as ET

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
WORKSPACE = ROOT.parent.parent.parent
SOURCE_DIR = (
    WORKSPACE
    / "references"
    / "radxa-cm-projects"
    / "cm5"
    / "radxa-cm5-io-board"
)
SOURCE_ASC = SOURCE_DIR / "radxa_cm5_io_board_v2200.asc"
SOURCE_BOM = SOURCE_DIR / "radxa_cm5_io_board_v2200_BOM.xlsx"
HIROSE_DRAWING = (
    WORKSPACE
    / "references"
    / "components"
    / "hirose"
    / "DF40C-100DS-0.4V"
    / "DF40C-100DS-0.4V-51_2D_drawing.pdf"
)
HIROSE_DRAWING_SHA256 = "30b4c54ab38ddd18a65e1f3b631b8bcfb64efd1799c2a4a10936f010070521cd"
SCHEMATIC = ROOT / "CM5-Core-Allocated.kicad_sch"
FOOTPRINT_DIR = ROOT / "CM5Carrier.pretty"
KICAD_CLI = Path(
    "/Applications/KiCad/KiCad.app/Contents/MacOS/kicad-cli"
)

FOOTPRINTS = {
    "J501": "Radxa_CM5_U33A_DF40C_100DS_OFFICIAL",
    "J502": "Radxa_CM5_U33B_DF40C_100DS_OFFICIAL",
    "J503": "Radxa_CM5_J24_DF40C_100DS_OFFICIAL",
}


def check(name: str, condition: bool, detail: str) -> bool:
    print(f"{'PASS' if condition else 'FAIL'}  {name}: {detail}")
    return condition


def balanced_sexpr(text: str, start: int) -> str:
    depth = 0
    quoted = False
    escaped = False
    for index in range(start, len(text)):
        character = text[index]
        if quoted:
            if escaped:
                escaped = False
            elif character == "\\":
                escaped = True
            elif character == '"':
                quoted = False
            continue
        if character == '"':
            quoted = True
        elif character == "(":
            depth += 1
        elif character == ")":
            depth -= 1
            if depth == 0:
                return text[start:index + 1]
    raise RuntimeError("Unbalanced KiCad S-expression")


def top_level_sexprs(expression: str) -> list[str]:
    children: list[str] = []
    depth = 0
    quoted = False
    escaped = False
    child_start: int | None = None
    for index, character in enumerate(expression):
        if quoted:
            if escaped:
                escaped = False
            elif character == "\\":
                escaped = True
            elif character == '"':
                quoted = False
            continue
        if character == '"':
            quoted = True
        elif character == "(":
            depth += 1
            if depth == 2:
                child_start = index
        elif character == ")":
            if depth == 2 and child_start is not None:
                children.append(expression[child_start:index + 1])
                child_start = None
            depth -= 1
    return children


def expression_name(expression: str) -> str:
    match = re.match(r"\(\s*([^\s()]+)", expression)
    if not match:
        raise RuntimeError(f"Cannot identify expression: {expression[:80]!r}")
    return match.group(1)


def footprint_by_reference(board_text: str, reference: str) -> str:
    marker = f'(property "Reference" "{reference}"'
    property_offset = board_text.find(marker)
    if property_offset < 0:
        raise RuntimeError(f"Imported official board has no {reference}")
    start = board_text.rfind("(footprint ", 0, property_offset)
    return balanced_sexpr(board_text, start)


def restore_native_j24_pad_angle(expression: str) -> str:
    """Restore the 0-degree angle from Radxa's native PADS decal."""
    at_match = re.search(
        r"\(at\s+(-?[0-9.]+)\s+(-?[0-9.]+)\s+(-?[0-9.]+)\)",
        expression,
    )
    if not at_match or float(at_match.group(3)) % 360.0 != 270.0:
        raise RuntimeError("Unexpected J24 pad angle in KiCad PADS import")
    replacement = f"(at {at_match.group(1)} {at_match.group(2)})"
    return expression[:at_match.start()] + replacement + expression[at_match.end():]


def pad_map(footprint: str, *, restore_j24_angle: bool = False) -> dict[int, str]:
    pads: dict[int, str] = {}
    for child in top_level_sexprs(footprint):
        if expression_name(child) != "pad":
            continue
        match = re.match(r'\(pad "(\d+)"', child)
        if match:
            if restore_j24_angle:
                child = restore_native_j24_pad_angle(child)
            pads[int(match.group(1))] = normalized_geometry(child)
    return pads


def pad_dimensions(footprint: str) -> dict[int, tuple[float, float, float, float, float]]:
    """Return pad x, y, local angle, width and height."""
    pads: dict[int, tuple[float, float, float, float, float]] = {}
    for child in top_level_sexprs(footprint):
        if expression_name(child) != "pad":
            continue
        number = re.match(r'\(pad "(\d+)"', child)
        at = re.search(
            r"\(at\s+(-?[0-9.]+)\s+(-?[0-9.]+)(?:\s+(-?[0-9.]+))?\)",
            child,
        )
        size = re.search(r"\(size\s+([0-9.]+)\s+([0-9.]+)\)", child)
        if not number or not at or not size:
            raise RuntimeError("Cannot parse generated CM5 pad geometry")
        pads[int(number.group(1))] = (
            float(at.group(1)),
            float(at.group(2)),
            float(at.group(3) or 0.0) % 360.0,
            float(size.group(1)),
            float(size.group(2)),
        )
    return pads


def native_pads_decal_contract() -> bool:
    """Check pitch and land geometry directly in Radxa's PADS source."""
    source = SOURCE_ASC.read_text(errors="replace")
    start = re.search(r"^DF40C_100DS\s+M\s", source, re.MULTILINE)
    if not start:
        return False
    next_decal = re.search(
        r"^[A-Za-z0-9_+.-]+\s+M\s",
        source[start.end():],
        re.MULTILINE,
    )
    if not next_decal:
        return False
    decal = source[start.start():start.end() + next_decal.start()]
    coordinates = [
        (int(x), int(y), int(pin))
        for x, y, pin in re.findall(
            r"^T(-?\d+)\s+(-?\d+)\s+-?\d+\s+-?\d+\s+(\d+)\s*$",
            decal,
            re.MULTILINE,
        )
    ]
    if [pin for _, _, pin in coordinates] != list(range(1, 101)):
        return False
    pitch_units = {
        abs(coordinates[index + 2][1] - coordinates[index][1])
        for index in range(98)
    }
    row_spacing_units = {
        abs(coordinates[index + 1][0] - coordinates[index][0])
        for index in range(0, 100, 2)
    }
    pad_stack_ok = bool(
        re.search(
            r"^PAD 0 3\s*\n-2 300000 RF\s+0\.000 1710000\b",
            decal,
            re.MULTILINE,
        )
    )
    return pitch_units == {600000} and row_spacing_units == {3960000} and pad_stack_ok


def normalized_geometry(expression: str) -> str:
    expression = re.sub(
        r"\n\s*\((?:net|uuid|path)\s+[^\n()]*\)",
        "",
        expression,
    )
    return " ".join(expression.split())


def footprint_transform(footprint: str) -> tuple[float, float, float]:
    for child in top_level_sexprs(footprint):
        if expression_name(child) != "at":
            continue
        values = [float(value) for value in child[3:-1].split()]
        return values[0], values[1], values[2] if len(values) > 2 else 0.0
    raise RuntimeError("Imported footprint has no board position")


def generated_footprint(name: str) -> str:
    path = FOOTPRINT_DIR / f"{name}.kicad_mod"
    text = path.read_text()
    start = text.find("(footprint ")
    return balanced_sexpr(text, start)


def netlist_contract() -> tuple[dict[str, str], dict[tuple[str, str], str]]:
    with tempfile.TemporaryDirectory(prefix="cm5-mating-netlist-") as temp:
        netlist = Path(temp) / "cm5-core.xml"
        result = subprocess.run(
            [
                str(KICAD_CLI),
                "sch",
                "export",
                "netlist",
                "--format",
                "kicadxml",
                "--output",
                str(netlist),
                str(SCHEMATIC),
            ],
            check=False,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
        )
        if result.returncode:
            raise RuntimeError(f"CM5 netlist export failed:\n{result.stderr.strip()}")
        root = ET.parse(netlist)
    footprints = {
        component.attrib.get("ref", ""): component.findtext("footprint", "")
        for component in root.findall("./components/comp")
    }
    nets = {
        (node.attrib["ref"], node.attrib["pin"]): net.attrib["name"].lstrip("/")
        for net in root.findall("./nets/net")
        for node in net.findall("node")
    }
    return footprints, nets


def official_board_text() -> str:
    with tempfile.TemporaryDirectory(prefix="radxa-cm5-official-pcb-") as temp:
        output = Path(temp) / "radxa_cm5_io_board_v2200.kicad_pcb"
        result = subprocess.run(
            [
                str(KICAD_CLI),
                "pcb",
                "import",
                "--format",
                "pads",
                "--output",
                str(output),
                str(SOURCE_ASC),
            ],
            check=False,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
        )
        if result.returncode:
            raise RuntimeError(f"Official Radxa board import failed:\n{result.stderr.strip()}")
        return output.read_text()


def main() -> int:
    board_text = official_board_text()
    official_u33 = footprint_by_reference(board_text, "U33")
    official_j24 = footprint_by_reference(board_text, "J24")
    official_u33_pads = pad_map(official_u33)
    official_j24_pads = pad_map(official_j24, restore_j24_angle=True)
    generated = {
        reference: pad_map(generated_footprint(name))
        for reference, name in FOOTPRINTS.items()
    }
    checks: list[bool] = []

    checks.append(
        check(
            "native Radxa PADS decal",
            native_pads_decal_contract(),
            "100 pins; 0.40 mm pitch; 2.64 mm row spacing; 1.14 x 0.20 mm lands at 0 degrees",
        )
    )
    drawing_ok = (
        HIROSE_DRAWING.exists()
        and hashlib.sha256(HIROSE_DRAWING.read_bytes()).hexdigest() == HIROSE_DRAWING_SHA256
    )
    checks.append(
        check(
            "Hirose controlled drawing",
            drawing_ok,
            "manufacturer drawing is present and SHA-256 locked",
        )
    )

    expected_pins = {
        "J501": set(range(1, 101)) | set(range(201, 205)),
        "J502": set(range(101, 201)),
        "J503": set(range(1, 101)),
    }
    checks.append(
        check(
            "generated pad cardinality",
            all(set(generated[ref]) == pins for ref, pins in expected_pins.items()),
            "J501=104, J502=100 and J503=100 exact numbered pads",
        )
    )
    source_maps = {"J501": official_u33_pads, "J502": official_u33_pads, "J503": official_j24_pads}
    geometry_ok = all(
        generated[reference][pin] == source_maps[reference][pin]
        for reference, pins in expected_pins.items()
        for pin in pins
    )
    checks.append(
        check(
            "official pad/drill identity",
            geometry_ok,
            "all 304 generated pad records match Radxa V2.20 after the proved J24 importer-angle correction",
        )
    )

    j24_dimensions = pad_dimensions(generated_footprint(FOOTPRINTS["J503"]))
    j24_pattern_ok = all(
        abs(x - (2.73 if pin % 2 else 0.09)) < 1e-9
        and abs(y - (1.5 + ((pin - 1) // 2) * 0.4)) < 1e-9
        and angle == 0.0
        and abs(width - 1.14) < 1e-9
        and abs(height - 0.2) < 1e-9
        for pin, (x, y, angle, width, height) in j24_dimensions.items()
    )
    checks.append(
        check(
            "J24 physical land pattern",
            len(j24_dimensions) == 100 and j24_pattern_ok,
            "zero-overlap 0.40 mm-pitch rows with 1.50 mm inter-row copper gap",
        )
    )

    u33_x, u33_y, u33_angle = footprint_transform(official_u33)
    j24_x, j24_y, j24_angle = footprint_transform(official_j24)
    relative = (
        round(j24_x - u33_x, 6),
        round(j24_y - u33_y, 6),
        round(j24_angle - u33_angle, 6),
    )
    checks.append(
        check(
            "official J24 transform",
            relative == (11.405, -25.415, -90.0),
            f"J24 relative to U33 is {relative} mm/degrees",
        )
    )

    # This workbook has trailing records that openpyxl's streaming reader
    # omits, including the controlled DF40 row, so load the full worksheet.
    workbook = load_workbook(SOURCE_BOM, read_only=False, data_only=True)
    bom_rows = [
        row
        for row in workbook.worksheets[0].iter_rows(values_only=True)
        if any("DF40C-100DS-0.4V(51)" in str(value) for value in row if value is not None)
    ]
    bom_ok = len(bom_rows) == 1 and int(bom_rows[0][5]) == 3
    checks.append(
        check(
            "official connector BOM",
            bom_ok,
            "Radxa BOM specifies three Hirose DF40C-100DS-0.4V(51) receptacles",
        )
    )

    footprints, nets = netlist_contract()
    footprint_ok = all(
        footprints.get(reference) == f"CM5Carrier:{name}"
        for reference, name in FOOTPRINTS.items()
    )
    checks.append(
        check(
            "schematic footprint ownership",
            footprint_ok,
            "J501/J502/J503 own the three official-source local footprints",
        )
    )
    mount_ground_ok = all(nets.get(("J501", str(pin))) == "GND" for pin in range(201, 205))
    checks.append(
        check(
            "module mounting-pad ground",
            mount_ground_ok,
            "U33 mechanical pads 201-204 are explicitly tied to carrier GND",
        )
    )

    failures = len(checks) - sum(checks)
    print(f"\nRESULT: {len(checks)} checks, {failures} failures")
    return 1 if failures else 0


if __name__ == "__main__":
    raise SystemExit(main())
